"""
Advanced CFDI Reports
Generate Excel reports for CFDIs with filtering and analysis
"""

from typing import Any, Dict, List, Optional
from datetime import datetime
import logging
from pathlib import Path

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not available - Excel export disabled")

class CFDIReportsService:
    """Service for generating CFDI reports."""
    
    def __init__(self, core):
        self.core = core
    
    def generate_period_report(
        self,
        start_date: str,
        end_date: str,
        output_path: Optional[str] = None,
        include_cancelled: bool = False
    ) -> str:
        """
        Generate Excel report for period.
        
Args:
            start_date: Start date (YYYY-MM-DD)
            end_date: End date (YYYY-MM-DD)
            output_path: Optional output path
            include_cancelled: Include cancelled CFDIs
            
        Returns:
            Path to generated Excel file
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl required for Excel export")
        
        # Get CFDIs
        sql = """
            SELECT * FROM cfdis
            WHERE DATE(fecha_emision) BETWEEN %s AND %s
        """
        params = [start_date, end_date]
        
        if not include_cancelled:
            sql += " AND estado = 'vigente'"
        
        sql += " ORDER BY fecha_emision DESC"
        
        result = self.core.db.execute_query(sql, tuple(params))
        cfdis = [dict(row) for row in result] if result else []
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CFDIs"
        
        # Styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Headers
        headers = [
            'UUID', 'Folio', 'Serie', 'Fecha Emisión', 'RFC Receptor',
            'Nombre Receptor', 'Subtotal', 'Impuestos', 'Total', 'Estado'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data
        for row_idx, cfdi in enumerate(cfdis, 2):
            ws.cell(row=row_idx, column=1, value=cfdi.get('uuid', ''))
            ws.cell(row=row_idx, column=2, value=cfdi.get('folio', ''))
            ws.cell(row=row_idx, column=3, value=cfdi.get('serie', ''))
            ws.cell(row=row_idx, column=4, value=cfdi.get('fecha_emision', ''))
            ws.cell(row=row_idx, column=5, value=cfdi.get('rfc_receptor', ''))
            ws.cell(row=row_idx, column=6, value=cfdi.get('nombre_receptor', ''))
            ws.cell(row=row_idx, column=7, value=cfdi.get('subtotal', 0))
            ws.cell(row=row_idx, column=8, value=cfdi.get('impuestos', 0))
            ws.cell(row=row_idx, column=9, value=cfdi.get('total', 0))
            ws.cell(row=row_idx, column=10, value=cfdi.get('estado', ''))
            
            # Format currency columns
            for col in [7, 8, 9]:
                ws.cell(row=row_idx, column=col).number_format = '"$"#,##0.00'
        
        # Add summary
        summary_row = len(cfdis) + 3
        ws.cell(row=summary_row, column=6, value="TOTALES:")
        ws.cell(row=summary_row, column=6).font = Font(bold=True)
        
        ws.cell(row=summary_row, column=7, value=f'=SUM(G2:G{len(cfdis)+1})')
        ws.cell(row=summary_row, column=8, value=f'=SUM(H2:H{len(cfdis)+1})')
        ws.cell(row=summary_row, column=9, value=f'=SUM(I2:I{len(cfdis)+1})')
        
        for col in [7, 8, 9]:
            ws.cell(row=summary_row, column=col).font = Font(bold=True)
            ws.cell(row=summary_row, column=col).number_format = '"$"#,##0.00'
        
        # Autosize columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save
        if not output_path:
            from app.core import DATA_DIR
            reports_dir = Path(DATA_DIR) / "reports"
            reports_dir.mkdir(exist_ok=True)
            output_path = reports_dir / f"CFDIs_{start_date}_to_{end_date}.xlsx"
        
        wb.save(output_path)
        logger.info(f"Report generated: {output_path}")
        
        return str(output_path)
    
    def generate_customer_report(self, customer_rfc: str) -> str:
        """Generate report for specific customer."""
        sql = "SELECT * FROM cfdis WHERE rfc_receptor = %s ORDER BY fecha_emision DESC"
        result = self.core.db.execute_query(sql, (customer_rfc,))
        
        cfdis = [dict(row) for row in result] if result else []
        
        if not cfdis:
            raise ValueError(f"No CFDIs found for RFC: {customer_rfc}")
        
        # Similar Excel generation as period_report
        from app.core import DATA_DIR
        reports_dir = Path(DATA_DIR) / "reports"
        reports_dir.mkdir(exist_ok=True)
        output_path = reports_dir / f"CFDIs_{customer_rfc}.xlsx"
        
        # Create workbook (simplified version)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"CFDIs - {customer_rfc}"
        
        # Add data (same as period report)
        # ... (code similar to generate_period_report)
        
        wb.save(output_path)
        return str(output_path)
    
    def get_statistics(self, start_date: str, end_date: str) -> Dict[str, Any]:
        """Get CFDI statistics for period."""
        sql = """
            SELECT 
                COUNT(*) as total_cfdis,
                COALESCE(SUM(CASE WHEN estado = 'vigente' THEN 1 ELSE 0 END), 0) as vigentes,
                COALESCE(SUM(CASE WHEN estado = 'cancelado' THEN 1 ELSE 0 END), 0) as cancelados,
                COALESCE(SUM(CASE WHEN estado = 'vigente' THEN total ELSE 0 END), 0) as total_facturado,
                COALESCE(SUM(CASE WHEN estado = 'vigente' THEN impuestos ELSE 0 END), 0) as total_impuestos,
                COUNT(DISTINCT rfc_receptor) as clientes_unicos
            FROM cfdis
            WHERE DATE(fecha_emision) BETWEEN %s AND %s
        """
        
        result = self.core.db.execute_query(sql, (start_date, end_date))
        
        if result:
            row = dict(result[0])
            return {
                'total_cfdis': row.get('total_cfdis', 0),
                'vigentes': row.get('vigentes', 0),
                'cancelados': row.get('cancelados', 0),
                'total_facturado': row.get('total_facturado', 0),
                'total_impuestos': row.get('total_impuestos', 0),
                'clientes_unicos': row.get('clientes_unicos', 0),
                'promedio_por_cfdi': row.get('total_facturado', 0) / max(row.get('vigentes', 1), 1)
            }
        
        return {}
