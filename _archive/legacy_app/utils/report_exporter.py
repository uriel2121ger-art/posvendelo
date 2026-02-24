"""
Report export utilities for Excel and PDF formats.
Provides functions to export POS reports to various formats.
"""
from __future__ import annotations

from typing import Any
from datetime import datetime
import io
from pathlib import Path

from PyQt6 import QtWidgets


def export_to_excel(data: list[dict[str, Any]], filename: str, sheet_name: str = "Report") -> bool:
    """Export data to Excel format using openpyxl.
    
    Args:
        data: List of dictionaries with report data
        filename: Output filename
        sheet_name: Excel sheet name
        
    Returns:
        True if successful, False otherwise
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        if not data:
            wb.save(filename)
            return True
        
        # Write headers
        headers = list(data[0].keys())
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header.upper())
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Write data
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, header in enumerate(headers, start=1):
                value = row_data.get(header, "")
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass  # Error obteniendo longitud de celda
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save
        wb.save(filename)
        return True
        
    except ImportError:
        QtWidgets.QMessageBox.warning(
            None,
            "Librería no disponible",
            "La librería 'openpyxl' no está instalada.\nInstálala con: pip install openpyxl"
        )
        return False
    except Exception as e:
        QtWidgets.QMessageBox.critical(
            None,
            "Error",
            f"Error al exportar a Excel:\n{str(e)}"
        )
        return False

def export_to_pdf(data: list[dict[str, Any]], filename: str, title: str = "Reporte") -> bool:
    """Export data to PDF format using reportlab.
    
    Args:
        data: List of dictionaries with report data
        filename: Output filename
        title: Report title
        
    Returns:
        True if successful, False otherwise
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, letter
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        # Create PDF
        doc = SimpleDocTemplate(filename, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_para = Paragraph(f"<b>{title}</b>", styles['Title'])
        elements.append(title_para)
        elements.append(Spacer(1, 0.3*inch))
        
        # Date
        date_para = Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal'])
        elements.append(date_para)
        elements.append(Spacer(1, 0.2*inch))
        
        if not data:
            no_data_para = Paragraph("No hay datos para mostrar", styles['Normal'])
            elements.append(no_data_para)
            doc.build(elements)
            return True
        
        # Create table
        headers = list(data[0].keys())
        table_data = [[h.upper() for h in headers]]
        
        for row in data:
            table_data.append([str(row.get(h, "")) for h in headers])
        
        # Calculate column widths
        col_widths = [1.5*inch] * len(headers)
        
        table = Table(table_data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')])
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        return True
        
    except ImportError as e:
        QtWidgets.QMessageBox.warning(
            None,
            "Librería no disponible",
            f"La librería 'reportlab' no está instalada.\nInstálala con: pip install reportlab"
        )
        return False
    except Exception as e:
        QtWidgets.QMessageBox.critical(
            None,
            "Error",
            f"Error al exportar a PDF:\n{str(e)}"
        )
        return False

def prompt_export_format(parent=None) -> str | None:
    """Show dialog to select export format.
    
    Returns:
        'excel', 'pdf', or None if canceled
    """
    dialog = QtWidgets.QDialog(parent)
    dialog.setWindowTitle("Exportar Reporte")
    dialog.setModal(True)
    
    layout = QtWidgets.QVBoxLayout(dialog)
    
    layout.addWidget(QtWidgets.QLabel("Selecciona el formato de exportación:"))
    
    excel_btn = QtWidgets.QPushButton("📊 Excel (.xlsx)")
    excel_btn.clicked.connect(lambda: dialog.done(1))
    
    pdf_btn = QtWidgets.QPushButton("📄 PDF (.pdf)")
    pdf_btn.clicked.connect(lambda: dialog.done(2))
    
    cancel_btn = QtWidgets.QPushButton("Cancelar")
    cancel_btn.clicked.connect(dialog.reject)
    
    layout.addWidget(excel_btn)
    layout.addWidget(pdf_btn)
    layout.addWidget(cancel_btn)
    
    result = dialog.exec()
    
    if result == 1:
        return 'excel'
    elif result == 2:
        return 'pdf'
    else:
        return None

def export_report(data: list[dict[str, Any]], title: str = "Reporte", parent=None) -> bool:
    """Main export function with format selection.
    
    Args:
        data: Report data to export
        title: Report title
        parent: Parent widget for dialogs
        
    Returns:
        True if successful, False otherwise
    """
    if not data:
        QtWidgets.QMessageBox.information(
            parent,
            "Sin datos",
            "No hay datos para exportar"
        )
        return False
    
    # Select format
    format_type = prompt_export_format(parent)
    if not format_type:
        return False
    
    # Select filename
    if format_type == 'excel':
        ext = "Excel Files (*.xlsx)"
        default_name = f"{title}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    else:
        ext = "PDF Files (*.pdf)"
        default_name = f"{title}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    filename, _ = QtWidgets.QFileDialog.getSaveFileName(
        parent,
        "Guardar reporte",
        default_name,
        ext
    )
    
    if not filename:
        return False
    
    # Export
    if format_type == 'excel':
        success = export_to_excel(data, filename, title)
    else:
        success = export_to_pdf(data, filename, title)
    
    if success:
        QtWidgets.QMessageBox.information(
            parent,
            "Éxito",
            f"Reporte exportado exitosamente:\n{filename}"
        )
    
    return success
