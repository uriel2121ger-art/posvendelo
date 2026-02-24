import logging

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

def export_inventory_to_excel(data, filename):
    """
    Exporta datos de inventario/productos a Excel con TODAS las columnas.
    Ahora dinámico - exporta todas las columnas que vengan en los datos.
    """
    if not HAS_OPENPYXL:
        logger.error("openpyxl no está instalado. No se puede exportar a Excel.")
        return False
    
    if not data:
        logger.warning("No hay datos para exportar")
        return False
        
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario"
        
        # Obtener TODAS las columnas dinámicamente del primer registro
        all_keys = list(data[0].keys())
        
        # Mapeo de nombres técnicos a nombres amigables
        friendly_names = {
            'id': 'ID',
            'sku': 'SKU',
            'name': 'Nombre',
            'price': 'Precio',
            'price_wholesale': 'Precio Mayoreo',
            'cost': 'Costo',
            'stock': 'Stock',
            'category_id': 'ID Categoría',
            'category': 'Categoría',
            'department': 'Departamento',
            'provider': 'Proveedor',
            'min_stock': 'Stock Mínimo',
            'max_stock': 'Stock Máximo',
            'is_active': 'Activo',
            'is_kit': 'Es Kit',
            'is_favorite': 'Favorito',
            'tax_scheme': 'Esquema Fiscal',
            'tax_rate': 'Tasa Impuesto',
            'sale_type': 'Tipo Venta',
            'barcode': 'Código Barras',
            'description': 'Descripción',
            'notes': 'Notas',
            'shadow_stock': 'Stock Oculto',
            'cost_a': 'Costo Serie A',
            'cost_b': 'Costo Serie B',
            'qty_from_a': 'Cantidad Serie A',
            'qty_from_b': 'Cantidad Serie B',
            'cost_price': 'Precio de Costo',
            'status': 'Estado',
            'supplier_id': 'ID Proveedor',
            'visible': 'Visible',
            'entry_date': 'Fecha Entrada',
            'sat_clave_prod_serv': 'SAT Clave Prod/Serv',
            'sat_clave_unidad': 'SAT Clave Unidad',
            'sat_descripcion': 'SAT Descripción',
            'sat_code': 'Código SAT',
            'sat_unit': 'Unidad SAT',
            'created_at': 'Fecha Creación',
            'updated_at': 'Fecha Actualización'
        }
        
        # Crear encabezados con nombres amigables
        headers = [friendly_names.get(key, key.replace('_', ' ').title()) for key in all_keys]
        ws.append(headers)
        
        # Estilo encabezado
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            
        # Ajustar ancho de columnas automáticamente
        for idx, key in enumerate(all_keys, 1):
            column_letter = get_column_letter(idx)
            # Ancho basado en el nombre del header + padding
            max_length = len(headers[idx-1])
            ws.column_dimensions[column_letter].width = min(max_length + 5, 50)
            
        # Agregar datos fila por fila
        for item in data:
            row = [item.get(key, '') for key in all_keys]
            ws.append(row)
            
        # Congelar primera fila (encabezados)
        ws.freeze_panes = "A2"
        
        # Auto-filtro en encabezados
        ws.auto_filter.ref = ws.dimensions
        
        wb.save(filename)
        logger.info(f"✅ Inventario exportado a {filename} con {len(all_keys)} columnas y {len(data)} registros")
        return True
        
    except Exception as e:
        logger.error(f"❌ Error exportando inventario a Excel: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False

def export_product_catalog_to_excel(data, filename):
    """Exporta catálogo de productos a Excel con todas las columnas."""
    return export_inventory_to_excel(data, filename)

def export_sales_to_excel(data, filename):
    """
    Exporta datos de ventas a Excel con todas las columnas.
    Dinámico - adapta a las columnas disponibles.
    """
    if not HAS_OPENPYXL:
        logger.error("openpyxl no está instalado.")
        return False
    
    if not data:
        logger.warning("No hay ventas para exportar")
        return False
        
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Ventas"
        
        # Obtener todas las columnas
        all_keys = list(data[0].keys())
        
        # Nombres amigables para ventas
        friendly_names = {
            'id': 'ID',
            'uuid': 'UUID',
            'folio': 'Folio',
            'serie': 'Serie',
            'timestamp': 'Fecha/Hora',
            'created_at': 'Fecha Creación',
            'subtotal': 'Subtotal',
            'tax': 'Impuesto',
            'total': 'Total',
            'discount': 'Descuento',
            'payment_method': 'Método Pago',
            'customer_id': 'ID Cliente',
            'customer_name': 'Cliente',
            'user_id': 'ID Usuario',
            'username': 'Usuario',
            'turn_id': 'ID Turno',
            'branch_id': 'ID Sucursal',
            'terminal_id': 'ID Terminal',
            'status': 'Estado',
            'notes': 'Notas',
            'cfdi_uuid': 'UUID CFDI',
            'invoice_status': 'Estado Factura'
        }
        
        headers = [friendly_names.get(key, key.replace('_', ' ').title()) for key in all_keys]
        ws.append(headers)
        
        # Estilo
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2E75B5", end_color="2E75B5", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        # Auto-width
        for idx, key in enumerate(all_keys, 1):
            ws.column_dimensions[get_column_letter(idx)].width = min(len(headers[idx-1]) + 5, 40)
            
        # Datos
        for item in data:
            row = [item.get(key, '') for key in all_keys]
            ws.append(row)
            
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        
        wb.save(filename)
        logger.info(f"✅ Ventas exportadas a {filename} con {len(all_keys)} columnas")
        return True
        
    except Exception as e:
        logger.error(f"❌ Error exportando ventas: {e}")
        return False

def export_customers_to_excel(data, filename):
    """Exporta clientes a Excel con todas las columnas."""
    if not HAS_OPENPYXL or not data:
        return False
        
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes"
        
        all_keys = list(data[0].keys())
        
        friendly_names = {
            'id': 'ID',
            'name': 'Nombre',
            'phone': 'Teléfono',
            'email': 'Email',
            'address': 'Dirección',
            'rfc': 'RFC',
            'credit_limit': 'Límite Crédito',
            'credit_balance': 'Saldo Crédito',
            'points': 'Puntos',
            'notes': 'Notas',
            'is_active': 'Activo',
            'created_at': 'Fecha Creación',
            'updated_at': 'Última Actualización'
        }
        
        headers = [friendly_names.get(key, key.replace('_', ' ').title()) for key in all_keys]
        ws.append(headers)
        
        # Estilo
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            
        for item in data:
            ws.append([item.get(key, '') for key in all_keys])
            
        ws.freeze_panes = "A2"
        wb.save(filename)
        logger.info(f"✅ Clientes exportados con {len(all_keys)} columnas")
        return True
        
    except Exception as e:
        logger.error(f"❌ Error exportando clientes: {e}")
        return False
