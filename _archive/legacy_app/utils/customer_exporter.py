import csv
import logging

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

def export_customers_to_csv(customers, filename):
    """Exporta clientes a CSV con TODOS los campos."""
    try:
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            # Encabezados completos
            writer.writerow([
                "ID", "Nombre", "Apellido", "Teléfono", "Email", "Email Fiscal",
                "RFC", "Razón Social", "Régimen Fiscal",
                "Domicilio 1", "Domicilio 2", "Colonia", "Municipio", "Estado", "País", "Código Postal",
                "VIP", "Crédito Autorizado", "Límite de Crédito", "Saldo de Crédito",
                "Puntos", "Saldo Monedero", "Notas"
            ])
            for c in customers:
                writer.writerow([
                    c.get('id', ''),
                    c.get('first_name', ''),
                    c.get('last_name', ''),
                    c.get('phone', ''),
                    c.get('email', ''),
                    c.get('email_fiscal', ''),
                    c.get('rfc', ''),
                    c.get('razon_social', ''),
                    c.get('regimen_fiscal', ''),
                    c.get('domicilio1', ''),
                    c.get('domicilio2', ''),
                    c.get('colonia', ''),
                    c.get('municipio', ''),
                    c.get('estado', ''),
                    c.get('pais', 'México'),
                    c.get('codigo_postal', ''),
                    'Sí' if c.get('vip') else 'No',
                    'Sí' if c.get('credit_authorized') else 'No',
                    c.get('credit_limit', 0),
                    c.get('credit_balance', 0),
                    c.get('points', 0),
                    c.get('wallet_balance', 0),
                    c.get('notes', '')
                ])
        return True
    except Exception as e:
        logger.error(f"Error exportando CSV: {e}")
        return False

def export_customers_to_excel(customers, filename):
    """Exporta clientes a Excel con TODOS los campos."""
    if not HAS_OPENPYXL:
        logger.error("openpyxl no instalado")
        return False
        
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes"
        
        # Encabezados completos
        headers = [
            "ID", "Nombre", "Apellido", "Teléfono", "Email", "Email Fiscal",
            "RFC", "Razón Social", "Régimen Fiscal",
            "Domicilio 1", "Domicilio 2", "Colonia", "Municipio", "Estado", "País", "Código Postal",
            "VIP", "Crédito Autorizado", "Límite de Crédito", "Saldo de Crédito",
            "Puntos", "Saldo Monedero", "Notas"
        ]
        ws.append(headers)
        
        # Estilo para encabezados
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Datos
        for c in customers:
            ws.append([
                c.get('id', ''),
                c.get('first_name', ''),
                c.get('last_name', ''),
                c.get('phone', ''),
                c.get('email', ''),
                c.get('email_fiscal', ''),
                c.get('rfc', ''),
                c.get('razon_social', ''),
                c.get('regimen_fiscal', ''),
                c.get('domicilio1', ''),
                c.get('domicilio2', ''),
                c.get('colonia', ''),
                c.get('municipio', ''),
                c.get('estado', ''),
                c.get('pais', 'México'),
                c.get('codigo_postal', ''),
                'Sí' if c.get('vip') else 'No',
                'Sí' if c.get('credit_authorized') else 'No',
                float(c.get('credit_limit', 0)),
                float(c.get('credit_balance', 0)),
                float(c.get('points', 0)),
                float(c.get('wallet_balance', 0)),
                c.get('notes', '')
            ])
            
        wb.save(filename)
        return True
    except Exception as e:
        logger.error(f"Error exportando Excel: {e}")
        return False
