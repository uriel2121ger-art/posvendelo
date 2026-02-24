"""
Report Export Dialog - Export reports to CSV, Excel, or PDF
"""

import csv
from datetime import datetime
from pathlib import Path

from PyQt6 import QtCore, QtWidgets
from PyQt6.QtWidgets import (
    QComboBox,
    QDialog,
    QFileDialog,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QMessageBox,
    QPushButton,
    QVBoxLayout,
    QWidget,
)


class ReportExportDialog(QDialog):
    """Dialog for exporting report data to various formats."""
    
    def __init__(self, title=None, headers=None, rows=None, pdf_exporter=None, parent=None):
        super().__init__(parent)
        self.title = title or "Reporte"
        self.headers = headers or []
        self.rows = rows or []
        self.pdf_exporter = pdf_exporter
        
        self.setWindowTitle(f"Exportar: {self.title}")
        self.setMinimumSize(400, 250)
        
        self._build_ui()
    
    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Header
        header = QLabel(f"📊 Exportar {self.title}")
        header.setStyleSheet("font-size: 18px; font-weight: bold;")
        layout.addWidget(header)
        
        # Info
        info = QLabel(f"Registros a exportar: {len(self.rows)}")
        info.setStyleSheet("color: #888;")
        layout.addWidget(info)
        
        # Format selection
        format_group = QGroupBox("Formato de exportación")
        format_layout = QVBoxLayout(format_group)
        
        self.format_combo = QComboBox()
        self.format_combo.addItem("📄 CSV (Valores separados por coma)", "csv")
        self.format_combo.addItem("📊 Excel (XLSX)", "xlsx")
        if self.pdf_exporter:
            self.format_combo.addItem("📕 PDF", "pdf")
        self.format_combo.setFixedHeight(35)
        format_layout.addWidget(self.format_combo)
        
        layout.addWidget(format_group)
        
        layout.addStretch()
        
        # Buttons
        btn_layout = QHBoxLayout()
        
        cancel_btn = QPushButton("Cancelar")
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setFixedHeight(40)
        btn_layout.addWidget(cancel_btn)
        
        export_btn = QPushButton("📥 Exportar")
        export_btn.clicked.connect(self._do_export)
        export_btn.setFixedHeight(40)
        export_btn.setStyleSheet("font-weight: bold;")
        btn_layout.addWidget(export_btn)
        
        layout.addLayout(btn_layout)
    
    def _do_export(self):
        """Perform the export based on selected format."""
        if not self.rows:
            QMessageBox.warning(self, "Sin datos", "No hay datos para exportar.")
            return
        
        format_type = self.format_combo.currentData()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"{self.title.replace(' ', '_')}_{timestamp}"
        
        if format_type == "csv":
            self._export_csv(default_name)
        elif format_type == "xlsx":
            self._export_excel(default_name)
        elif format_type == "pdf":
            self._export_pdf(default_name)
    
    def _export_csv(self, default_name: str):
        """Export to CSV format."""
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar CSV",
            f"{default_name}.csv",
            "CSV (*.csv)"
        )
        
        if not file_path:
            return
        
        try:
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(self.headers)
                writer.writerows(self.rows)
            
            QMessageBox.information(
                self, "Exportación exitosa",
                f"✅ Archivo guardado:\n{file_path}"
            )
            self.accept()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al exportar: {e}")
    
    def _export_excel(self, default_name: str):
        """Export to Excel format."""
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar Excel",
            f"{default_name}.xlsx",
            "Excel (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            # Try openpyxl first
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Alignment, Font, PatternFill
                
                wb = Workbook()
                ws = wb.active
                ws.title = self.title[:31]  # Excel sheet name limit
                
                # Style for headers
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                
                # Write headers
                for col, header in enumerate(self.headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                # Write data
                for row_idx, row in enumerate(self.rows, 2):
                    for col_idx, value in enumerate(row, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Auto-adjust column widths
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except Exception:
                            pass
                    ws.column_dimensions[column].width = min(max_length + 2, 50)
                
                wb.save(file_path)
                
            except ImportError:
                # Fallback to CSV if openpyxl not available
                QMessageBox.warning(
                    self, "openpyxl no disponible",
                    "La librería openpyxl no está instalada. Exportando como CSV..."
                )
                file_path = file_path.replace('.xlsx', '.csv')
                with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
                    writer = csv.writer(f)
                    writer.writerow(self.headers)
                    writer.writerows(self.rows)
            
            QMessageBox.information(
                self, "Exportación exitosa",
                f"✅ Archivo guardado:\n{file_path}"
            )
            self.accept()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al exportar: {e}")
    
    def _export_pdf(self, default_name: str):
        """Export to PDF format."""
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar PDF",
            f"{default_name}.pdf",
            "PDF (*.pdf)"
        )
        
        if not file_path:
            return
        
        try:
            if self.pdf_exporter:
                self.pdf_exporter(file_path)
                QMessageBox.information(
                    self, "Exportación exitosa",
                    f"✅ PDF guardado:\n{file_path}"
                )
                self.accept()
            else:
                QMessageBox.warning(
                    self, "No disponible",
                    "Exportación a PDF no disponible para este reporte."
                )
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al exportar PDF: {e}")

    def showEvent(self, event):
        """Apply theme colors when dialog is shown."""
        super().showEvent(event)
        try:
            from app.utils.theme_manager import theme_manager
            c = theme_manager.get_colors()
            
            self.setStyleSheet(f"""
                QDialog {{
                    background: {c['bg_card']};
                    color: {c['text_primary']};
                }}
                QGroupBox {{
                    border: 1px solid {c['border']};
                    border-radius: 8px;
                    margin-top: 10px;
                    padding-top: 15px;
                    font-weight: bold;
                    color: {c['text_primary']};
                }}
                QGroupBox::title {{
                    subcontrol-origin: margin;
                    left: 10px;
                    padding: 0 5px;
                }}
                QComboBox {{
                    background: {c['input_bg']};
                    border: 1px solid {c['border']};
                    border-radius: 5px;
                    padding: 8px;
                    color: {c['text_primary']};
                }}
            """)
            
            for btn in self.findChildren(QPushButton):
                text = btn.text().lower()
                if 'exportar' in text:
                    btn.setStyleSheet(f"background: {c['btn_success']}; color: white; font-weight: bold; padding: 8px; border-radius: 5px;")
                elif 'cancelar' in text:
                    btn.setStyleSheet(f"background: {c['btn_danger']}; color: white; font-weight: bold; padding: 8px; border-radius: 5px;")
                    
        except Exception:
            pass
