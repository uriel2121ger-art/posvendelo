from __future__ import annotations

from typing import Any
from datetime import datetime, timedelta

from PyQt6 import QtCore, QtGui, QtWidgets

from app.core import STATE, POSCore
from app.dialogs.product_editor import ProductEditorDialog
from app.dialogs.product_search import ProductSearchDialog
from app.utils.export_csv import export_inventory_to_csv
from app.utils.theme_manager import theme_manager


class InventoryTab(QtWidgets.QWidget):
    """Módulo de Control de Inventario con 6 pestañas especializadas"""
    
    def __init__(self, core: POSCore, parent: QtWidgets.QWidget | None = None):
        super().__init__(parent)
        self.core = core
        self.load_assets()
        self._build_ui()
    
    def load_assets(self):
        self.icons = {}
        try:
            self.icons["inventory"] = QtGui.QIcon("assets/icon_inventory.png")
            self.icons["search"] = QtGui.QIcon("assets/icon_search.png")
            self.icons["save"] = QtGui.QIcon("assets/icon_add.png")
            self.icons["list"] = QtGui.QIcon("assets/icon_shifts.png")
        except Exception as e:
            pass  # Icons are optional, fail silently
    
    def _build_ui(self) -> None:
        theme = (self.core.get_app_config() or {}).get("theme", "Light")
        c = theme_manager.get_colors(theme)
        
        main_layout = QtWidgets.QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        
        # === HEADER ===
        self.header = QtWidgets.QFrame()
        self.header.setFixedHeight(70)
        self.header.setStyleSheet(f"""
            QFrame {{
                background: {c['bg_header']};
                border-bottom: 1px solid {c['border']};
            }}
        """)
        header_layout = QtWidgets.QHBoxLayout(self.header)
        header_layout.setContentsMargins(20, 0, 20, 0)
        
        if "inventory" in self.icons:
            icon_lbl = QtWidgets.QLabel()
            icon_lbl.setPixmap(self.icons["inventory"].pixmap(32, 32))
            header_layout.addWidget(icon_lbl)
        
        self.title_label = QtWidgets.QLabel("CONTROL DE INVENTARIO")
        self.title_label.setStyleSheet(f"color: {c['text_header']}; font-size: 20px; font-weight: 800; letter-spacing: 1px; background: transparent;")
        header_layout.addWidget(self.title_label)
        header_layout.addStretch()
        
        main_layout.addWidget(self.header)
        
        # === TAB WIDGET ===
        self.tab_widget = QtWidgets.QTabWidget()
        self.tab_widget.setStyleSheet(f"""
            QTabWidget::pane {{
                border: none;
                background: {c['bg_main']};
            }}
            QTabBar::tab {{
                background: {c['bg_card']};
                color: {c['text_secondary']};
                padding: 12px 20px;
                margin-right: 2px;
                border: 1px solid {c['border']};
                border-bottom: none;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                font-weight: 600;
            }}
            QTabBar::tab:selected {{
                background: {c['bg_main']};
                color: {c['text_primary']};
                border-bottom: 2px solid {c['btn_primary']};
            }}
            QTabBar::tab:hover {{
                background: {c['border']};
            }}
        """)
        
        # Crear las 6 pestañas
        self.tab_consulta = ConsultaTab(self.core, self.icons)
        self.tab_ajustes = AjustesTab(self.core, self.icons)
        self.tab_movimientos = MovimientosTab(self.core)
        self.tab_transferencias = TransferenciasTab(self.core)
        self.tab_alertas = AlertasTab(self.core)
        self.tab_analisis = AnalisisTab(self.core)
        
        # Agregar pestañas
        self.tab_widget.addTab(self.tab_consulta, "📊 Consulta")
        self.tab_widget.addTab(self.tab_ajustes, "⚙️ Ajustes")
        self.tab_widget.addTab(self.tab_movimientos, "📜 Movimientos")
        # self.tab_widget.addTab(self.tab_transferencias, "📦 Transferencias") # Removed by request
        self.tab_widget.addTab(self.tab_alertas, "🔔 Alertas")
        self.tab_widget.addTab(self.tab_analisis, "📈 Análisis")
        
        # Conectar señal de cambio de pestaña
        self.tab_widget.currentChanged.connect(self._on_tab_changed)
        
        main_layout.addWidget(self.tab_widget)
    
    def _on_tab_changed(self, index: int):
        """Refrescar datos cuando se cambia de pestaña"""
        current_tab = self.tab_widget.widget(index)
        if hasattr(current_tab, 'refresh_data'):
            current_tab.refresh_data()
    
    def showEvent(self, event):
        """Aplicar tema cuando se muestra el tab."""
        super().showEvent(event)
        if hasattr(self, 'update_theme'):
            self.update_theme()
    
    def update_theme(self) -> None:
        """Actualizar tema en todas las pestañas"""
        cfg = self.core.read_local_config()
        theme = cfg.get("theme", "Light")
        c = theme_manager.get_colors(theme)
        
        if hasattr(self, "header"):
            self.header.setStyleSheet(f"""
                QFrame {{
                    background: {c['bg_header']};
                    border-bottom: 1px solid {c['border']};
                }}
            """)
        
        if hasattr(self, "title_label"):
            self.title_label.setStyleSheet(f"color: {c['text_header']}; font-size: 20px; font-weight: 800; letter-spacing: 1px; background: transparent;")
        
        if hasattr(self, "tab_widget"):
            self.tab_widget.setStyleSheet(f"""
                QTabWidget::pane {{
                    border: none;
                    background: {c['bg_main']};
                }}
                QTabBar::tab {{
                    background: {c['bg_card']};
                    color: {c['text_secondary']};
                    padding: 12px 20px;
                    margin-right: 2px;
                    border: 1px solid {c['border']};
                    border-bottom: none;
                    border-top-left-radius: 8px;
                    border-top-right-radius: 8px;
                    font-weight: 600;
                }}
                QTabBar::tab:selected {{
                    background: {c['bg_main']};
                    color: {c['text_primary']};
                    border-bottom: 2px solid {c['btn_primary']};
                }}
                QTabBar::tab:hover {{
                    background: {c['border']};
                }}
            """)
        
        # Actualizar tema en cada pestaña
        for i in range(self.tab_widget.count()):
            tab = self.tab_widget.widget(i)
            if hasattr(tab, 'update_theme'):
                tab.update_theme()

# ============================================================================
# PESTAÑA 1: CONSULTA
# ============================================================================

class ConsultaTab(QtWidgets.QWidget):
    """Pestaña de consulta de inventario (solo lectura)"""
    
    def __init__(self, core: POSCore, icons: dict):
        super().__init__()
        self.core = core
        self.icons = icons
        self._build_ui()
        self.refresh_data()
    
    def _build_ui(self):
        theme = (self.core.get_app_config() or {}).get("theme", "Light")
        c = theme_manager.get_colors(theme)
        
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(20)
        
        # Cards de resumen
        cards_layout = QtWidgets.QHBoxLayout()
        cards_layout.setSpacing(15)
        
        self.card_total = self._create_card("Total Productos", "0", "#4CAF50", c)
        self.card_low = self._create_card("Bajo Stock", "0", "#FF9800", c)
        self.card_value = self._create_card("Valor Total", "$0.00", "#2196F3", c)
        
        cards_layout.addWidget(self.card_total)
        cards_layout.addWidget(self.card_low)
        cards_layout.addWidget(self.card_value)
        
        layout.addLayout(cards_layout)
        
        # Búsqueda
        search_frame = QtWidgets.QFrame()
        search_frame.setStyleSheet(f"""
            QFrame {{
                background: {c['bg_card']};
                border: 1px solid {c['border']};
                border-radius: 10px;
            }}
        """)
        search_layout = QtWidgets.QHBoxLayout(search_frame)
        search_layout.setContentsMargins(15, 15, 15, 15)
        
        self.search_input = QtWidgets.QLineEdit()
        self.search_input.setPlaceholderText("🔍 Buscar en inventario (Nombre o SKU)...")
        self.search_input.setFixedHeight(40)
        self.search_input.textChanged.connect(self.filter_table)
        self.search_input.setStyleSheet(f"""
            QLineEdit {{
                background: {c['input_bg']};
                border: 1px solid {c['input_border']};
                border-radius: 6px;
                padding: 0 15px;
                font-size: 13px;
                color: {c['text_primary']};
            }}
            QLineEdit:focus {{
                border: 2px solid {c['input_focus']};
                background: {c['bg_card']};
            }}
        """)
        search_layout.addWidget(self.search_input)
        
        layout.addWidget(search_frame)
        
        # Tabla
        self.table = QtWidgets.QTableWidget(0, 7)
        self.table.setEditTriggers(QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setHorizontalHeaderLabels(["ID", "SKU", "Nombre", "Tipo", "Stock", "Mín", "Máx"])
        self.table.setSelectionBehavior(QtWidgets.QTableWidget.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QtWidgets.QTableWidget.SelectionMode.SingleSelection)
        self.table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeMode.Stretch)
        self.table.verticalHeader().setVisible(False)
        self.table.setAlternatingRowColors(False)
        self.table.itemDoubleClicked.connect(self._edit_product)
        
        self.table.setStyleSheet(f"""
            QTableWidget {{
                background: {c['bg_card']};
                border: 1px solid {c['border']};
                border-radius: 10px;
                gridline-color: transparent;
                font-size: 13px;
            }}
            QHeaderView::section {{
                background: {c['table_header_bg']};
                color: {c['table_header_text']};
                padding: 12px;
                border: none;
                border-bottom: 1px solid {c['border']};
                font-weight: bold;
                font-size: 12px;
            }}
            QTableWidget::item {{
                padding: 10px;
                border-bottom: 1px solid {c['border']};
                color: {c['table_text']};
            }}
            QTableWidget::item:selected {{
                background: {c['table_selected']};
                color: {c['bg_header']};
            }}
        """)
        
        layout.addWidget(self.table)
    
    def _create_card(self, title: str, value: str, color: str, c: dict) -> QtWidgets.QFrame:
        card = QtWidgets.QFrame()
        card.setStyleSheet(f"""
            QFrame {{
                background-color: {c['bg_card']};
                border: 1px solid {c['border']};
                border-radius: 10px;
                border-left: 5px solid {color};
            }}
        """)
        
        layout = QtWidgets.QVBoxLayout(card)
        layout.setSpacing(5)
        
        title_label = QtWidgets.QLabel(title)
        title_label.setStyleSheet(f"color: {c['text_secondary']}; font-size: 13px; font-weight: 600; border: none;")
        
        value_label = QtWidgets.QLabel(value)
        value_label.setObjectName("value_label")
        value_label.setStyleSheet(f"color: {color}; font-size: 28px; font-weight: bold; border: none;")
        
        layout.addWidget(title_label)
        layout.addWidget(value_label)
        
        return card
    
    def refresh_data(self):
        """Refrescar datos de la tabla"""
        products = self.core.list_products_for_export()
        
        # Actualizar cards
        total = len(products)
        low_stock = sum(1 for p in products if float(p.get('stock', 0) or 0) < float(p.get('min_stock', 0) or 0))
        total_value = sum(float(p.get('price', 0) or 0) * float(p.get('stock', 0) or 0) for p in products)
        
        self.card_total.findChild(QtWidgets.QLabel, "value_label").setText(str(total))
        self.card_low.findChild(QtWidgets.QLabel, "value_label").setText(str(low_stock))
        
        if STATE.role == "admin":
            self.card_value.findChild(QtWidgets.QLabel, "value_label").setText(f"${total_value:,.2f}")
        else:
            self.card_value.findChild(QtWidgets.QLabel, "value_label").setText("$***")
        
        # Actualizar tabla
        self.table.setRowCount(0)
        self.table.setRowCount(len(products))
        
        for row_idx, row in enumerate(products):
            values = [
                str(row.get("id") or row.get("product_id")),
                row.get("sku") or "",
                row.get("name"),
                row.get("sale_type") or "unidad",
                f"{float(row.get('stock', 0.0) or 0.0):.2f}",
                f"{float(row.get('min_stock', 0.0) or 0.0):.2f}",
                f"{float(row.get('max_stock', 0.0) or 0.0):.2f}",
            ]
            for col, value in enumerate(values):
                cell = QtWidgets.QTableWidgetItem(str(value))
                cell.setFlags(cell.flags() & ~QtCore.Qt.ItemFlag.ItemIsEditable)
                self.table.setItem(row_idx, col, cell)
    
    def filter_table(self, text: str):
        text = text.lower().strip()
        for i in range(self.table.rowCount()):
            sku = self.table.item(i, 1).text().lower()
            name = self.table.item(i, 2).text().lower()
            if not text or text in sku or text in name:
                self.table.setRowHidden(i, False)
            else:
                self.table.setRowHidden(i, True)
    
    def _edit_product(self):
        # Restriction: Only admin can edit products
        if STATE.role != "admin":
            return
            
        row = self.table.currentRow()
        if row < 0:
            return
        try:
            product_id = int(self.table.item(row, 0).text())
        except (ValueError, TypeError, AttributeError):
            return
        product_data = self.core.get_product_by_id(product_id)
        if not product_data:
            QtWidgets.QMessageBox.warning(self, "Error", "No se pudo cargar el producto.")
            return
        dlg = ProductEditorDialog(core=self.core, product=dict(product_data), parent=self)
        if dlg.exec() == QtWidgets.QDialog.DialogCode.Accepted:
            self.refresh_data()
    
    def update_theme(self):
        theme = (self.core.get_app_config() or {}).get("theme", "Light")
        c = theme_manager.get_colors(theme)
        
        # Cards (need to rebuild or update stylesheet)
        # For simplicity, we assume cards frame style update is enough if we kept references, 
        # but _create_card returns a new object. We should have stored references.
        # Since we stored references (self.card_total, etc.), we can update them.
        
        for card, color in [(self.card_total, "#4CAF50"), (self.card_low, "#FF9800"), (self.card_value, "#2196F3")]:
            card.setStyleSheet(f"""
                QFrame {{
                    background-color: {c['bg_card']};
                    border: 1px solid {c['border']};
                    border-radius: 10px;
                    border-left: 5px solid {color};
                }}
            """)
            # Update labels inside card
            # Using findChildren might be needed if we didn't store label refs
            for lbl in card.findChildren(QtWidgets.QLabel):
                if lbl.objectName() == "value_label":
                    lbl.setStyleSheet(f"color: {color}; font-size: 28px; font-weight: bold; border: none;")
                else:
                    lbl.setStyleSheet(f"color: {c['text_secondary']}; font-size: 13px; font-weight: 600; border: none;")

        # Search
        self.search_input.setStyleSheet(f"""
            QLineEdit {{
                background: {c['input_bg']};
                border: 1px solid {c['input_border']};
                border-radius: 6px;
                padding: 0 15px;
                font-size: 13px;
                color: {c['text_primary']};
            }}
            QLineEdit:focus {{
                border: 2px solid {c['input_focus']};
                background: {c['bg_card']};
            }}
        """)
        
        # Table
        self.table.setStyleSheet(f"""
            QTableWidget {{
                background: {c['bg_card']};
                border: 1px solid {c['border']};
                border-radius: 10px;
                gridline-color: transparent;
                font-size: 13px;
            }}
            QHeaderView::section {{
                background: {c['table_header_bg']};
                color: {c['table_header_text']};
                padding: 12px;
                border: none;
                border-bottom: 1px solid {c['border']};
                font-weight: bold;
                font-size: 12px;
                color: {c['table_header_text']};
            }}
            QTableWidget::item {{
                padding: 10px;
                border-bottom: 1px solid {c['border']};
                color: {c['table_text']};
            }}
            QTableWidget::item:selected {{
                background: {c['table_selected']};
                color: {c['bg_header']};
            }}
        """)

# ============================================================================
# PESTAÑA 2: AJUSTES
# ============================================================================

class AjustesTab(QtWidgets.QWidget):
    """Pestaña para ajustar stock manualmente"""
    
    def __init__(self, core: POSCore, icons: dict):
        super().__init__()
        self.core = core
        self.icons = icons
        self._build_ui()
        
        # Initial permission check
        if STATE.role != "admin":
            # These are created in _build_ui, ensuring they exist now
            self.apply_btn.setEnabled(False)
            self.apply_btn.setToolTip("Requiere permisos de Administrador")
            self.qty_input.setEnabled(False)
            self.sku_input.setPlaceholderText("Búsqueda bloqueada - Solo Admin")
            self.search_btn.setEnabled(False)
    
    def _build_ui(self):
        theme = (self.core.get_app_config() or {}).get("theme", "Light")
        c = theme_manager.get_colors(theme)
        
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(20)
        
        # Formulario de ajuste
        form_frame = QtWidgets.QFrame()
        form_frame.setStyleSheet(f"""
            QFrame {{
                background: {c['bg_card']};
                border: 1px solid {c['border']};
                border-radius: 10px;
            }}
        """)
        form_layout = QtWidgets.QHBoxLayout(form_frame)
        form_layout.setContentsMargins(15, 15, 15, 15)
        form_layout.setSpacing(15)
        
        # SKU
        sku_container = QtWidgets.QVBoxLayout()
        sku_label = QtWidgets.QLabel("SKU / Código")
        sku_label.setStyleSheet(f"font-weight: bold; color: {c['text_secondary']}; font-size: 12px;")
        self.sku_input = QtWidgets.QLineEdit()
        self.sku_input.setPlaceholderText("Ingresa SKU...")
        self.sku_input.setFixedHeight(40)
        self.sku_input.setStyleSheet(f"""
            QLineEdit {{
                background: {c['input_bg']};
                border: 1px solid {c['input_border']};
                border-radius: 6px;
                padding: 0 10px;
                font-size: 13px;
                color: {c['text_primary']};
            }}
            QLineEdit:focus {{
                border: 2px solid {c['input_focus']};
                background: {c['bg_card']};
            }}
        """)
        sku_container.addWidget(sku_label)
        sku_container.addWidget(self.sku_input)
        form_layout.addLayout(sku_container, 2)
        
        # Botón Buscar
        self.search_btn = QtWidgets.QPushButton(" Buscar")
        if "search" in self.icons:
            self.search_btn.setIcon(self.icons["search"])
        self.search_btn.clicked.connect(self._pick_product)
        self.search_btn.setCursor(QtCore.Qt.CursorShape.PointingHandCursor)
        self.search_btn.setFixedHeight(40)
        self.search_btn.setMinimumWidth(100)
        self.search_btn.setStyleSheet(f"""
            QPushButton {{
                background: {c['btn_primary']};
                color: white;
                border: none;
                border-radius: 6px;
                font-weight: bold;
                margin-top: 18px;
                padding: 0 15px;
            }}
            QPushButton:hover {{
                opacity: 0.9;
            }}
        """)
        form_layout.addWidget(self.search_btn)
        
        # Cantidad
        qty_container = QtWidgets.QVBoxLayout()
        qty_label = QtWidgets.QLabel("Cantidad")
        qty_label.setStyleSheet(f"font-weight: bold; color: {c['text_secondary']}; font-size: 12px;")
        self.qty_input = QtWidgets.QSpinBox()
        self.qty_input.setRange(-1_000_000, 1_000_000)
        self.qty_input.setFixedHeight(40)
        self.qty_input.setStyleSheet(f"""
            QSpinBox {{
                background: {c['input_bg']};
                border: 1px solid {c['input_border']};
                border-radius: 6px;
                padding: 0 10px;
                font-size: 13px;
                color: {c['text_primary']};
            }}
            QSpinBox:focus {{
                border: 2px solid {c['input_focus']};
                background: {c['bg_card']};
            }}
        """)
        qty_container.addWidget(qty_label)
        qty_container.addWidget(self.qty_input)
        form_layout.addLayout(qty_container, 1)
        
        # Razón
        reason_container = QtWidgets.QVBoxLayout()
        reason_label = QtWidgets.QLabel("Razón")
        reason_label.setStyleSheet(f"font-weight: bold; color: {c['text_secondary']}; font-size: 12px;")
        self.reason_combo = QtWidgets.QComboBox()
        self.reason_combo.addItems([
            "Corrección",
            "Merma",
            "Robo",
            "Devolución",
            "Inventario Físico",
            "Otro"
        ])
        self.reason_combo.setFixedHeight(40)
        self.reason_combo.setStyleSheet(f"""
            QComboBox {{
                background: {c['input_bg']};
                border: 1px solid {c['input_border']};
                border-radius: 6px;
                padding: 0 10px;
                font-size: 13px;
                color: {c['text_primary']};
            }}
            QComboBox:focus {{
                border: 2px solid {c['input_focus']};
                background: {c['bg_card']};
            }}
        """)
        reason_container.addWidget(reason_label)
        reason_container.addWidget(self.reason_combo)
        form_layout.addLayout(reason_container, 2)
        
        # Botón Aplicar
        self.apply_btn = QtWidgets.QPushButton(" Aplicar Ajuste")
        if "save" in self.icons:
            self.apply_btn.setIcon(self.icons["save"])
        self.apply_btn.clicked.connect(self.adjust_stock)
        self.apply_btn.setCursor(QtCore.Qt.CursorShape.PointingHandCursor)
        self.apply_btn.setFixedHeight(40)
        self.apply_btn.setMinimumWidth(140)
        self.apply_btn.setStyleSheet(f"""
            QPushButton {{
                background: {c['btn_success']};
                color: white;
                border: none;
                border-radius: 6px;
                font-weight: bold;
                margin-top: 18px;
                padding: 0 15px;
            }}
            QPushButton:hover {{
                opacity: 0.9;
            }}
        """)
        form_layout.addWidget(self.apply_btn)
        
        layout.addWidget(form_frame)
        
        # Tabla de ajustes recientes
        table_label = QtWidgets.QLabel("Ajustes Recientes")
        table_label.setStyleSheet(f"color: {c['text_primary']}; font-size: 16px; font-weight: bold;")
        layout.addWidget(table_label)
        
        self.table = QtWidgets.QTableWidget(0, 6)
        self.table.setEditTriggers(QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setHorizontalHeaderLabels(["Fecha", "SKU", "Producto", "Cantidad", "Razón", "Usuario"])
        self.table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeMode.Stretch)
        self.table.verticalHeader().setVisible(False)
        self.table.setStyleSheet(f"""
            QTableWidget {{
                background: {c['bg_card']};
                border: 1px solid {c['border']};
                border-radius: 10px;
                gridline-color: transparent;
                font-size: 13px;
            }}
            QHeaderView::section {{
                background: {c['table_header_bg']};
                color: {c['table_header_text']};
                padding: 12px;
                border: none;
                font-weight: bold;
            }}
            QTableWidget::item {{
                padding: 10px;
                border-bottom: 1px solid {c['border']};
            }}
        """)
        
        layout.addWidget(self.table)
        
        self.refresh_data()
    
    def _pick_product(self):
        dialog = ProductSearchDialog(core=self.core, parent=self)
        if dialog.exec() == QtWidgets.QDialog.DialogCode.Accepted and dialog.selected_product:
            self.sku_input.setText(dialog.selected_product.get("sku", ""))
    
    def adjust_stock(self):
        sku = self.sku_input.text().strip()
        qty = float(self.qty_input.value())
        reason = self.reason_combo.currentText()
        
        if not sku or qty == 0:
            QtWidgets.QMessageBox.warning(self, "Datos faltantes", "SKU y cantidad son obligatorios")
            return
        
        product = self.core.get_product_by_sku_or_barcode(sku)
        if not product:
            QtWidgets.QMessageBox.warning(self, "No encontrado", "Producto no existe")
            return
        
        product_id = product["id"]
        old_stock = float(product.get('stock', 0))
        new_stock = old_stock + qty # Calculate new stock based on current stock + adjustment quantity
        
        # Realizar ajuste
        self.core.add_stock(product_id, qty, reason=reason)
        
        # AUDIT LOG - Inventory adjustment
        try:
            self.core.audit.log_inventory_adjustment(
                product_id, sku, old_stock, new_stock, reason
            )
        except Exception as e:
            print(f"Error logging inventory adjustment: {e}")
        
        QtWidgets.QMessageBox.information(self, "Ajuste Aplicado", f"Stock ajustado: {qty:+.2f}")
        
        self.qty_input.setValue(0)
        self.refresh_data()
    
    def refresh_data(self):
        """Cargar ajustes recientes"""
        try:
            movements = self.core.db.execute_query("""
                SELECT il.timestamp, p.sku, p.name, il.qty_change, il.reason, il.user_id
                FROM inventory_log il
                JOIN products p ON il.product_id = p.id
                WHERE il.reason NOT IN ('venta', 'sale')
                ORDER BY il.timestamp DESC
                LIMIT 50
            """)
            
            self.table.setRowCount(len(movements))
            for row_idx, mov in enumerate(movements):
                mov_dict = dict(mov)
                self.table.setItem(row_idx, 0, QtWidgets.QTableWidgetItem(mov_dict.get('timestamp', '')[:19]))
                self.table.setItem(row_idx, 1, QtWidgets.QTableWidgetItem(mov_dict.get('sku', '')))
                self.table.setItem(row_idx, 2, QtWidgets.QTableWidgetItem(mov_dict.get('name', '')))
                self.table.setItem(row_idx, 3, QtWidgets.QTableWidgetItem(f"{mov_dict.get('qty_change', 0):+.2f}"))
                self.table.setItem(row_idx, 4, QtWidgets.QTableWidgetItem(mov_dict.get('reason', '')))
                self.table.setItem(row_idx, 5, QtWidgets.QTableWidgetItem(str(mov_dict.get('user_id', ''))))
        except Exception as e:
            print(f"Error loading adjustments: {e}")
    
    def update_theme(self):
        """Actualizar tema en esta pestaña"""
        # Re-construir UI con nuevos colores
        # Método simplificado: limpiar layout y reconstruir
        # O mejor: actualizar estilos de componentes existentes
        theme = (self.core.get_app_config() or {}).get("theme", "Light")
        c = theme_manager.get_colors(theme)
        
        # Actualizar inputs
        input_style = f"""
            QLineEdit {{
                background: {c['input_bg']};
                border: 1px solid {c['input_border']};
                border-radius: 6px;
                padding: 0 10px;
                font-size: 13px;
                color: {c['text_primary']};
            }}
            QLineEdit:focus {{
                border: 2px solid {c['input_focus']};
                background: {c['bg_card']};
            }}
        """
        self.sku_input.setStyleSheet(input_style)
        
        # SpinBox
        spin_style = f"""
            QSpinBox {{
                background: {c['input_bg']};
                border: 1px solid {c['input_border']};
                border-radius: 6px;
                padding: 0 10px;
                font-size: 13px;
                color: {c['text_primary']};
            }}
            QSpinBox:focus {{
                border: 2px solid {c['input_focus']};
                background: {c['bg_card']};
            }}
        """
        self.qty_input.setStyleSheet(spin_style)
        
        # Combo
        combo_style = f"""
            QComboBox {{
                background: {c['input_bg']};
                border: 1px solid {c['input_border']};
                border-radius: 6px;
                padding: 0 10px;
                font-size: 13px;
                color: {c['text_primary']};
            }}
            QComboBox:focus {{
                border: 2px solid {c['input_focus']};
                background: {c['bg_card']};
            }}
        """
        self.reason_combo.setStyleSheet(combo_style)
        
        # Tabla
        self.table.setStyleSheet(f"""
            QTableWidget {{
                background: {c['bg_card']};
                border: 1px solid {c['border']};
                border-radius: 10px;
                gridline-color: transparent;
                font-size: 13px;
            }}
            QHeaderView::section {{
                background: {c['table_header_bg']};
                color: {c['table_header_text']};
                padding: 12px;
                border: none;
                font-weight: bold;
            }}
            QTableWidget::item {{
                padding: 10px;
                border-bottom: 1px solid {c['border']};
                color: {c['table_text']};
            }}
        """)

# ============================================================================
# PESTAÑA 3: MOVIMIENTOS
# ============================================================================

class MovimientosTab(QtWidgets.QWidget):
    """Pestaña de historial de movimientos"""
    
    def __init__(self, core: POSCore):
        super().__init__()
        self.core = core
        self._build_ui()
        self.refresh_data()  # Load data on init
    
    def _build_ui(self):
        theme = (self.core.get_app_config() or {}).get("theme", "Light")
        c = theme_manager.get_colors(theme)
        
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(20)
        
        # Filtros
        filters_frame = QtWidgets.QFrame()
        filters_frame.setStyleSheet(f"""
            QFrame {{
                background: {c['bg_card']};
                border: 1px solid {c['border']};
                border-radius: 10px;
            }}
        """)
        filters_layout = QtWidgets.QHBoxLayout(filters_frame)
        filters_layout.setContentsMargins(15, 15, 15, 15)
        
        # Fecha Desde
        from_label = QtWidgets.QLabel("Desde:")
        from_label.setStyleSheet(f"color: {c['text_secondary']}; font-weight: bold;")
        self.date_from = QtWidgets.QDateEdit()
        self.date_from.setDate(QtCore.QDate.currentDate().addDays(-30))
        self.date_from.setCalendarPopup(True)
        self.date_from.setFixedHeight(35)
        
        # Fecha Hasta
        to_label = QtWidgets.QLabel("Hasta:")
        to_label.setStyleSheet(f"color: {c['text_secondary']}; font-weight: bold;")
        self.date_to = QtWidgets.QDateEdit()
        self.date_to.setDate(QtCore.QDate.currentDate())
        self.date_to.setCalendarPopup(True)
        self.date_to.setFixedHeight(35)
        
        # Botón Filtrar
        self.filter_btn = QtWidgets.QPushButton("🔍 Filtrar")
        self.filter_btn.clicked.connect(self.refresh_data)
        self.filter_btn.setFixedHeight(35)
        self.filter_btn.setCursor(QtCore.Qt.CursorShape.PointingHandCursor)
        self.filter_btn.setStyleSheet(f"""
            QPushButton {{
                background: {c['btn_primary']};
                color: white;
                border: none;
                border-radius: 6px;
                padding: 0 20px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background: #2980b9;
            }}
        """)
        
        # Botón Exportar
        self.export_btn = QtWidgets.QPushButton("📊 Exportar Excel")
        self.export_btn.clicked.connect(self._export_movements)
        self.export_btn.setFixedHeight(35)
        self.export_btn.setStyleSheet(f"""
            QPushButton {{
                background: {c['btn_success']};
                color: white;
                border: none;
                border-radius: 6px;
                padding: 0 20px;
                font-weight: bold;
            }}
        """)
        
        filters_layout.addWidget(from_label)
        filters_layout.addWidget(self.date_from)
        filters_layout.addSpacing(10)
        filters_layout.addWidget(to_label)
        filters_layout.addWidget(self.date_to)
        filters_layout.addSpacing(10)
        filters_layout.addWidget(self.filter_btn)
        filters_layout.addStretch()
        filters_layout.addWidget(self.export_btn)
        
        layout.addWidget(filters_frame)
        
        # Tabla
        self.table = QtWidgets.QTableWidget(0, 7)
        self.table.setEditTriggers(QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setHorizontalHeaderLabels(["Fecha", "SKU", "Producto", "Tipo", "Cantidad", "Razón", "Usuario"])
        self.table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeMode.Stretch)
        self.table.verticalHeader().setVisible(False)
        self.table.setStyleSheet(f"""
            QTableWidget {{
                background: {c['bg_card']};
                border: 1px solid {c['border']};
                border-radius: 10px;
                gridline-color: transparent;
                font-size: 13px;
            }}
            QHeaderView::section {{
                background: {c['table_header_bg']};
                color: {c['table_header_text']};
                padding: 12px;
                border: none;
                font-weight: bold;
            }}
            QTableWidget::item {{
                padding: 10px;
                border-bottom: 1px solid {c['border']};
            }}
        """)
        
        layout.addWidget(self.table)
    
    def refresh_data(self):
        """Cargar movimientos filtrados"""
        try:
            date_from = self.date_from.date().toString("yyyy-MM-dd")
            date_to = self.date_to.date().toString("yyyy-MM-dd")
            
            # Use CAST for PostgreSQL compatibility (works in both SQLite and PostgreSQL)
            movements = self.core.db.execute_query("""
                SELECT il.timestamp, p.sku, p.name, il.qty_change, il.reason, il.user_id
                FROM inventory_log il
                JOIN products p ON il.product_id = p.id
                WHERE CAST(il.timestamp AS DATE) BETWEEN %s AND %s
                ORDER BY il.timestamp DESC
                LIMIT 1000
            """, (date_from, date_to))
            
            self.table.setRowCount(len(movements))
            for row_idx, mov in enumerate(movements):
                mov_dict = dict(mov)
                self.table.setItem(row_idx, 0, QtWidgets.QTableWidgetItem(mov_dict.get('timestamp', '')[:19]))
                self.table.setItem(row_idx, 1, QtWidgets.QTableWidgetItem(mov_dict.get('sku', '')))
                self.table.setItem(row_idx, 2, QtWidgets.QTableWidgetItem(mov_dict.get('name', '')))
                tipo = 'Entrada' if mov_dict.get('qty_change', 0) > 0 else 'Salida'
                self.table.setItem(row_idx, 3, QtWidgets.QTableWidgetItem(tipo))
                self.table.setItem(row_idx, 4, QtWidgets.QTableWidgetItem(f"{mov_dict.get('qty_change', 0):+.2f}"))
                self.table.setItem(row_idx, 5, QtWidgets.QTableWidgetItem(mov_dict.get('reason', '')))
                self.table.setItem(row_idx, 6, QtWidgets.QTableWidgetItem(str(mov_dict.get('user_id', ''))))
        except Exception as e:
            print(f"Error loading movements: {e}")
    
    def _export_movements(self):
        """Exportar movimientos de inventario a Excel/CSV"""
        from datetime import datetime
        
        if self.table.rowCount() == 0:
            QtWidgets.QMessageBox.warning(self, "Sin datos", "No hay movimientos para exportar")
            return
        
        # Pedir ubicación de guardado
        default_name = f"movimientos_inventario_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        file_path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "Guardar Exportación",
            default_name,
            "Excel (*.xlsx);;CSV (*.csv)"
        )
        
        if not file_path:
            return
        
        try:
            # Obtener datos de la tabla
            headers = []
            for col in range(self.table.columnCount()):
                headers.append(self.table.horizontalHeaderItem(col).text())
            
            rows = []
            for row in range(self.table.rowCount()):
                row_data = []
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    row_data.append(item.text() if item else "")
                rows.append(row_data)
            
            if file_path.endswith('.xlsx'):
                # Exportar a Excel usando openpyxl
                try:
                    from openpyxl import Workbook
                    from openpyxl.styles import Alignment, Font, PatternFill
                    
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Movimientos Inventario"
                    
                    # Encabezados con estilo
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    
                    for col_idx, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col_idx, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")
                    
                    # Datos
                    for row_idx, row_data in enumerate(rows, 2):
                        for col_idx, value in enumerate(row_data, 1):
                            ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    # Ajustar anchos de columna
                    for col_idx, header in enumerate(headers, 1):
                        ws.column_dimensions[chr(64 + col_idx)].width = max(len(header) + 2, 15)
                    
                    wb.save(file_path)
                    
                except ImportError:
                    # Fallback a CSV si no hay openpyxl
                    file_path = file_path.replace('.xlsx', '.csv')
                    self._export_to_csv(file_path, headers, rows)
            else:
                # Exportar a CSV
                self._export_to_csv(file_path, headers, rows)
            
            QtWidgets.QMessageBox.information(
                self, 
                "Exportación Exitosa", 
                f"✅ Se exportaron {len(rows)} movimientos a:\n{file_path}"
            )
            
        except Exception as e:
            QtWidgets.QMessageBox.critical(
                self, 
                "Error de Exportación", 
                f"No se pudo exportar los datos:\n{str(e)}"
            )
    
    def _export_to_csv(self, file_path: str, headers: list, rows: list):
        """Exportar a CSV como fallback"""
        import csv
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(rows)
    
    def update_theme(self):
        theme = (self.core.get_app_config() or {}).get("theme", "Light")
        c = theme_manager.get_colors(theme)
        
        # Update frame styles by finding QFrame children
        # This is a bit broad, but if we assume the first frame is the filter frame...
        frames = self.findChildren(QtWidgets.QFrame)
        if frames:
            frames[0].setStyleSheet(f"""
                QFrame {{
                    background: {c['bg_card']};
                    border: 1px solid {c['border']};
                    border-radius: 10px;
                }}
            """)

        # Botones
        for btn in self.findChildren(QtWidgets.QPushButton):
             if "Exportar" in btn.text():
                 btn.setStyleSheet(f"""
                    QPushButton {{
                        background: {c['btn_success']};
                        color: white;
                        border: none;
                        border-radius: 6px;
                        padding: 0 20px;
                        font-weight: bold;
                    }}
                """)
             elif "Filtrar" in btn.text():
                 btn.setStyleSheet(f"""
                    QPushButton {{
                        background: {c['btn_primary']};
                        color: white;
                        border: none;
                        border-radius: 6px;
                        padding: 0 20px;
                        font-weight: bold;
                    }}
                """)
        
        # Labels
        for lbl in self.findChildren(QtWidgets.QLabel):
            lbl.setStyleSheet(f"color: {c['text_secondary']}; font-weight: bold;")
            
        # Tabla
        self.table.setStyleSheet(f"""
            QTableWidget {{
                background: {c['bg_card']};
                border: 1px solid {c['border']};
                border-radius: 10px;
                gridline-color: transparent;
                font-size: 13px;
            }}
            QHeaderView::section {{
                background: {c['table_header_bg']};
                color: {c['table_header_text']};
                padding: 12px;
                border: none;
                font-weight: bold;
            }}
            QTableWidget::item {{
                padding: 10px;
                border-bottom: 1px solid {c['border']};
                 color: {c['table_text']};
            }}
        """)

# ============================================================================
# PESTAÑA 4: TRANSFERENCIAS
# ============================================================================

class TransferenciasTab(QtWidgets.QWidget):
    """Pestaña de transferencias entre sucursales"""
    
    def __init__(self, core: POSCore):
        super().__init__()
        self.core = core
        self._build_ui()
    
    def _build_ui(self):
        theme = (self.core.get_app_config() or {}).get("theme", "Light")
        c = theme_manager.get_colors(theme)
        
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(20)
        
        # Botón Nueva Transferencia
        btn_layout = QtWidgets.QHBoxLayout()
        new_btn = QtWidgets.QPushButton("+ Nueva Transferencia")
        new_btn.clicked.connect(self._create_transfer)
        new_btn.setFixedHeight(40)
        new_btn.setStyleSheet(f"""
            QPushButton {{
                background: {c['btn_success']};
                color: white;
                border: none;
                border-radius: 6px;
                padding: 0 20px;
                font-weight: bold;
            }}
        """)
        btn_layout.addWidget(new_btn)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        
        # Tabs para Pendientes / Historial
        self.sub_tabs = QtWidgets.QTabWidget()
        self.sub_tabs.setStyleSheet(f"""
            QTabWidget::pane {{
                border: 1px solid {c['border']};
                border-radius: 8px;
            }}
            QTabBar::tab {{
                background: {c['bg_card']};
                color: {c['text_secondary']};
                padding: 10px 15px;
                border: 1px solid {c['border']};
                border-bottom: none;
            }}
            QTabBar::tab:selected {{
                background: {c['bg_main']};
                color: {c['text_primary']};
            }}
        """)
        
        # Tab Pendientes
        pending_widget = QtWidgets.QWidget()
        pending_layout = QtWidgets.QVBoxLayout(pending_widget)
        
        self.pending_table = QtWidgets.QTableWidget(0, 6)
        self.pending_table.setEditTriggers(QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)
        self.pending_table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows)
        self.pending_table.setHorizontalHeaderLabels(["ID", "De", "Para", "Productos", "Creado", "Acciones"])
        self.pending_table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeMode.Stretch)
        self.pending_table.verticalHeader().setVisible(False)
        self.pending_table.setStyleSheet(f"""
            QTableWidget {{
                background: {c['bg_card']};
                border: none;
                gridline-color: transparent;
                font-size: 13px;
            }}
            QHeaderView::section {{
                background: {c['table_header_bg']};
                color: {c['table_header_text']};
                padding: 12px;
                border: none;
                font-weight: bold;
            }}
            QTableWidget::item {{
                padding: 10px;
                border-bottom: 1px solid {c['border']};
            }}
        """)
        pending_layout.addWidget(self.pending_table)
        
        # Tab Historial
        history_widget = QtWidgets.QWidget()
        history_layout = QtWidgets.QVBoxLayout(history_widget)
        
        self.history_table = QtWidgets.QTableWidget(0, 7)
        self.history_table.setEditTriggers(QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)
        self.history_table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows)
        self.history_table.setHorizontalHeaderLabels(["ID", "De", "Para", "Productos", "Estado", "Creado", "Completado"])
        self.history_table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeMode.Stretch)
        self.history_table.verticalHeader().setVisible(False)
        self.history_table.setStyleSheet(self.pending_table.styleSheet())
        history_layout.addWidget(self.history_table)
        
        self.sub_tabs.addTab(pending_widget, "Pendientes")
        self.sub_tabs.addTab(history_widget, "Historial")
        
        layout.addWidget(self.sub_tabs)
        
        self.refresh_data()
    
    def _create_transfer(self):
        QtWidgets.QMessageBox.information(self, "Nueva Transferencia", 
            "Funcionalidad de crear transferencia:\n\n" +
            "1. Seleccionar sucursal origen\n" +
            "2. Seleccionar sucursal destino\n" +
            "3. Agregar productos y cantidades\n" +
            "4. Crear solicitud\n\n" +
            "Esta función requiere configuración de sucursales.")
    
    def refresh_data(self):
        """Cargar transferencias"""
        try:
            # Pendientes
            pending = self.core.db.execute_query("""
                SELECT id, from_branch_id, to_branch_id, created_at
                FROM inventory_transfers
                WHERE status = 'pending'
                ORDER BY created_at DESC
            """)
            
            self.pending_table.setRowCount(len(pending))
            for row_idx, transfer in enumerate(pending):
                t = dict(transfer)
                transfer_id = t.get('id')

                # FIX 2026-01-30: Contar items del transfer
                item_count = 0
                try:
                    count_result = self.core.db.execute_query(
                        "SELECT COUNT(*) as cnt FROM inventory_transfer_items WHERE transfer_id = %s",
                        (transfer_id,)
                    )
                    # FIX 2026-02-01: Validar count_result con len() antes de acceder a [0]
                    if count_result and len(count_result) > 0 and count_result[0]:
                        item_count = count_result[0].get('cnt', 0) or 0
                except Exception:
                    pass

                self.pending_table.setItem(row_idx, 0, QtWidgets.QTableWidgetItem(str(transfer_id)))
                self.pending_table.setItem(row_idx, 1, QtWidgets.QTableWidgetItem(f"Sucursal {t.get('from_branch_id', 'N/A')}"))
                self.pending_table.setItem(row_idx, 2, QtWidgets.QTableWidgetItem(f"Sucursal {t.get('to_branch_id', 'N/A')}"))
                self.pending_table.setItem(row_idx, 3, QtWidgets.QTableWidgetItem(str(item_count)))
                self.pending_table.setItem(row_idx, 4, QtWidgets.QTableWidgetItem(t.get('created_at', '')[:19]))
                self.pending_table.setItem(row_idx, 5, QtWidgets.QTableWidgetItem("Aprobar/Rechazar"))
            
            # Historial
            history = self.core.db.execute_query("""
                SELECT id, from_branch_id, to_branch_id, status, created_at, completed_at
                FROM inventory_transfers
                WHERE status IN ('approved', 'rejected', 'completed')
                ORDER BY created_at DESC
                LIMIT 50
            """)
            
            self.history_table.setRowCount(len(history))
            for row_idx, transfer in enumerate(history):
                t = dict(transfer)
                self.history_table.setItem(row_idx, 0, QtWidgets.QTableWidgetItem(str(t.get('id'))))
                self.history_table.setItem(row_idx, 1, QtWidgets.QTableWidgetItem(f"Sucursal {t.get('from_branch_id', 'N/A')}"))
                self.history_table.setItem(row_idx, 2, QtWidgets.QTableWidgetItem(f"Sucursal {t.get('to_branch_id', 'N/A')}"))
                self.history_table.setItem(row_idx, 3, QtWidgets.QTableWidgetItem("0"))
                self.history_table.setItem(row_idx, 4, QtWidgets.QTableWidgetItem(t.get('status', '').upper()))
                self.history_table.setItem(row_idx, 5, QtWidgets.QTableWidgetItem(t.get('created_at', '')[:19]))
                self.history_table.setItem(row_idx, 6, QtWidgets.QTableWidgetItem(t.get('completed_at', '')[:19] if t.get('completed_at') else '-'))
        except Exception as e:
            print(f"Error loading transfers: {e}")
    
    def update_theme(self):
        pass

# ============================================================================
# PESTAÑA 5: ALERTAS
# ============================================================================

class AlertasTab(QtWidgets.QWidget):
    """Pestaña de alertas de inventario"""
    
    def __init__(self, core: POSCore):
        super().__init__()
        self.core = core
        self._build_ui()
        self.refresh_data()  # Load data on init
    
    def _build_ui(self):
        theme = (self.core.get_app_config() or {}).get("theme", "Light")
        c = theme_manager.get_colors(theme)
        
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(20)
        
        # Cards de alertas
        cards_layout = QtWidgets.QHBoxLayout()
        cards_layout.setSpacing(15)
        
        self.card_low_stock = self._create_card("Stock Bajo", "0", "#FF9800", c)
        self.card_no_movement = self._create_card("Sin Movimiento", "0", "#9C27B0", c)
        self.card_overstock = self._create_card("Sobre Stock", "0", "#2196F3", c)
        
        cards_layout.addWidget(self.card_low_stock)
        cards_layout.addWidget(self.card_no_movement)
        cards_layout.addWidget(self.card_overstock)
        
        layout.addLayout(cards_layout)
        
        # Tabla de alertas
        table_label = QtWidgets.QLabel("Productos que Requieren Atención")
        table_label.setStyleSheet(f"color: {c['text_primary']}; font-size: 16px; font-weight: bold;")
        layout.addWidget(table_label)
        
        self.table = QtWidgets.QTableWidget(0, 6)
        self.table.setEditTriggers(QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setHorizontalHeaderLabels(["SKU", "Producto", "Stock", "Mín", "Máx", "Alerta"])
        self.table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeMode.Stretch)
        self.table.verticalHeader().setVisible(False)
        self.table.setStyleSheet(f"""
            QTableWidget {{
                background: {c['bg_card']};
                border: 1px solid {c['border']};
                border-radius: 10px;
                gridline-color: transparent;
                font-size: 13px;
            }}
            QHeaderView::section {{
                background: {c['table_header_bg']};
                color: {c['table_header_text']};
                padding: 12px;
                border: none;
                font-weight: bold;
            }}
            QTableWidget::item {{
                padding: 10px;
                border-bottom: 1px solid {c['border']};
            }}
        """)
        
        layout.addWidget(self.table)
        
        # Botón Actualizar
        btn_layout = QtWidgets.QHBoxLayout()
        self.btn_refresh = QtWidgets.QPushButton("🔄 Actualizar Alertas")
        self.btn_refresh.setFixedHeight(40)
        self.btn_refresh.clicked.connect(self.refresh_data)
        self.btn_refresh.setCursor(QtCore.Qt.CursorShape.PointingHandCursor)
        self.btn_refresh.setStyleSheet(f"""
            QPushButton {{
                background: {c['btn_primary']};
                color: white;
                border: none;
                border-radius: 6px;
                padding: 0 20px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background: #2980b9;
            }}
        """)
        btn_layout.addWidget(self.btn_refresh)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
    
    def _create_card(self, title: str, value: str, color: str, c: dict) -> QtWidgets.QFrame:
        card = QtWidgets.QFrame()
        card.setStyleSheet(f"""
            QFrame {{
                background-color: {c['bg_card']};
                border: 1px solid {c['border']};
                border-radius: 10px;
                border-left: 5px solid {color};
            }}
        """)
        
        layout = QtWidgets.QVBoxLayout(card)
        layout.setSpacing(5)
        
        title_label = QtWidgets.QLabel(title)
        title_label.setStyleSheet(f"color: {c['text_secondary']}; font-size: 13px; font-weight: 600; border: none;")
        
        value_label = QtWidgets.QLabel(value)
        value_label.setObjectName("value_label")
        value_label.setStyleSheet(f"color: {color}; font-size: 28px; font-weight: bold; border: none;")
        
        layout.addWidget(title_label)
        layout.addWidget(value_label)
        
        return card
    
    def refresh_data(self):
        """Cargar alertas"""
        try:
            products = self.core.list_products_for_export()
            
            low_stock = []
            overstock = []
            no_movement = []
            
            # Obtener productos con movimiento en últimos 30 días
            from datetime import datetime, timedelta
            cutoff_date = (datetime.now() - timedelta(days=30)).isoformat()
            
            try:
                recent_movements = self.core.db.execute_query("""
                    SELECT DISTINCT product_id FROM inventory_log
                    WHERE timestamp >= %s
                """, (cutoff_date,))
                products_with_movement = {m['product_id'] for m in recent_movements}
            except Exception:
                products_with_movement = set()
            
            for p in products:
                stock = float(p.get('stock', 0) or 0)
                min_stock = float(p.get('min_stock', 0) or 0)
                max_stock = float(p.get('max_stock', 0) or 0)
                product_id = p.get('id')
                
                if stock <= min_stock and min_stock > 0:
                    low_stock.append(p)
                elif max_stock > 0 and stock >= max_stock:
                    overstock.append(p)
                
                # Verificar si tiene movimiento reciente
                if product_id not in products_with_movement and stock > 0:
                    no_movement.append(p)
            
            # Actualizar cards
            self.card_low_stock.findChild(QtWidgets.QLabel, "value_label").setText(str(len(low_stock)))
            self.card_no_movement.findChild(QtWidgets.QLabel, "value_label").setText(str(len(no_movement)))
            self.card_overstock.findChild(QtWidgets.QLabel, "value_label").setText(str(len(overstock)))
            
            # Actualizar tabla - incluir también los sin movimiento
            all_alerts = []
            for p in low_stock:
                all_alerts.append({**p, 'alert_type': 'Stock Bajo'})
            for p in overstock:
                all_alerts.append({**p, 'alert_type': 'Sobre Stock'})
            for p in no_movement[:20]:  # Limitar a 20 para no saturar
                all_alerts.append({**p, 'alert_type': 'Sin Movimiento (30d)'})
            
            self.table.setRowCount(len(all_alerts))
            
            for row_idx, p in enumerate(all_alerts):
                stock = float(p.get('stock', 0) or 0)
                min_stock = float(p.get('min_stock', 0) or 0)
                max_stock = float(p.get('max_stock', 0) or 0)
                
                self.table.setItem(row_idx, 0, QtWidgets.QTableWidgetItem(p.get('sku', '')))
                self.table.setItem(row_idx, 1, QtWidgets.QTableWidgetItem(p.get('name', '')))
                self.table.setItem(row_idx, 2, QtWidgets.QTableWidgetItem(f"{stock:.2f}"))
                self.table.setItem(row_idx, 3, QtWidgets.QTableWidgetItem(f"{min_stock:.2f}"))
                self.table.setItem(row_idx, 4, QtWidgets.QTableWidgetItem(f"{max_stock:.2f}"))
                self.table.setItem(row_idx, 5, QtWidgets.QTableWidgetItem(p.get('alert_type', '')))
        except Exception as e:
            print(f"Error loading alerts: {e}")
    
    def update_theme(self):
        pass

# ============================================================================
# PESTAÑA 6: ANÁLISIS
# ============================================================================

class AnalisisTab(QtWidgets.QWidget):
    """Pestaña de análisis y reportes"""
    
    def __init__(self, core: POSCore):
        super().__init__()
        self.core = core
        self._build_ui()
    
    def _build_ui(self):
        theme = (self.core.get_app_config() or {}).get("theme", "Light")
        c = theme_manager.get_colors(theme)
        
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(20)
        
        # Cards de métricas
        cards_layout = QtWidgets.QHBoxLayout()
        cards_layout.setSpacing(15)
        
        self.card_total_value = self._create_card("Valor Total", "$0.00", "#4CAF50", c)
        self.card_products = self._create_card("Total Productos", "0", "#2196F3", c)
        self.card_avg_value = self._create_card("Valor Promedio", "$0.00", "#FF9800", c)
        
        cards_layout.addWidget(self.card_total_value)
        cards_layout.addWidget(self.card_products)
        cards_layout.addWidget(self.card_avg_value)
        
        layout.addLayout(cards_layout)
        
        # Tabs para diferentes análisis
        self.sub_tabs = QtWidgets.QTabWidget()
        self.sub_tabs.setStyleSheet(f"""
            QTabWidget::pane {{
                border: 1px solid {c['border']};
                border-radius: 8px;
            }}
            QTabBar::tab {{
                background: {c['bg_card']};
                color: {c['text_secondary']};
                padding: 10px 15px;
                border: 1px solid {c['border']};
                border-bottom: none;
            }}
            QTabBar::tab:selected {{
                background: {c['bg_main']};
                color: {c['text_primary']};
            }}
        """)
        
        # Tab Top Productos
        top_widget = QtWidgets.QWidget()
        top_layout = QtWidgets.QVBoxLayout(top_widget)
        
        self.top_table = QtWidgets.QTableWidget(0, 5)
        self.top_table.setEditTriggers(QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)
        self.top_table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows)
        self.top_table.setHorizontalHeaderLabels(["SKU", "Producto", "Stock", "Valor Unit.", "Valor Total"])
        self.top_table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeMode.Stretch)
        self.top_table.verticalHeader().setVisible(False)
        self.top_table.setStyleSheet(f"""
            QTableWidget {{
                background: {c['bg_card']};
                border: none;
                gridline-color: transparent;
                font-size: 13px;
            }}
            QHeaderView::section {{
                background: {c['table_header_bg']};
                color: {c['table_header_text']};
                padding: 12px;
                border: none;
                font-weight: bold;
            }}
            QTableWidget::item {{
                padding: 10px;
                border-bottom: 1px solid {c['border']};
            }}
        """)
        top_layout.addWidget(self.top_table)
        
        # Tab Baja Rotación
        low_widget = QtWidgets.QWidget()
        low_layout = QtWidgets.QVBoxLayout(low_widget)
        
        self.low_table = QtWidgets.QTableWidget(0, 4)
        self.low_table.setEditTriggers(QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)
        self.low_table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows)
        self.low_table.setHorizontalHeaderLabels(["SKU", "Producto", "Stock", "Días sin Movimiento"])
        self.low_table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeMode.Stretch)
        self.low_table.verticalHeader().setVisible(False)
        self.low_table.setStyleSheet(self.top_table.styleSheet())
        low_layout.addWidget(self.low_table)
        
        self.sub_tabs.addTab(top_widget, "Top 10 por Valor")
        self.sub_tabs.addTab(low_widget, "Baja Rotación")
        
        layout.addWidget(self.sub_tabs)
        
        self.refresh_data()
    
    def _create_card(self, title: str, value: str, color: str, c: dict) -> QtWidgets.QFrame:
        card = QtWidgets.QFrame()
        card.setStyleSheet(f"""
            QFrame {{
                background-color: {c['bg_card']};
                border: 1px solid {c['border']};
                border-radius: 10px;
                border-left: 5px solid {color};
            }}
        """)
        
        layout = QtWidgets.QVBoxLayout(card)
        layout.setSpacing(5)
        
        title_label = QtWidgets.QLabel(title)
        title_label.setStyleSheet(f"color: {c['text_secondary']}; font-size: 13px; font-weight: 600; border: none;")
        
        value_label = QtWidgets.QLabel(value)
        value_label.setObjectName("value_label")
        value_label.setStyleSheet(f"color: {color}; font-size: 28px; font-weight: bold; border: none;")
        
        layout.addWidget(title_label)
        layout.addWidget(value_label)
        
        return card
    
    def refresh_data(self):
        """Cargar análisis"""
        try:
            products = self.core.list_products_for_export()
            
            # Calcular métricas
            total_value = sum(float(p.get('price', 0) or 0) * float(p.get('stock', 0) or 0) for p in products)
            total_products = len(products)
            avg_value = total_value / total_products if total_products > 0 else 0
            
            # Actualizar cards
            if STATE.role == "admin":
                self.card_total_value.findChild(QtWidgets.QLabel, "value_label").setText(f"${total_value:,.2f}")
                self.card_avg_value.findChild(QtWidgets.QLabel, "value_label").setText(f"${avg_value:,.2f}")
            else:
                self.card_total_value.findChild(QtWidgets.QLabel, "value_label").setText("$***")
                self.card_avg_value.findChild(QtWidgets.QLabel, "value_label").setText("$***")
            
            self.card_products.findChild(QtWidgets.QLabel, "value_label").setText(str(total_products))
            
            # Top 10 por valor
            products_with_value = [
                {
                    **p,
                    'total_value': float(p.get('price', 0) or 0) * float(p.get('stock', 0) or 0)
                }
                for p in products
            ]
            top_products = sorted(products_with_value, key=lambda x: x['total_value'], reverse=True)[:10]
            
            self.top_table.setRowCount(len(top_products))
            for row_idx, p in enumerate(top_products):
                self.top_table.setItem(row_idx, 0, QtWidgets.QTableWidgetItem(p.get('sku', '')))
                self.top_table.setItem(row_idx, 1, QtWidgets.QTableWidgetItem(p.get('name', '')))
                self.top_table.setItem(row_idx, 2, QtWidgets.QTableWidgetItem(f"{float(p.get('stock', 0)):.2f}"))
                if STATE.role == "admin":
                    self.top_table.setItem(row_idx, 3, QtWidgets.QTableWidgetItem(f"${float(p.get('price', 0)):.2f}"))
                    self.top_table.setItem(row_idx, 4, QtWidgets.QTableWidgetItem(f"${p['total_value']:.2f}"))
                else:
                    self.top_table.setItem(row_idx, 3, QtWidgets.QTableWidgetItem("$***"))
                    self.top_table.setItem(row_idx, 4, QtWidgets.QTableWidgetItem("$***"))
            
            # Baja rotación - productos sin movimientos recientes
            # Consultar último movimiento para cada producto
            last_movements = {}
            try:
                movements_query = self.core.db.execute_query("""
                    SELECT product_id, MAX(timestamp) as last_movement
                    FROM inventory_log
                    GROUP BY product_id
                """)
                for m in movements_query:
                    md = dict(m)
                    last_movements[md.get('product_id')] = md.get('last_movement')
            except Exception:
                pass
            
            from datetime import datetime
            today = datetime.now()
            
            # Calcular días sin movimiento para cada producto
            products_with_days = []
            for p in products:
                product_id = p.get('id')
                last_mov = last_movements.get(product_id)
                
                if last_mov:
                    try:
                        last_date = datetime.fromisoformat(last_mov.replace('T', ' ')[:19])
                        days = (today - last_date).days
                    except Exception:
                        days = 999  # Si no se puede parsear, considerarlo muy viejo
                else:
                    days = 999  # Sin movimiento registrado
                
                products_with_days.append({**p, 'days_without_movement': days})
            
            # Ordenar por días sin movimiento (descendente) y tomar los primeros 10
            low_rotation = sorted(products_with_days, key=lambda x: x['days_without_movement'], reverse=True)[:10]
            
            self.low_table.setRowCount(len(low_rotation))
            for row_idx, p in enumerate(low_rotation):
                days = p['days_without_movement']
                days_text = f"{days} días" if days < 999 else "Sin registro"
                
                self.low_table.setItem(row_idx, 0, QtWidgets.QTableWidgetItem(p.get('sku', '')))
                self.low_table.setItem(row_idx, 1, QtWidgets.QTableWidgetItem(p.get('name', '')))
                self.low_table.setItem(row_idx, 2, QtWidgets.QTableWidgetItem(f"{float(p.get('stock', 0)):.2f}"))
                self.low_table.setItem(row_idx, 3, QtWidgets.QTableWidgetItem(days_text))
        except Exception as e:
            print(f"Error loading analysis: {e}")
    
    def update_theme(self):
        pass
