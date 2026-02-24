from __future__ import annotations

import logging

from PyQt6 import QtCore, QtGui, QtWidgets

logger = logging.getLogger(__name__)

from app.core import POSCore
from app.utils import ticket_engine  # FIX: Add missing import
from app.utils.export_csv import export_sales_to_csv
from app.utils.path_utils import get_debug_log_path_str, get_debug_log_path, agent_log_enabled
from app.utils.theme_manager import theme_manager


class HistoryTab(QtWidgets.QWidget):
    def __init__(self, core: POSCore, parent: QtWidgets.QWidget | None = None):
        super().__init__(parent)
        self.core = core
        self.selected_sale_id: int | None = None
        self.load_assets()
        self._build_ui()
        self.refresh_sales()

    def load_assets(self):
        self.icons = {}
        try:
            self.icons["search"] = QtGui.QIcon("assets/icon_search.png")
            self.icons["print"] = QtGui.QIcon("assets/icon_products.png")
            self.icons["cancel"] = QtGui.QIcon("assets/icon_exit.png")
            self.icons["invoice"] = QtGui.QIcon("assets/icon_excel.png")
            self.icons["history"] = QtGui.QIcon("assets/icon_shifts.png")
        except Exception as e:
            # FIX 2026-02-01: Usar logger
            logger.debug("Error loading icons: %s", e)

    def reprint_ticket(self) -> None:
        if not self.selected_sale_id:
            QtWidgets.QMessageBox.warning(self, "Aviso", "Seleccione una venta para reimprimir.")
            return
        try:
            ticket_engine.print_ticket(self.core, self.selected_sale_id)
            QtWidgets.QMessageBox.information(self, "Éxito", "Ticket reimpreso correctamente")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Error", f"No se pudo reimprimir: {e}")

    def _build_ui(self) -> None:
        main_layout = QtWidgets.QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # === HEADER ===
        self.header = QtWidgets.QFrame()
        self.header.setFixedHeight(70)
        header_layout = QtWidgets.QHBoxLayout(self.header)
        header_layout.setContentsMargins(20, 0, 20, 0)
        
        if "history" in self.icons:
            icon_lbl = QtWidgets.QLabel()
            icon_lbl.setPixmap(self.icons["history"].pixmap(32, 32))
            header_layout.addWidget(icon_lbl)
            
        self.title_label = QtWidgets.QLabel("HISTORIAL DE VENTAS")
        header_layout.addWidget(self.title_label)
        header_layout.addStretch()
        
        main_layout.addWidget(self.header)

        # === CONTENEDOR PRINCIPAL ===
        self.content_widget = QtWidgets.QWidget()
        content_layout = QtWidgets.QVBoxLayout(self.content_widget)
        content_layout.setContentsMargins(20, 20, 20, 20)
        content_layout.setSpacing(20)

        # === DASHBOARD DE CONTROL ===
        self.dashboard_frame = QtWidgets.QFrame()
        dashboard_layout = QtWidgets.QVBoxLayout(self.dashboard_frame)
        dashboard_layout.setContentsMargins(15, 15, 15, 15)
        dashboard_layout.setSpacing(15)
        
        # --- FILA 1: Búsqueda ---
        row1 = QtWidgets.QHBoxLayout()
        self.search_input = QtWidgets.QLineEdit()
        self.search_input.setPlaceholderText("🔍 Buscar por folio o ticket...")
        self.search_input.setFixedHeight(45)
        self.search_input.textChanged.connect(self.refresh_sales)
        row1.addWidget(self.search_input)
        dashboard_layout.addLayout(row1)
        
        # --- FILA 2: Filtros y Acciones ---
        row2 = QtWidgets.QHBoxLayout()
        row2.setSpacing(15)
        
        # Filtro Fecha Desde
        self.lbl_date_from = QtWidgets.QLabel("📅 Desde:")
        row2.addWidget(self.lbl_date_from)
        
        self.date_from = QtWidgets.QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_from.setDate(QtCore.QDate.currentDate())
        self.date_from.setFixedHeight(40)
        self.date_from.dateChanged.connect(self.refresh_sales)
        row2.addWidget(self.date_from)
        
        # Filtro Fecha Hasta
        self.lbl_date_to = QtWidgets.QLabel("Hasta:")
        row2.addWidget(self.lbl_date_to)
        
        self.date_to = QtWidgets.QDateEdit()
        self.date_to.setCalendarPopup(True)
        self.date_to.setDate(QtCore.QDate.currentDate())
        self.date_to.setFixedHeight(40)
        self.date_to.dateChanged.connect(self.refresh_sales)
        row2.addWidget(self.date_to)
        
        self.today_btn = QtWidgets.QPushButton("Hoy")
        self.today_btn.clicked.connect(self._set_today)
        self.today_btn.setCursor(QtCore.Qt.CursorShape.PointingHandCursor)
        self.today_btn.setFixedHeight(40)
        row2.addWidget(self.today_btn)
        
        self.week_btn = QtWidgets.QPushButton("Semana")
        self.week_btn.clicked.connect(self._set_week)
        self.week_btn.setCursor(QtCore.Qt.CursorShape.PointingHandCursor)
        self.week_btn.setFixedHeight(40)
        row2.addWidget(self.week_btn)
        
        self.month_btn = QtWidgets.QPushButton("Mes")
        self.month_btn.clicked.connect(self._set_month)
        self.month_btn.setCursor(QtCore.Qt.CursorShape.PointingHandCursor)
        self.month_btn.setFixedHeight(40)
        row2.addWidget(self.month_btn)
        
        # Checkbox Crédito
        self.credit_checkbox = QtWidgets.QCheckBox("Solo Crédito")
        self.credit_checkbox.stateChanged.connect(self.refresh_sales)
        row2.addWidget(self.credit_checkbox)
        
        # ComboBox Método de Pago
        self.lbl_payment = QtWidgets.QLabel("💳 Método:")
        row2.addWidget(self.lbl_payment)
        
        self.payment_filter = QtWidgets.QComboBox()
        self.payment_filter.addItems([
            "Todos",
            "Efectivo",
            "Tarjeta", 
            "Crédito",
            "Mixto",
            "Transferencia",
            "USD",
            "Vales",
            "Cheque",
            "Monedero"
        ])
        self.payment_filter.setFixedHeight(40)
        self.payment_filter.setMinimumWidth(120)
        self.payment_filter.currentTextChanged.connect(self.refresh_sales)
        row2.addWidget(self.payment_filter)
        
        row2.addStretch()
        
        # Botones Acción
        self.action_buttons_data = []
        
        def make_btn(text, icon_key, color_key, callback):
            btn = QtWidgets.QPushButton(f" {text}")
            if icon_key in self.icons:
                btn.setIcon(self.icons[icon_key])
                btn.setIconSize(QtCore.QSize(18, 18))
            btn.clicked.connect(callback)
            btn.setFixedHeight(40)
            btn.setCursor(QtCore.Qt.CursorShape.PointingHandCursor)
            self.action_buttons_data.append((btn, color_key))
            return btn
            
        self.btn_reprint = make_btn("Reimprimir", "print", "btn_primary", self.reprint_ticket)
        self.btn_cancel = make_btn("Cancelar", "cancel", "btn_danger", self.cancel_sale)
        self.btn_return = make_btn("Devoluciones", "history", "btn_warning", self.cancel_sale) # Reuses cancel dialog which handles partials
        self.btn_factura = make_btn("Facturar", "invoice", "#9b59b6", self._issue_cfdi)
        self.btn_export = make_btn("📊 Exportar", "invoice", "btn_success", self._export_sales)
        self.btn_import = make_btn("📥 Importar", "history", "#e67e22", self._import_sales)

        row2.addWidget(self.btn_reprint)
        row2.addWidget(self.btn_cancel)
        row2.addWidget(self.btn_return)
        row2.addWidget(self.btn_factura)
        row2.addWidget(self.btn_export)
        row2.addWidget(self.btn_import)
        
        dashboard_layout.addLayout(row2)
        content_layout.addWidget(self.dashboard_frame)
        
        # === CONTENIDO PRINCIPAL (SPLITTER) ===
        self.splitter = QtWidgets.QSplitter(QtCore.Qt.Orientation.Horizontal)
        self.splitter.setHandleWidth(1)
        
        # --- LADO IZQUIERDO: LISTA VENTAS ---
        self.left_card = QtWidgets.QFrame()
        left_layout = QtWidgets.QVBoxLayout(self.left_card)
        left_layout.setContentsMargins(0, 0, 0, 0)
        
        # FIX B4 2026-01-30: Tabla tiene 6 columnas (Folio, Arts, Hora, PC/Origen, Total, CFDI)
        self.sales_table = QtWidgets.QTableWidget(0, 6)
        self.sales_table.setEditTriggers(QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)  # Read-only
        self.sales_table.setHorizontalHeaderLabels(["Folio", "Arts", "Hora", "PC/Origen", "Total", "CFDI"])
        self.sales_table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeMode.Stretch)
        self.sales_table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows)
        self.sales_table.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.SingleSelection)
        self.sales_table.setAlternatingRowColors(False)
        self.sales_table.itemSelectionChanged.connect(self.refresh_items)
        left_layout.addWidget(self.sales_table)
        
        # --- LADO DERECHO: DETALLE ---
        self.right_card = QtWidgets.QFrame()
        right_layout = QtWidgets.QVBoxLayout(self.right_card)
        right_layout.setContentsMargins(0, 0, 0, 0)
        
        # Header Detalle
        self.detail_header = QtWidgets.QLabel("📦 Detalle del Ticket")
        right_layout.addWidget(self.detail_header)
        
        self.items_table = QtWidgets.QTableWidget(0, 4)
        self.items_table.setEditTriggers(QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)  # Read-only
        self.items_table.setHorizontalHeaderLabels(["Cant.", "Descripción", "Importe", "Total"])
        self.items_table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeMode.Stretch)
        self.items_table.verticalHeader().setVisible(False)
        right_layout.addWidget(self.items_table)
        
        # Info Pago
        self.payment_frame = QtWidgets.QFrame()
        payment_layout = QtWidgets.QVBoxLayout(self.payment_frame)
        payment_layout.setContentsMargins(15, 15, 15, 15)
        
        self.payment_label = QtWidgets.QLabel("Pago con: -")
        self.total_label = QtWidgets.QLabel("Total: $0.00")
        self.total_label.setAlignment(QtCore.Qt.AlignmentFlag.AlignRight)
        
        payment_layout.addWidget(self.payment_label)
        payment_layout.addWidget(self.total_label)
        
        right_layout.addWidget(self.payment_frame)
        
        self.splitter.addWidget(self.left_card)
        self.splitter.addWidget(self.right_card)
        self.splitter.setStretchFactor(0, 60)
        self.splitter.setStretchFactor(1, 40)
        
        content_layout.addWidget(self.splitter)
        
        # === BARRA DE TOTALES ===
        self.totals_frame = QtWidgets.QFrame()
        totals_layout = QtWidgets.QHBoxLayout(self.totals_frame)
        totals_layout.setContentsMargins(15, 10, 15, 10)
        
        self.total_sales_label = QtWidgets.QLabel("💰 Total Ventas: $0.00")
        self.count_sales_label = QtWidgets.QLabel("📊 Cantidad: 0")
        self.cancelled_label = QtWidgets.QLabel("🚫 Canceladas: 0")
        
        totals_layout.addWidget(self.total_sales_label)
        totals_layout.addWidget(self.count_sales_label)
        totals_layout.addWidget(self.cancelled_label)
        totals_layout.addStretch()
        
        content_layout.addWidget(self.totals_frame)
        main_layout.addWidget(self.content_widget)

        self.update_theme()
    
    def _set_today(self):
        """Set date range to today"""
        today = QtCore.QDate.currentDate()
        self.date_from.setDate(today)
        self.date_to.setDate(today)
    
    def _set_week(self):
        """Set date range to last 7 days"""
        today = QtCore.QDate.currentDate()
        self.date_from.setDate(today.addDays(-6))
        self.date_to.setDate(today)
    
    def _set_month(self):
        """Set date range to current month"""
        today = QtCore.QDate.currentDate()
        self.date_from.setDate(QtCore.QDate(today.year(), today.month(), 1))
        self.date_to.setDate(today)
    
    def _export_sales(self):
        """Export sales - show options dialog"""
        from PyQt6.QtWidgets import QFileDialog, QInputDialog

        # Ask export type
        options = [
            "📊 Reporte Simple (sin detalle de productos)",
            "📋 Reporte DETALLADO (incluye cada producto comprado)",
            "📈 Reporte Excel con Dashboard Ejecutivo",
            "🧾 Exportar venta seleccionada (Individual)"
        ]
        
        option, ok = QInputDialog.getItem(
            self,
            "Tipo de Exportación",
            "¿Qué tipo de reporte deseas generar?",
            options,
            1,  # Default: Detallado
            False
        )
        
        if not ok:
            return
        
        if "Individual" in option:
            self._export_single_sale()
        elif "DETALLADO" in option:
            self._export_all_sales_detailed()
        elif "Dashboard" in option or "Excel" in option:
            self._export_all_sales_excel()
        else:
            self._export_all_sales()
    
    def _export_single_sale(self):
        """Export single selected sale with full details"""
        from datetime import datetime

        from PyQt6.QtWidgets import QFileDialog
        
        if not self.selected_sale_id:
            QtWidgets.QMessageBox.warning(self, "Selecciona una venta", "Primero selecciona una venta de la lista")
            return
        
        sale = self.core.get_sale(self.selected_sale_id)
        if not sale:
            return
        
        items = self.core.get_sale_items(self.selected_sale_id)
        
        folio = sale.get('folio_visible') or f"#{sale['id']}"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Exportar Venta Individual",
            f"Venta_{folio.replace('#', '')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
            "Excel (*.xlsx);;CSV (*.csv)"
        )
        
        if not file_path:
            return
        
        try:
            if file_path.endswith('.xlsx'):
                self._export_sale_to_excel(sale, items, file_path)
            else:
                self._export_sale_to_csv(sale, items, file_path)
            
            QtWidgets.QMessageBox.information(
                self,
                "Exportación Exitosa",
                f"✅ Venta {folio} exportada a:\n{file_path}"
            )
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Error", f"Error al exportar: {e}")
    
    def _export_sale_to_excel(self, sale, items, file_path):
        """Export single sale to Excel with styling"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        except ImportError:
            # Fallback to CSV
            file_path = file_path.replace('.xlsx', '.csv')
            self._export_sale_to_csv(sale, items, file_path)
            return
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Detalle Venta"
        
        # Styles
        header_font = Font(bold=True, size=14)
        label_font = Font(bold=True)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Header
        ws['A1'] = "DETALLE DE VENTA"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        # Sale info
        folio = sale.get('folio_visible') or f"#{sale['id']}"
        info = [
            ("Folio:", folio),
            ("Serie:", sale.get('serie', 'A')),
            ("Fecha:", str(sale.get('timestamp', ''))[:19]),
            ("Método de Pago:", sale.get('payment_method', '')),
            ("Subtotal:", f"${float(sale.get('subtotal', 0)):.2f}"),
            ("IVA:", f"${float(sale.get('tax', 0)):.2f}"),
            ("Total:", f"${float(sale.get('total', 0)):.2f}"),
        ]
        
        row = 3
        for label, value in info:
            ws.cell(row=row, column=1, value=label).font = label_font
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # Products header
        row += 1
        ws.cell(row=row, column=1, value="PRODUCTOS").font = header_font
        row += 1
        
        headers = ["Cantidad", "Producto", "Precio Unit.", "Total", "Código SAT"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Products data
        row += 1
        for item in items:
            ws.cell(row=row, column=1, value=float(item.get('qty', 0)))
            ws.cell(row=row, column=2, value=item.get('name', ''))
            ws.cell(row=row, column=3, value=f"${float(item.get('price', 0)):.2f}")
            ws.cell(row=row, column=4, value=f"${float(item.get('total', 0)):.2f}")
            ws.cell(row=row, column=5, value=item.get('sat_clave_prod_serv', ''))
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        
        wb.save(file_path)
    
    def _export_sale_to_csv(self, sale, items, file_path):
        """Export single sale to CSV"""
        import csv
        
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Header info
            folio = sale.get('folio_visible') or f"#{sale['id']}"
            writer.writerow(["DETALLE DE VENTA"])
            writer.writerow(["Folio", folio])
            writer.writerow(["Serie", sale.get('serie', 'A')])
            writer.writerow(["Fecha", str(sale.get('timestamp', ''))[:19]])
            writer.writerow(["Método", sale.get('payment_method', '')])
            writer.writerow(["Subtotal", f"${float(sale.get('subtotal', 0)):.2f}"])
            writer.writerow(["IVA", f"${float(sale.get('tax', 0)):.2f}"])
            writer.writerow(["Total", f"${float(sale.get('total', 0)):.2f}"])
            writer.writerow([])
            
            # Products
            writer.writerow(["PRODUCTOS"])
            writer.writerow(["Cantidad", "Producto", "Precio Unit.", "Total", "Código SAT"])
            for item in items:
                writer.writerow([
                    float(item.get('qty', 0)),
                    item.get('name', ''),
                    f"${float(item.get('price', 0)):.2f}",
                    f"${float(item.get('total', 0)):.2f}",
                    item.get('sat_clave_prod_serv', '')
                ])
    
    def _export_all_sales(self):
        """Export all filtered sales to Excel/CSV"""
        from datetime import datetime

        from PyQt6.QtWidgets import QFileDialog
        
        date_from = self.date_from.date().toString("yyyy-MM-dd")
        date_to = self.date_to.date().toString("yyyy-MM-dd")
        
        # Get filter
        payment_filter_text = self.payment_filter.currentText() if hasattr(self, 'payment_filter') else "Todos"
        payment_method_map = {
            "Todos": None, "Efectivo": "cash", "Tarjeta": "card", "Crédito": "credit",
            "Mixto": "mixed", "Transferencia": "transfer", "USD": "usd",
            "Vales": "voucher", "Cheque": "check", "Monedero": "wallet"
        }
        selected_payment = payment_method_map.get(payment_filter_text)
        
        # Get data
        sales_data = self.core.get_sales_by_range(date_from, date_to)
        
        # Protección contra None
        if not sales_data:
            QtWidgets.QMessageBox.warning(self, "Sin datos", "No hay ventas en el período seleccionado")
            return
        
        if selected_payment:
            sales_data = [s for s in sales_data if s.get("payment_method") == selected_payment]
        
        if not sales_data:
            QtWidgets.QMessageBox.warning(self, "Sin datos", "No hay ventas con los filtros seleccionados")
            return
        
        suffix = f"_{payment_filter_text}" if selected_payment else ""
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Exportar Reporte de Ventas",
            f"Reporte_Ventas_{date_from}_a_{date_to}{suffix}.xlsx",
            "Excel (*.xlsx);;CSV (*.csv)"
        )
        
        if not file_path:
            return
        
        try:
            if file_path.endswith('.xlsx'):
                self._export_global_to_excel(sales_data, file_path, date_from, date_to)
            else:
                from app.utils.export_csv import export_sales_to_csv
                export_sales_to_csv(sales_data, file_path)
            
            QtWidgets.QMessageBox.information(
                self,
                "Exportación Exitosa",
                f"✅ {len(sales_data)} ventas exportadas a:\n{file_path}"
            )
        except Exception as e:
            import traceback
            traceback.print_exc()  # Imprimir en consola para debug
            QtWidgets.QMessageBox.critical(self, "Error", f"Error al exportar: {e}")
    
    def _export_all_sales_detailed(self):
        """Export all sales with FULL PRODUCT DETAILS to CSV"""
        from datetime import datetime

        from PyQt6.QtWidgets import QFileDialog
        
        date_from = self.date_from.date().toString("yyyy-MM-dd")
        date_to = self.date_to.date().toString("yyyy-MM-dd")
        
        # Get data
        sales_data = self.core.get_sales_by_range(date_from, date_to)
        
        if not sales_data:
            QtWidgets.QMessageBox.warning(self, "Sin datos", "No hay ventas en el período seleccionado")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Exportar Reporte DETALLADO de Ventas",
            f"Ventas_Detalladas_{date_from}_a_{date_to}.csv",
            "CSV (*.csv)"
        )
        
        if not file_path:
            return
        
        try:
            from app.utils.export_csv import export_sales_detailed_to_csv

            # FIX 2026-02-04: Optimized - pre-load customer data with JOIN instead of N+1 queries
            customer_ids = [s['customer_id'] for s in sales_data if s.get('customer_id')]
            customers_map = {}
            if customer_ids and self.core.db:
                customers_result = self.core.db.execute_query(
                    "SELECT id, name, rfc FROM customers WHERE id IN %s",
                    (tuple(customer_ids),)
                )
                customers_map = {c['id']: c for c in customers_result}

            for sale in sales_data:
                if sale.get('customer_id') and sale['customer_id'] in customers_map:
                    cust = customers_map[sale['customer_id']]
                    sale['customer_name'] = cust.get('name', '')
                    sale['customer_rfc'] = cust.get('rfc', '')

            # FIX 2026-02-04: Optimized - pre-load item counts with GROUP BY instead of N+1 queries
            sale_ids = [s['id'] for s in sales_data]
            items_count_map = {}
            if sale_ids and self.core.db:
                items_count_result = self.core.db.execute_query(
                    "SELECT sale_id, COUNT(*) as cnt FROM sale_items WHERE sale_id IN %s GROUP BY sale_id",
                    (tuple(sale_ids),)
                )
                items_count_map = {r['sale_id']: r['cnt'] for r in items_count_result}

            # Export with product details
            success = export_sales_detailed_to_csv(
                sales_data,
                self.core.get_sale_items,  # Function to get items per sale
                file_path
            )

            if success:
                # Count total products exported using pre-loaded data
                total_products = sum(items_count_map.get(s['id'], 0) for s in sales_data)

                QtWidgets.QMessageBox.information(
                    self,
                    "Exportación Detallada Exitosa",
                    f"✅ {len(sales_data)} ventas exportadas\n"
                    f"📦 {total_products} productos detallados\n"
                    f"📁 Archivo: {file_path}"
                )
            else:
                QtWidgets.QMessageBox.critical(self, "Error", "Error al exportar ventas detalladas")
                
        except Exception as e:
            import traceback
            traceback.print_exc()
            QtWidgets.QMessageBox.critical(self, "Error", f"Error al exportar: {e}")
    
    def _export_all_sales_excel(self):
        """Export all sales to Excel with Dashboard"""
        from datetime import datetime

        from PyQt6.QtWidgets import QFileDialog
        
        date_from = self.date_from.date().toString("yyyy-MM-dd")
        date_to = self.date_to.date().toString("yyyy-MM-dd")
        
        # Get data
        sales_data = self.core.get_sales_by_range(date_from, date_to)
        
        if not sales_data:
            QtWidgets.QMessageBox.warning(self, "Sin datos", "No hay ventas en el período seleccionado")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Exportar Reporte Excel con Dashboard",
            f"Reporte_Ejecutivo_{date_from}_a_{date_to}.xlsx",
            "Excel (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            self._export_global_to_excel(sales_data, file_path, date_from, date_to)
            
            QtWidgets.QMessageBox.information(
                self,
                "Exportación Excel Exitosa",
                f"✅ {len(sales_data)} ventas exportadas a Excel\n"
                f"📊 Incluye Dashboard Ejecutivo\n"
                f"📁 Archivo: {file_path}"
            )
        except Exception as e:
            import traceback
            traceback.print_exc()
            QtWidgets.QMessageBox.critical(self, "Error", f"Error al exportar: {e}")
    
    def _export_global_to_excel(self, sales_data, file_path, date_from, date_to):
        """
        Export all sales to Excel with Smart Accountant Template:
        - Auto filters on all columns
        - Formulas for automatic totals
        - Conditional formatting (cancelled = red)
        - Executive dashboard with summary
        - Protected cells
        """
        try:
            from openpyxl import Workbook
            from openpyxl.formatting.rule import FormulaRule
            from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.worksheet.protection import SheetProtection
        except ImportError:
            from app.utils.export_csv import export_sales_to_csv
            file_path = file_path.replace('.xlsx', '.csv')
            export_sales_to_csv(sales_data, file_path)
            return
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Ventas"
        
        # ==================== ESTILOS ====================
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=16, color="1F4E79")
        subtitle_font = Font(bold=True, size=12, color="2F5496")
        money_font = Font(bold=True, size=14, color="228B22")
        warning_font = Font(bold=True, size=12, color="C00000")
        
        dashboard_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
        serie_a_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        serie_b_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        cancelled_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        cancelled_font = Font(color="9C0006", strikethrough=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ==================== TÍTULO ====================
        ws.merge_cells('A1:I1')
        ws['A1'] = f"📊 REPORTE EJECUTIVO DE VENTAS"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal="center")
        
        ws.merge_cells('A2:I2')
        ws['A2'] = f"Período: {date_from} al {date_to}"
        ws['A2'].font = subtitle_font
        ws['A2'].alignment = Alignment(horizontal="center")
        
        # ==================== DASHBOARD EJECUTIVO ====================
        # Pre-calculate values
        total_amount = sum(float(s.get('total', 0)) for s in sales_data if s.get('status') != 'cancelled')
        total_subtotal = sum(float(s.get('subtotal', 0)) for s in sales_data if s.get('status') != 'cancelled')
        total_iva = sum(float(s.get('tax', 0)) for s in sales_data if s.get('status') != 'cancelled')
        serie_a_count = len([s for s in sales_data if s.get('serie') == 'A' and s.get('status') != 'cancelled'])
        serie_b_count = len([s for s in sales_data if s.get('serie') == 'B' and s.get('status') != 'cancelled'])
        serie_a_amount = sum(float(s.get('total', 0)) for s in sales_data if s.get('serie') == 'A' and s.get('status') != 'cancelled')
        serie_b_amount = sum(float(s.get('total', 0)) for s in sales_data if s.get('serie') == 'B' and s.get('status') != 'cancelled')
        cancelled_count = len([s for s in sales_data if s.get('status') == 'cancelled'])
        
        # Payment method breakdown
        payment_breakdown = {}
        for s in sales_data:
            if s.get('status') != 'cancelled':
                method = s.get('payment_method', 'cash')
                payment_breakdown[method] = payment_breakdown.get(method, 0) + float(s.get('total', 0))
        
        # Dashboard Row 4-5
        dashboard_data = [
            ("A4", "📈 RESUMEN EJECUTIVO", subtitle_font, None),
            ("A5", "Total Transacciones:", None, None), ("B5", len([s for s in sales_data if s.get('status') != 'cancelled']), Font(bold=True, size=14), None),
            ("C5", "Canceladas:", None, None), ("D5", cancelled_count, warning_font, None),
            ("E5", "💰 INGRESO TOTAL:", subtitle_font, None), ("F5", f"${total_amount:,.2f}", money_font, None),
        ]
        
        for cell_ref, value, font, fill in dashboard_data:
            ws[cell_ref] = value
            if font:
                ws[cell_ref].font = font
            if fill:
                ws[cell_ref].fill = fill
        
        # Dashboard Row 6-7: Serie breakdown
        ws['A7'] = "📗 SERIE A (Fiscal):"
        ws['A7'].font = Font(bold=True, color="006400")
        ws['B7'] = f"{serie_a_count} ventas"
        ws['C7'] = f"${serie_a_amount:,.2f}"
        ws['C7'].font = Font(bold=True, color="006400")
        for cell in ['A7', 'B7', 'C7']:
            ws[cell].fill = serie_a_fill
        
        ws['E7'] = "📙 SERIE B (Público):"
        ws['E7'].font = Font(bold=True, color="996600")
        ws['F7'] = f"{serie_b_count} ventas"
        ws['G7'] = f"${serie_b_amount:,.2f}"
        ws['G7'].font = Font(bold=True, color="996600")
        for cell in ['E7', 'F7', 'G7']:
            ws[cell].fill = serie_b_fill
        
        # Dashboard Row 8: IVA
        ws['A9'] = "🏛️ IVA POR ENTERAR (SAT):"
        ws['A9'].font = Font(bold=True, color="8B0000")
        ws['C9'] = f"${total_iva:,.2f}"
        ws['C9'].font = Font(bold=True, size=14, color="8B0000")
        
        # Dashboard Row 9: Payment methods
        ws['E9'] = "💳 Por Método:"
        ws['E9'].font = Font(bold=True)
        method_str = " | ".join([f"{m}: ${v:,.2f}" for m, v in list(payment_breakdown.items())[:4]])
        ws['F9'] = method_str
        
        # ==================== TABLA DE DATOS ====================
        headers = ["Folio", "Serie", "Fecha", "Hora", "Método", "Subtotal", "IVA", "Total", "Estado"]
        header_row = 11
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        
        # Data rows
        data_start_row = 12
        for idx, sale in enumerate(sales_data):
            row = data_start_row + idx
            folio = sale.get('folio_visible') or f"#{sale.get('id', '')}"
            timestamp = str(sale.get('timestamp', ''))
            date_part = timestamp.split()[0] if ' ' in timestamp else timestamp[:10]
            timestamp_parts = timestamp.split()
            time_part = timestamp_parts[1][:8] if len(timestamp_parts) > 1 and timestamp_parts[1] else ''
            status = sale.get('status', 'completed')
            
            values = [
                folio,
                sale.get('serie', 'A'),
                date_part,
                time_part,
                sale.get('payment_method', ''),
                float(sale.get('subtotal', 0)),
                float(sale.get('tax', 0)),
                float(sale.get('total', 0)),
                'CANCELADA' if status == 'cancelled' else 'Completada'
            ]
            
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                
                # Center align most columns
                if col not in [6, 7, 8]:  # Not money columns
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = '$#,##0.00'
                
                # Serie coloring
                if col == 2:
                    if value == 'A':
                        cell.fill = serie_a_fill
                    elif value == 'B':
                        cell.fill = serie_b_fill
                
                # Cancelled row styling
                if status == 'cancelled':
                    cell.fill = cancelled_fill
                    cell.font = cancelled_font
        
        last_data_row = data_start_row + len(sales_data) - 1
        
        # ==================== FÓRMULAS DE TOTALES ====================
        totals_row = last_data_row + 2
        ws.cell(row=totals_row, column=5, value="TOTALES:").font = Font(bold=True)
        
        # Subtotal sum formula
        ws.cell(row=totals_row, column=6, value=f"=SUM(F{data_start_row}:F{last_data_row})")
        ws.cell(row=totals_row, column=6).font = Font(bold=True)
        ws.cell(row=totals_row, column=6).number_format = '$#,##0.00'
        
        # IVA sum formula
        ws.cell(row=totals_row, column=7, value=f"=SUM(G{data_start_row}:G{last_data_row})")
        ws.cell(row=totals_row, column=7).font = Font(bold=True, color="8B0000")
        ws.cell(row=totals_row, column=7).number_format = '$#,##0.00'
        
        # Total sum formula
        ws.cell(row=totals_row, column=8, value=f"=SUM(H{data_start_row}:H{last_data_row})")
        ws.cell(row=totals_row, column=8).font = money_font
        ws.cell(row=totals_row, column=8).number_format = '$#,##0.00'
        
        # ==================== FILTROS AUTOMÁTICOS ====================
        ws.auto_filter.ref = f"A{header_row}:I{last_data_row}"
        
        # ==================== ANCHOS DE COLUMNA ====================
        widths = {'A': 18, 'B': 8, 'C': 12, 'D': 10, 'E': 14, 'F': 14, 'G': 12, 'H': 14, 'I': 12}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        
        # ==================== FREEZE PANES ====================
        ws.freeze_panes = f"A{data_start_row}"
        
        # ==================== PROTECCIÓN (opcional) ====================
        # Protect formulas but allow filtering (sin contraseña)
        ws.protection.sheet = True
        ws.protection.autoFilter = False
        ws.protection.sort = False
        
        # ==================== METADATOS ====================
        from datetime import datetime
        ws[f'A{totals_row + 2}'] = f"📅 Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws[f'A{totals_row + 2}'].font = Font(italic=True, color="808080", size=9)
        ws[f'A{totals_row + 3}'] = f"🏪 TITAN POS - Sistema de Punto de Venta"
        ws[f'A{totals_row + 3}'].font = Font(italic=True, color="808080", size=9)
        
        wb.save(file_path)

    def refresh_sales(self) -> None:
        # 1. Obtener filtros
        try:
            search_text = self.search_input.text().strip().lower() if hasattr(self, 'search_input') else ""
            date_from = self.date_from.date().toString("yyyy-MM-dd") if hasattr(self, 'date_from') else ""
            date_to = self.date_to.date().toString("yyyy-MM-dd") if hasattr(self, 'date_to') else ""
            only_credit = self.credit_checkbox.isChecked() if hasattr(self, 'credit_checkbox') else False
            
            # Obtener filtro de método de pago
            payment_filter_text = self.payment_filter.currentText() if hasattr(self, 'payment_filter') else "Todos"
            payment_method_map = {
                "Todos": None,
                "Efectivo": "cash",
                "Tarjeta": "card",
                "Crédito": "credit",
                "Mixto": "mixed",
                "Transferencia": "transfer",
                "USD": "usd",
                "Vales": "voucher",
                "Cheque": "check",
                "Monedero": "wallet"
            }
            selected_payment = payment_method_map.get(payment_filter_text)
        except RuntimeError:
            return
        
        # 2. Consultar DB con rango de fechas (solo fecha, la función agrega hora)
        # #region agent log
        if agent_log_enabled():
            try:
                import json, time
                with open(get_debug_log_path_str(), "a") as f:
                    f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"REV3","location":"history_tab.py:refresh_sales","message":"Querying sales with date range","data":{"date_from":date_from,"date_to":date_to},"timestamp":int(time.time()*1000)})+"\n")
            except Exception as e:
                logger.debug("Writing debug log for sales query: %s", e)
        # #endregion
        
        sales = self.core.get_sales_by_range(date_from, date_to)
        
        # #region agent log
        if agent_log_enabled():
            try:
                with open(get_debug_log_path_str(), "a") as f:
                    f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"REV3","location":"history_tab.py:refresh_sales","message":"Sales retrieved","data":{"count":len(sales),"first_sale_has_origin_pc":"origin_pc" in (sales[0] if sales else {})},"timestamp":int(time.time()*1000)})+"\n")
            except Exception as e:
                logger.debug("Writing debug log for sales retrieval: %s", e)
        # #endregion
        
        # 3. Filtrado en memoria adicional
        filtered_sales = []
        total_amount = 0.0
        cancelled_count = 0
        
        for s in sales:
            # Filtro Texto (Folio o Ticket) - search by folio_visible or ID
            folio = s.get("folio_visible") or f"#{s.get('id', '')}"
            s_id = str(s.get("id", ""))
            if search_text and search_text not in folio.lower() and search_text not in s_id:
                continue
                
            # Filtro Crédito (legacy checkbox)
            if only_credit and s.get("payment_method") != "credit":
                continue
            
            # Filtro Método de Pago (nuevo combo)
            if selected_payment and s.get("payment_method") != selected_payment:
                continue
            
            # Count totals
            status = s.get("status", "completed")
            if status != "cancelled":
                total_amount += float(s.get("total", 0))
            else:
                cancelled_count += 1
                
            filtered_sales.append(s)
            
        sales = filtered_sales
        
        # Update totals bar
        if hasattr(self, 'total_sales_label'):
            self.total_sales_label.setText(f"💰 Total Ventas: ${total_amount:,.2f}")
        if hasattr(self, 'count_sales_label'):
            self.count_sales_label.setText(f"📊 Cantidad: {len(sales)}")
        if hasattr(self, 'cancelled_label'):
            self.cancelled_label.setText(f"🚫 Canceladas: {cancelled_count}")
        
        self.sales_table.setUpdatesEnabled(False)
        self.sales_table.setRowCount(0)
        self.sales_table.setRowCount(len(sales))
        
        # OPTIMIZATION: Pre-fetch all item counts and CFDI status in 2 batch queries
        sale_ids = [int(s["id"]) for s in sales]
        items_count_map = self.core.get_sale_items_count_batch(sale_ids) if sale_ids else {}
        cfdi_map = self.core.get_sales_cfdi_batch(sale_ids) if sale_ids else {}
        
        for row_idx, sale in enumerate(sales):
            sale_id = int(sale["id"])
            
            # Use pre-fetched data (O(1) lookup instead of SQL query)
            items_count = items_count_map.get(sale_id, 0)
            cfdi = cfdi_map.get(sale_id)
            cfdi_flag = "✓ Sí" if cfdi else "No"
            
            # Extract time from timestamp
            ts = str(sale.get("timestamp", ""))
            ts_parts = ts.split()
            time_part = ts_parts[1] if len(ts_parts) > 1 and ts_parts[1] else ts
            
            # Status check (si existe columna status)
            status = sale.get("status", "completed")
            status_text = ""
            if status == "cancelled":
                status_text = " (CANCELADA)"
            
            # Folio, Arts, Hora, PC/Origen, Total, CFDI
            # Use folio_visible if available, otherwise show #ID
            folio_display = sale.get("folio_visible") or f"#{sale['id']}"
            origin_pc = sale.get("origin_pc") or "N/A"
            
            # #region agent log
            if agent_log_enabled():
                try:
                    import json, time
                    with open(get_debug_log_path_str(), "a") as f:
                        f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"REV3","location":"history_tab.py:refresh_sales","message":"Displaying sale with origin_pc","data":{"sale_id":sale_id,"origin_pc":origin_pc,"has_origin_pc":sale.get("origin_pc") is not None},"timestamp":int(time.time()*1000)})+"\n")
                except Exception as e:
                    logger.debug("Writing debug log for sale display: %s", e)
            # #endregion
            
            values = [
                folio_display + status_text,
                str(items_count),
                time_part,
                origin_pc,
                f"${float(sale['total']):.2f}",
                cfdi_flag
            ]
            
            for col, value in enumerate(values):
                item = QtWidgets.QTableWidgetItem(str(value))
                item.setFlags(item.flags() & ~QtCore.Qt.ItemFlag.ItemIsEditable)
                
                # Store actual sale_id in first column's UserRole for later retrieval
                if col == 0:
                    item.setData(QtCore.Qt.ItemDataRole.UserRole, sale_id)
                
                # Center align all columns
                item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
                
                # FIX B5 2026-01-30: Índices correctos - col 4 = Total, col 5 = CFDI
                # Format total column
                if col == 4:  # Total column (índice 4, no 3)
                    font = item.font()
                    font.setBold(True)
                    item.setFont(font)
                    item.setForeground(QtGui.QColor("#27ae60"))

                # Format CFDI column
                if col == 5 and cfdi:  # CFDI column (índice 5, no 4)
                    item.setForeground(QtGui.QColor("#4CAF50"))
                    
                if status == "cancelled":
                    item.setForeground(QtGui.QColor("#e74c3c"))
                    font = item.font()
                    font.setStrikeOut(True)
                    item.setFont(font)
                
                self.sales_table.setItem(row_idx, col, item)
        
        self.sales_table.setUpdatesEnabled(True)

    def cancel_sale(self) -> None:
        if not self.selected_sale_id:
            QtWidgets.QMessageBox.warning(self, "Aviso", "Seleccione una venta para cancelar.")
            return
            
        # Verificar si ya está cancelada
        sale = self.core.get_sale(self.selected_sale_id)
        if sale.get("status") == "cancelled":
            QtWidgets.QMessageBox.warning(self, "Aviso", "Esta venta ya está cancelada.")
            return
        
        # Verificar si el usuario actual tiene permiso para cancelar
        from app.core import STATE
        current_role = getattr(STATE, 'role', 'cashier').lower()
        
        # Si es cajero y no tiene permiso, requiere autorización de supervisor
        # Verificar primero con el sistema de permisos
        from app.utils.permissions import can_cancel_sale
        if not can_cancel_sale() or current_role == 'cashier':
            from app.dialogs.supervisor_override_dialog import require_supervisor_override
            
            total = float(sale.get('total', 0))
            folio = sale.get('folio_visible', f"#{self.selected_sale_id}")
            
            authorized, supervisor = require_supervisor_override(
                core=self.core,
                action_description=f"Cancelar Venta {folio}\nTotal: ${total:,.2f}",
                required_permission="cancel_sale",
                min_role="encargado",
                parent=self
            )
            
            if not authorized:
                return  # Usuario canceló o no se autorizó

        from app.dialogs.cancel_sale_dialog import CancelSaleDialog
        dlg = CancelSaleDialog(self.core, self.selected_sale_id, parent=self)
        if dlg.exec() == QtWidgets.QDialog.DialogCode.Accepted:
            # Abrir cajón después de cancelar venta
            try:
                from app.utils.ticket_engine import open_cash_drawer_safe
                open_cash_drawer_safe(core=self.core)
            except Exception as e:
                logging.warning(f"No se pudo abrir cajón después de cancelar venta: {e}")
            self.refresh_sales()
            self.refresh_items()

    def refresh_items(self) -> None:
        row = self.sales_table.currentRow()
        if row < 0:
            self.items_table.setRowCount(0)
            self.payment_label.setText("Pago con: -")
            self.total_label.setText("Total: $0.00")
            self.selected_sale_id = None
            return
            
        # Get sale_id from UserRole (stored as integer during table population)
        first_cell = self.sales_table.item(row, 0)
        sale_id = first_cell.data(QtCore.Qt.ItemDataRole.UserRole)
        if sale_id is None:
            # Fallback for legacy: try to parse numeric ID from text
            sale_id_text = first_cell.text()
            # Handle #123 format for legacy sales without folio_visible
            if sale_id_text.startswith('#'):
                sale_id_parts = sale_id_text.split()
                if sale_id_parts and len(sale_id_parts[0]) > 1:
                    try:
                        sale_id = int(sale_id_parts[0][1:])  # Remove # prefix
                    except (ValueError, IndexError):
                        self.items_table.setRowCount(0)
                        return
                else:
                    self.items_table.setRowCount(0)
                    return
            else:
                # This shouldn't happen with new sales
                self.items_table.setRowCount(0)
                return
        self.selected_sale_id = sale_id
        
        # Get sale info
        sale = self.core.get_sale(sale_id)
        payment_method = sale["payment_method"] if sale and "payment_method" in sale.keys() else "cash"
        total = float(sale["total"]) if sale and "total" in sale.keys() else 0.0
        
        # Map payment method to Spanish
        payment_map = {
            "cash": "Efectivo 💵",
            "card": "Tarjeta 💳",
            "transfer": "Transferencia 🏦", 
            "usd": "Dólares 💵",
            "voucher": "Vales 🎫",
            "check": "Cheque 💷",
            "credit": "Crédito 📋",
            "wallet": "Monedero 👛",
            "mixed": "Mixto 🔀"
        }
        
        payment_text = payment_map.get(payment_method, payment_method.title())
        self.payment_label.setText(f"Pago con: {payment_text}")
        self.total_label.setText(f"Total: ${total:.2f}")
        
        # Get and display items
        items = self.core.get_sale_items(sale_id)
        self.items_table.setUpdatesEnabled(False)
        self.items_table.setRowCount(0)
        self.items_table.setRowCount(len(items))
        
        for idx, item in enumerate(items):
            qty = float(item.get("qty") or 0)
            price = float(item.get("price") or 0)
            item_total = float(item.get("total") or item.get("subtotal") or (qty * price))
            
            # CRITICAL FIX: El precio guardado en sale_items.price es SIN IVA (porque price_includes_tax era True)
            # Por lo tanto, debemos calcular el precio con IVA para mostrarlo correctamente
            TAX_RATE = 0.16
            price_to_display = price * (1 + TAX_RATE)
            item_total_with_tax = (qty * price_to_display) - float(item.get('discount', 0))
            
            # Cant., Descripción, Importe, Total
            values = [
                f"{qty:.0f}" if qty == int(qty) else f"{qty:.2f}",
                item["name"] if "name" in item.keys() else "",
                f"${price_to_display:.2f}",
                f"${item_total_with_tax:.2f}"
            ]
            
            for col, value in enumerate(values):
                cell = QtWidgets.QTableWidgetItem(str(value))
                cell.setFlags(cell.flags() & ~QtCore.Qt.ItemFlag.ItemIsEditable)
                
                # Format columns
                if col == 0:  # Quantity - center
                    cell.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
                elif col in [2, 3]:  # Price and Total - right align and bold
                    cell.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignRight | QtCore.Qt.AlignmentFlag.AlignVCenter)
                    font = cell.font()
                    font.setBold(True)
                    cell.setFont(font)
                    if col == 3:  # Total column - green
                        cell.setForeground(QtGui.QColor("#27ae60"))
                
                self.items_table.setItem(idx, col, cell)
        
        self.items_table.setUpdatesEnabled(True)

    def _selected_sale_id(self) -> int | None:
        row = self.sales_table.currentRow()
        if row < 0:
            return None
        # Get sale_id from UserRole (stored during table population)
        first_cell = self.sales_table.item(row, 0)
        if not first_cell:
            return None
        sale_id = first_cell.data(QtCore.Qt.ItemDataRole.UserRole)
        if sale_id is not None:
            return int(sale_id)
        # Fallback for legacy: try to parse #123 format
        sale_id_text = first_cell.text()
        if sale_id_text.startswith('#'):
            sale_id_parts = sale_id_text.split()
            if sale_id_parts and len(sale_id_parts[0]) > 1:
                try:
                    return int(sale_id_parts[0][1:])
                except (ValueError, IndexError):
                    return None
        return None

    def _issue_cfdi(self) -> None:
        """Generate CFDI for selected sale."""
        sale_id = self._selected_sale_id()
        if not sale_id:
            QtWidgets.QMessageBox.warning(self, "Facturar", "Selecciona una venta")
            return
        
        # Check if already has CFDI
        existing_cfdi = self.core.get_cfdi_by_sale(sale_id)
        if existing_cfdi:
            uuid = existing_cfdi.get('uuid', 'N/A')
            QtWidgets.QMessageBox.information(
                self, 
                "Ya facturada",
                f"Esta venta ya tiene CFDI\n\nUUID: {uuid}"
            )
            return
        
        # Check fiscal configuration
        fiscal_config = self.core.get_fiscal_config()
        if not fiscal_config or not fiscal_config.get('rfc_emisor'):
            QtWidgets.QMessageBox.warning(
                self,
                "Configuración requerida",
                "Configure primero los datos fiscales en:\n"
                "Settings → Facturación"
            )
            return
        
        # Show customer dialog
        from app.dialogs.cfdi_customer_dialog import CFDICustomerDialog
        
        dialog = CFDICustomerDialog(self)
        if dialog.exec() != QtWidgets.QDialog.DialogCode.Accepted:
            return
        
        customer_data = dialog.get_customer_data()
        
        # Show progress dialog
        progress = QtWidgets.QProgressDialog(
            "Generando CFDI...\n\nEsto puede tardar unos segundos.",
            None,  # No cancel button
            0,
            0,  # Indeterminate
            self
        )
        progress.setWindowTitle("Facturación Electrónica")
        progress.setWindowModality(QtCore.Qt.WindowModality.WindowModal)
        progress.setMinimumDuration(0)
        progress.show()
        
        # Process events to show progress
        QtWidgets.QApplication.processEvents()
        
        try:
            # Generate CFDI
            result = self.core.issue_cfdi_for_sale(
                sale_id=sale_id,
                customer_rfc=customer_data['rfc'],
                customer_name=customer_data['nombre'],
                customer_regime=customer_data['regimen'],
                uso_cfdi=customer_data['uso_cfdi'],
                forma_pago=customer_data.get('forma_pago', '01'),
                customer_zip=customer_data.get('codigo_postal', '00000')
            )
            
            progress.close()
            
            if result.get('success'):
                uuid = result.get('uuid', 'N/A')
                xml_path = result.get('xml_path', '')
                
                # Success message
                msg = f"✅ CFDI generado correctamente\n\n"
                msg += f"UUID: {uuid}\n"
                msg += f"RFC: {customer_data['rfc']}\n"
                msg += f"Cliente: {customer_data['nombre']}\n\n"
                
                if xml_path:
                    msg += f"XML guardado en:\n{xml_path}"
                
                reply = QtWidgets.QMessageBox.information(
                    self,
                    "CFDI Generado",
                    msg,
                    QtWidgets.QMessageBox.StandardButton.Ok
                )
                
                # Refresh table
                self.refresh_sales()
                
            else:
                error = result.get('error', 'Error desconocido')
                QtWidgets.QMessageBox.critical(
                    self,
                    "Error al generar CFDI",
                    f"No se pudo generar el CFDI:\n\n{error}\n\n"
                    f"Verifique:\n"
                    f"- Configuración fiscal completa\n"
                    f"- Certificados CSD válidos\n"
                    f"- Credenciales del PAC"
                )
            
        except Exception as exc:
            progress.close()
            logging.exception("CFDI generation failed")
            QtWidgets.QMessageBox.critical(
                self,
                "Error",
                f"Error al generar CFDI:\n\n{str(exc)}"
            )

    def _view_cfdi(self) -> None:
        """View CFDI XML for selected sale."""
        sale_id = self._selected_sale_id()
        if not sale_id:
            return
        
        cfdi = self.core.get_cfdi_by_sale(sale_id)
        if not cfdi:
            QtWidgets.QMessageBox.information(
                self, 
                "Sin CFDI", 
                "Esta venta no tiene CFDI generado"
            )
            return
        
        # Show XML in dialog
        xml_content = cfdi.get('xml_timbrado') or cfdi.get('xml_original', '')
        if not xml_content:
            QtWidgets.QMessageBox.warning(self, "Error", "No se encontró el XML")
            return
        
        # Create dialog to show XML
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle(f"CFDI XML - UUID: {cfdi.get('uuid', 'N/A')}")
        dialog.setMinimumSize(800, 600)
        
        layout = QtWidgets.QVBoxLayout(dialog)
        
        # Info header
        info_text = f"UUID: {cfdi.get('uuid', 'N/A')}\n"
        info_text += f"RFC: {cfdi.get('rfc_receptor', 'N/A')}\n"
        info_text += f"Total: ${cfdi.get('total', 0):.2f}\n"
        info_text += f"Estado: {cfdi.get('estado', 'N/A')}"
        
        info_label = QtWidgets.QLabel(info_text)
        info_label.setStyleSheet("font-weight: bold; padding: 10px; background: #ecf0f1; border-radius: 5px;")
        layout.addWidget(info_label)
        
        # XML viewer
        xml_viewer = QtWidgets.QTextEdit()
        xml_viewer.setReadOnly(True)
        xml_viewer.setPlainText(xml_content)
        xml_viewer.setStyleSheet("font-family: monospace; font-size: 10pt;")
        layout.addWidget(xml_viewer)
        
        # Buttons
        btn_layout = QtWidgets.QHBoxLayout()
        btn_layout.addStretch()
        
        # Download PDF button (if facturapi_id exists)
        facturapi_id = cfdi.get('facturapi_id')
        if facturapi_id:
            btn_pdf = QtWidgets.QPushButton("📄 Descargar PDF")
            btn_pdf.setStyleSheet("background: #e74c3c; color: white; font-weight: bold; padding: 8px;")
            btn_pdf.clicked.connect(lambda: self._download_cfdi_pdf(facturapi_id, cfdi.get('uuid', 'factura')))
            btn_layout.addWidget(btn_pdf)
            
            btn_xml = QtWidgets.QPushButton("📝 Descargar XML")
            btn_xml.setStyleSheet("background: #3498db; color: white; font-weight: bold; padding: 8px;")
            btn_xml.clicked.connect(lambda: self._download_cfdi_xml(facturapi_id, cfdi.get('uuid', 'factura')))
            btn_layout.addWidget(btn_xml)
            
            btn_zip = QtWidgets.QPushButton("📦 Descargar ZIP")
            btn_zip.setStyleSheet("background: #27ae60; color: white; font-weight: bold; padding: 8px;")
            btn_zip.clicked.connect(lambda: self._download_cfdi_zip(facturapi_id, cfdi.get('uuid', 'factura')))
            btn_layout.addWidget(btn_zip)
        
        btn_copy = QtWidgets.QPushButton("📋 Copiar al portapapeles")
        btn_copy.clicked.connect(lambda: QtWidgets.QApplication.clipboard().setText(xml_content))
        btn_layout.addWidget(btn_copy)
        
        btn_close = QtWidgets.QPushButton("Cerrar")
        btn_close.clicked.connect(dialog.accept)
        btn_layout.addWidget(btn_close)
        
        layout.addLayout(btn_layout)
        
        dialog.exec()
    
    def _download_cfdi_pdf(self, facturapi_id: str, filename: str) -> None:
        """Download CFDI PDF from Facturapi."""
        try:
            from app.fiscal.facturapi_connector import Facturapi
            fiscal_config = self.core.get_fiscal_config()
            api_key = fiscal_config.get('facturapi_api_key', '')
            
            if not api_key:
                QtWidgets.QMessageBox.warning(self, "Error", "API Key de Facturapi no configurada")
                return
            
            facturapi = Facturapi(api_key)
            pdf_content = facturapi.invoices.download_pdf(facturapi_id)
            
            # Save file dialog
            save_path, _ = QtWidgets.QFileDialog.getSaveFileName(
                self, "Guardar PDF", f"{filename}.pdf", "PDF Files (*.pdf)"
            )
            
            if save_path:
                with open(save_path, 'wb') as f:
                    f.write(pdf_content)
                QtWidgets.QMessageBox.information(self, "Éxito", f"PDF guardado en:\n{save_path}")
                
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Error", f"Error descargando PDF:\n{str(e)}")
    
    def _download_cfdi_xml(self, facturapi_id: str, filename: str) -> None:
        """Download CFDI XML from Facturapi."""
        try:
            from app.fiscal.facturapi_connector import Facturapi
            fiscal_config = self.core.get_fiscal_config()
            api_key = fiscal_config.get('facturapi_api_key', '')
            
            if not api_key:
                QtWidgets.QMessageBox.warning(self, "Error", "API Key de Facturapi no configurada")
                return
            
            facturapi = Facturapi(api_key)
            xml_content = facturapi.invoices.download_xml(facturapi_id)
            
            # Save file dialog
            save_path, _ = QtWidgets.QFileDialog.getSaveFileName(
                self, "Guardar XML", f"{filename}.xml", "XML Files (*.xml)"
            )
            
            if save_path:
                with open(save_path, 'w', encoding='utf-8') as f:
                    f.write(xml_content)
                QtWidgets.QMessageBox.information(self, "Éxito", f"XML guardado en:\n{save_path}")
                
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Error", f"Error descargando XML:\n{str(e)}")
    
    def _download_cfdi_zip(self, facturapi_id: str, filename: str) -> None:
        """Download CFDI ZIP (PDF+XML) from Facturapi."""
        try:
            from app.fiscal.facturapi_connector import Facturapi
            fiscal_config = self.core.get_fiscal_config()
            api_key = fiscal_config.get('facturapi_api_key', '')
            
            if not api_key:
                QtWidgets.QMessageBox.warning(self, "Error", "API Key de Facturapi no configurada")
                return
            
            facturapi = Facturapi(api_key)
            zip_content = facturapi.invoices.download_zip(facturapi_id)
            
            # Save file dialog
            save_path, _ = QtWidgets.QFileDialog.getSaveFileName(
                self, "Guardar ZIP", f"{filename}.zip", "ZIP Files (*.zip)"
            )
            
            if save_path:
                with open(save_path, 'wb') as f:
                    f.write(zip_content)
                QtWidgets.QMessageBox.information(self, "Éxito", f"ZIP guardado en:\n{save_path}\n\nContiene PDF y XML")
                
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Error", f"Error descargando ZIP:\n{str(e)}")

    def _cancel_cfdi(self) -> None:
        """Cancel CFDI for selected sale."""
        sale_id = self._selected_sale_id()
        if not sale_id:
            return
        
        cfdi = self.core.get_cfdi_by_sale(sale_id)
        if not cfdi:
            QtWidgets.QMessageBox.warning(
                self, 
                "Sin CFDI", 
                "No hay CFDI para cancelar"
            )
            return
        
        if cfdi.get('estado') == 'cancelado':
            QtWidgets.QMessageBox.information(
                self,
                "Ya cancelado",
                f"Este CFDI ya está cancelado\n\nUUID: {cfdi.get('uuid')}"
            )
            return
        
        # Ask for cancellation reason
        motivos = [
            "01 - Comprobante emitido con errores con relación",
            "02 - Comprobante emitido con errores sin relación",
            "03 - No se llevó a cabo la operación",
            "04 - Operación nominativa relacionada"
        ]
        
        motivo, ok = QtWidgets.QInputDialog.getItem(
            self,
            "Cancelar CFDI",
            "Seleccione el motivo de cancelación:",
            motivos,
            0,
            False
        )
        
        if not ok:
            return
        
        motivo_code = motivo.split(' -')[0].strip()
        
        # Confirm
        reply = QtWidgets.QMessageBox.question(
            self,
            "Confirmar cancelación",
            f"¿Está seguro de cancelar este CFDI?\n\n"
            f"UUID: {cfdi.get('uuid')}\n"
            f"Motivo: {motivo}\n\n"
            f"Esta acción se registrará en el SAT.",
            QtWidgets.QMessageBox.StandardButton.Yes | QtWidgets.QMessageBox.StandardButton.No
        )
        
        if reply != QtWidgets.QMessageBox.StandardButton.Yes:
            return
        
        try:
            result = self.core.cancel_cfdi(cfdi.get('uuid'), motivo_code)
            
            if result.get('success'):
                QtWidgets.QMessageBox.information(
                    self,
                    "CFDI Cancelado",
                    f"El CFDI ha sido cancelado correctamente\n\nUUID: {cfdi.get('uuid')}"
                )
                self.refresh_sales()
            else:
                error = result.get('error', 'Error desconocido')
                QtWidgets.QMessageBox.critical(
                    self,
                    "Error al cancelar",
                    f"No se pudo cancelar el CFDI:\n\n{error}"
                )
                
        except Exception as exc:
            logging.exception("CFDI cancellation failed")
            QtWidgets.QMessageBox.critical(
                self,
                "Error",
                f"Error al cancelar CFDI:\n\n{str(exc)}"
            )

    def update_theme(self) -> None:
        cfg = self.core.read_local_config()
        theme = cfg.get("theme", "Light")
        c = theme_manager.get_colors(theme)
        
        self.setStyleSheet(f"background-color: {c['bg_main']}; color: {c['text_primary']};")
        
        if hasattr(self, "header"):
            self.header.setStyleSheet(f"""
                QFrame {{
                    background: {c['bg_header']};
                    border-bottom: 1px solid {c['border']};
                }}
            """)
            
        if hasattr(self, "title_label"):
            self.title_label.setStyleSheet(f"color: {c['text_header']}; font-size: 20px; font-weight: 800; letter-spacing: 1px; background: transparent;")
            
        if hasattr(self, "content_widget"):
            self.content_widget.setStyleSheet(f"background-color: {c['bg_main']};")
            
        if hasattr(self, "dashboard_frame"):
            self.dashboard_frame.setStyleSheet(f"""
                QFrame {{
                    background: {c['bg_card']};
                    border: 1px solid {c['border']};
                    border-bottom: 2px solid {c['border']};
                    border-radius: 10px;
                }}
            """)
            
        if hasattr(self, "search_input"):
            self.search_input.setStyleSheet(f"""
                QLineEdit {{
                    background: {c['input_bg']};
                    border: 1px solid {c['input_border']};
                    border-radius: 8px;
                    padding: 0 15px;
                    font-size: 14px;
                    color: {c['text_primary']};
                }}
                QLineEdit:focus {{
                    background: {c['bg_card']};
                    border: 2px solid {c['input_focus']};
                }}
            """)
            
        if hasattr(self, "credit_checkbox"):
            self.credit_checkbox.setStyleSheet(f"font-weight: bold; color: {c['text_secondary']}; font-size: 13px;")

        if hasattr(self, "action_buttons_data"):
            for btn, color_key in self.action_buttons_data:
                color = c.get(color_key, color_key)
                btn.setStyleSheet(f"""
                    QPushButton {{
                        background: {color}; color: white; border: none;
                        border-radius: 8px; padding: 0 15px; font-weight: bold; font-size: 13px;
                    }}
                    QPushButton:hover {{ opacity: 0.9; }}
                """)
        
        if hasattr(self, "splitter"):
            self.splitter.setStyleSheet("QSplitter::handle { background: transparent; }")

        if hasattr(self, "left_card"):
            self.left_card.setStyleSheet(f"""
                QFrame {{
                    background: {c['bg_card']};
                    border: 1px solid {c['border']};
                    border-bottom: 2px solid {c['border']};
                    border-radius: 10px;
                }}
            """)
            
        if hasattr(self, "sales_table"):
            self.sales_table.setStyleSheet(f"""
                QTableWidget {{
                    background: {c['bg_card']}; border: none; border-radius: 10px;
                    gridline-color: transparent; font-size: 13px;
                }}
                QHeaderView::section {{
                    background: {c['table_header_bg']};
                    color: {c['table_header_text']}; padding: 12px; border: none;
                    border-bottom: 1px solid {c['border']};
                    font-weight: bold; font-size: 12px;
                }}
                QTableWidget::item {{ padding: 10px; border-bottom: 1px solid {c['border']}; color: {c['table_text']}; }}
                QTableWidget::item:selected {{ background: {c['table_selected']}; color: {c['bg_header']}; }}
            """)
            
        if hasattr(self, "right_card"):
            self.right_card.setStyleSheet(f"""
                QFrame {{
                    background: {c['bg_card']};
                    border: 1px solid {c['border']};
                    border-bottom: 2px solid {c['border']};
                    border-radius: 10px;
                }}
            """)
            
        if hasattr(self, "detail_header"):
            self.detail_header.setStyleSheet(f"font-weight: bold; color: {c['text_secondary']}; font-size: 14px; padding: 15px; border-bottom: 1px solid {c['border']};")
            
        if hasattr(self, "items_table"):
            self.items_table.setStyleSheet(f"""
                QTableWidget {{
                    background: {c['bg_card']}; border: none;
                    gridline-color: transparent; font-size: 12px;
                }}
                QHeaderView::section {{
                    background: {c['table_header_bg']}; padding: 8px; border: none;
                    border-bottom: 1px solid {c['border']}; font-weight: bold; color: {c['table_header_text']};
                }}
                QTableWidget::item {{ padding: 6px; border-bottom: 1px solid {c['border']}; color: {c['table_text']}; }}
            """)
        
        if hasattr(self, "payment_frame"):
            self.payment_frame.setStyleSheet(f"background: {c['bg_card']}; border-top: 1px solid {c['border']}; border-bottom-left-radius: 10px; border-bottom-right-radius: 10px;")
            
        if hasattr(self, "payment_label"):
            self.payment_label.setStyleSheet(f"font-size: 13px; font-weight: bold; color: {c['text_secondary']};")
            
        if hasattr(self, "total_label"):
            self.total_label.setStyleSheet(f"font-size: 22px; font-weight: 900; color: {c['btn_success']};")
        
        # Style new date range elements
        for lbl_name in ['lbl_date_from', 'lbl_date_to']:
            if hasattr(self, lbl_name):
                getattr(self, lbl_name).setStyleSheet(f"font-weight: bold; color: {c['text_secondary']}; font-size: 13px;")
        
        for date_name in ['date_from', 'date_to']:
            if hasattr(self, date_name):
                getattr(self, date_name).setStyleSheet(f"""
                    QDateEdit {{
                        border: 1px solid {c['input_border']}; border-radius: 6px; padding: 5px 10px; 
                        background: {c['input_bg']}; font-size: 13px; color: {c['text_primary']};
                    }}
                    QDateEdit::drop-down {{ border: none; }}
                """)
        
        # Style quick filter buttons
        for btn_name in ['today_btn', 'week_btn', 'month_btn']:
            if hasattr(self, btn_name):
                getattr(self, btn_name).setStyleSheet(f"""
                    QPushButton {{
                        background: {c['bg_card']}; color: {c['btn_primary']}; border: 1px solid {c['btn_primary']};
                        border-radius: 6px; padding: 0 12px; font-weight: bold; font-size: 12px;
                    }}
                    QPushButton:hover {{ background: {c['btn_primary']}; color: white; }}
                """)
        
        # Style totals frame
        if hasattr(self, "totals_frame"):
            self.totals_frame.setStyleSheet(f"""
                QFrame {{
                    background: {c['bg_card']};
                    border: 1px solid {c['border']};
                    border-radius: 8px;
                }}
            """)
        
        for lbl_name in ['total_sales_label', 'count_sales_label', 'cancelled_label']:
            if hasattr(self, lbl_name):
                color = c['btn_success'] if 'total' in lbl_name else (c['text_secondary'] if 'count' in lbl_name else c['btn_danger'])
                getattr(self, lbl_name).setStyleSheet(f"font-size: 14px; font-weight: bold; color: {color}; padding: 5px 15px;")

        # Force style update
        self.style().unpolish(self)
        self.style().polish(self)
        self.update()

    def _import_sales(self):
        """Open the import wizard for sales."""
        try:
            from app.wizards.import_wizard import ImportWizard
            wizard = ImportWizard(parent=self, core=self.core)
            # Pre-select "Ventas (Formato Simple)" which is index 2
            if hasattr(wizard, 'page') and wizard.page(0):
                intro_page = wizard.page(0)
                if hasattr(intro_page, 'type_combo'):
                    intro_page.type_combo.setCurrentIndex(2)  # Ventas (Formato Simple)
            if wizard.exec():
                QtWidgets.QMessageBox.information(
                    self,
                    "Importación Exitosa",
                    "Las ventas se han importado correctamente."
                )
                self.refresh_sales()
        except Exception as e:
            logger.exception("Error opening import wizard")
            QtWidgets.QMessageBox.critical(
                self,
                "Error",
                f"Error al abrir wizard de importación:\n\n{str(e)}"
            )

    def showEvent(self, event):
        """Apply theme when tab is shown."""
        super().showEvent(event)
        self.update_theme()
