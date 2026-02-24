"""
Export Wizard for Products, Customers, and Sales
"""
import csv
import logging
from datetime import datetime

logger = logging.getLogger(__name__)

from PyQt6 import QtCore, QtWidgets

# FIX 2026-02-01: Import agent_log_enabled ANTES de usarla
try:
    from app.utils.path_utils import agent_log_enabled
except ImportError:
    def agent_log_enabled(): return False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ExportWizard(QtWidgets.QWizard):
    """Wizard for exporting data to CSV"""
    
    def __init__(self, parent=None, core=None):
        super().__init__(parent)
        self.core = core
        self.setWindowTitle("Asistente de Exportación")
        self.setWizardStyle(QtWidgets.QWizard.WizardStyle.ModernStyle)
        
        # Add pages
        self.addPage(ExportTypePage())
        self.addPage(ExportOptionsPage())
        self.addPage(ExportExecutionPage(core=self.core))
        
        self.resize(500, 350)
        self._apply_theme()
    
    def showEvent(self, event):
        """Apply theme when shown."""
        super().showEvent(event)
        self._apply_theme()
    
    def _apply_theme(self):
        """Apply current theme colors."""
        try:
            from app.utils.theme_manager import theme_manager
            c = theme_manager.get_colors()
            self.setStyleSheet(f"""
                QWizard {{ background: {c['bg_secondary']}; }}
                QWizardPage {{ background: {c['bg_primary']}; }}
                QLabel {{ color: {c['text_primary']}; }}
                QLineEdit, QTextEdit, QComboBox {{
                    background: {c['bg_secondary']}; color: {c['text_primary']};
                    border: 1px solid {c['border']}; padding: 8px; border-radius: 4px;
                }}
                QRadioButton, QCheckBox {{ color: {c['text_primary']}; }}
                QPushButton {{
                    background: {c['btn_primary']}; color: white;
                    padding: 8px 16px; border-radius: 4px; font-weight: bold;
                }}
                QPushButton:hover {{ background: {c['btn_success']}; }}
                QProgressBar {{
                    border: 1px solid {c['border']}; border-radius: 4px;
                    background: {c['bg_secondary']}; text-align: center;
                }}
                QProgressBar::chunk {{ background: {c['btn_success']}; }}
            """)
        except Exception as e:
            logger.debug("Applying export wizard theme: %s", e)

class ExportTypePage(QtWidgets.QWizardPage):
    """Select what to export"""
    
    def __init__(self):
        super().__init__()
        self.setTitle("Exportar Datos")
        self.setSubTitle("Selecciona qué datos deseas exportar")
        
        layout = QtWidgets.QVBoxLayout(self)
        
        info = QtWidgets.QLabel(
            "Exporta tus datos a CSV o Excel para:\n\n"
            "• Respaldo\n"
            "• Análisis en Excel\n"
            "• Migración a otro sistema"
        )
        layout.addWidget(info)
        
        # Export options
        self.products_radio = QtWidgets.QRadioButton("Productos")
        self.products_radio.setChecked(True)
        self.customers_radio = QtWidgets.QRadioButton("Clientes")
        self.sales_radio = QtWidgets.QRadioButton("Historial de Ventas")
        
        layout.addWidget(self.products_radio)
        layout.addWidget(self.customers_radio)
        layout.addWidget(self.sales_radio)
        
        layout.addStretch()
        
        self.registerField("export_products", self.products_radio)
        self.registerField("export_customers", self.customers_radio)
        self.registerField("export_sales", self.sales_radio)

class ExportOptionsPage(QtWidgets.QWizardPage):
    """Export options and destination"""
    
    def __init__(self):
        super().__init__()
        self.setTitle("Opciones de Exportación")
        self.setSubTitle("Configura la exportación")
        
        layout = QtWidgets.QVBoxLayout(self)
        
        # File selection
        file_layout = QtWidgets.QHBoxLayout()
        self.file_edit = QtWidgets.QLineEdit()
        self.file_edit.setText(f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")
        
        browse_btn = QtWidgets.QPushButton("Guardar como...")
        browse_btn.clicked.connect(self.choose_file)
        
        file_layout.addWidget(QtWidgets.QLabel("Archivo:"))
        file_layout.addWidget(self.file_edit)
        file_layout.addWidget(browse_btn)
        layout.addLayout(file_layout)
        
        # Format selection
        format_layout = QtWidgets.QHBoxLayout()
        format_layout.addWidget(QtWidgets.QLabel("Formato:"))
        self.format_combo = QtWidgets.QComboBox()
        self.format_combo.addItems(["CSV", "Excel (.xlsx)"])
        format_layout.addWidget(self.format_combo)
        format_layout.addStretch()
        layout.addLayout(format_layout)
        
        # Options
        self.include_header_check = QtWidgets.QCheckBox("Incluir encabezados")
        self.include_header_check.setChecked(True)
        layout.addWidget(self.include_header_check)
        
        layout.addStretch()
        
        self.registerField("export_file*", self.file_edit)
        self.registerField("export_format", self.format_combo, "currentText")
        self.registerField("include_headers", self.include_header_check)
    
    def choose_file(self):
        """Choose export file"""
        current_format = self.format_combo.currentText()
        if "Excel" in current_format:
            filter_str = "Excel Files (*.xlsx);;CSV Files (*.csv)"
        else:
            filter_str = "CSV Files (*.csv);;Excel Files (*.xlsx)"
        
        file_path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "Guardar Exportación", self.file_edit.text(), filter_str
        )
        if file_path:
            self.file_edit.setText(file_path)
            # Actualizar formato según extensión
            if file_path.lower().endswith('.xlsx'):
                self.format_combo.setCurrentText("Excel (.xlsx)")
            elif file_path.lower().endswith('.csv'):
                self.format_combo.setCurrentText("CSV")

class ExportExecutionPage(QtWidgets.QWizardPage):
    """Execute export"""
    
    def __init__(self, core=None):
        super().__init__()
        self.core = core
        self.setTitle("Exportando...")
        self.setSubTitle("Generando archivo CSV")
        
        layout = QtWidgets.QVBoxLayout(self)
        
        self.progress = QtWidgets.QProgressBar()
        layout.addWidget(self.progress)
        
        self.status_label = QtWidgets.QLabel("Preparando...")
        layout.addWidget(self.status_label)
        
        self.result_text = QtWidgets.QTextEdit()
        self.result_text.setReadOnly(True)
        self.result_text.setMaximumHeight(150)
        layout.addWidget(self.result_text)
        
        layout.addStretch()
        
        self.setCommitPage(True)
        self.setButtonText(QtWidgets.QWizard.WizardButton.CommitButton, "Exportar")
    
    def initializePage(self):
        """Start export"""
        # #region agent log
        if agent_log_enabled():
            import json, time
            try:
                from app.utils.path_utils import get_debug_log_path_str  # FIX: agent_log_enabled ya importado arriba
                with open(get_debug_log_path_str(), "a") as f:
                    f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"EXPORT_WIZARD","location":"export_wizard.py:initializePage","message":"Export wizard page initialized","data":{"core_db_is_none":self.core.db is None if self.core else None},"timestamp":int(time.time()*1000)})+"\n")
            except Exception as e: logger.debug("Writing export wizard init log: %s", e)
        # #endregion
        
        # Verificar que core y db estén disponibles
        if not self.core:
            self.result_text.append("❌ Error: Core no disponible")
            self.status_label.setText("❌ Error")
            return
        
        if not self.core.db:
            # #region agent log
            if agent_log_enabled():
                import json, time
                try:
                    from app.utils.path_utils import get_debug_log_path_str
                    with open(get_debug_log_path_str(), "a") as f:
                        f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"EXPORT_WIZARD","location":"export_wizard.py:initializePage","message":"DB is None, attempting to initialize","data":{},"timestamp":int(time.time()*1000)})+"\n")
                except Exception as e: logger.debug("Writing DB init attempt log: %s", e)
            # #endregion
            # Intentar inicializar la base de datos
            if not self.core.initialize_database():
                self.result_text.append("❌ Error: No se pudo conectar a la base de datos.\n")
                self.result_text.append("Verifica la configuración de PostgreSQL.")
                self.status_label.setText("❌ Error")
                return
        
        file_path = self.field("export_file")
        export_format = self.field("export_format")
        include_headers = self.field("include_headers")
        
        # Ajustar extensión según formato seleccionado
        if "Excel" in export_format and not file_path.lower().endswith('.xlsx'):
            file_path = file_path.rsplit('.', 1)[0] + '.xlsx'
        elif export_format == "CSV" and not file_path.lower().endswith('.csv'):
            file_path = file_path.rsplit('.', 1)[0] + '.csv'
        
        # #region agent log
        if agent_log_enabled():
            import json, time
            try:
                from app.utils.path_utils import get_debug_log_path_str
                with open(get_debug_log_path_str(), "a") as f:
                    export_type = "products" if self.field("export_products") else ("customers" if self.field("export_customers") else "sales")
                    f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"EXPORT_WIZARD","location":"export_wizard.py:initializePage","message":"Starting export","data":{"export_type":export_type,"file_path":file_path,"format":export_format},"timestamp":int(time.time()*1000)})+"\n")
            except Exception as e: logger.debug("Writing starting export log: %s", e)
        # #endregion
        
        if self.field("export_products"):
            self.export_products(file_path, include_headers, export_format)
        elif self.field("export_customers"):
            self.export_customers(file_path, include_headers, export_format)
        else:
            self.export_sales(file_path, include_headers, export_format)
    
    def export_products(self, file_path, include_headers, export_format="CSV"):
        """Export products to CSV or Excel"""
        self.status_label.setText("Exportando productos...")
        self.result_text.append("📦 Obteniendo productos...\n")
        
        try:
            products = self.core.engine.list_products_for_export()
            self.progress.setMaximum(len(products))
            
            if "Excel" in export_format:
                if not HAS_OPENPYXL:
                    self.result_text.append("❌ Error: Se requiere openpyxl para exportar a Excel.\n")
                    self.result_text.append("Instala con: pip install openpyxl")
                    self.status_label.setText("❌ Error")
                    return
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Productos"
                
                fieldnames = ['sku', 'name', 'price', 'stock', 'department', 'sat_clave_prod_serv', 'sat_clave_unidad']
                
                if include_headers:
                    headers = ['SKU', 'Nombre', 'Precio', 'Stock', 'Departamento', 'SAT Clave Prod/Serv', 'SAT Clave Unidad']
                    ws.append(headers)
                    
                    # Estilo encabezados
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")
                
                for i, product in enumerate(products):
                    row = [
                        product.get('sku', ''),
                        product.get('name', ''),
                        product.get('price', 0),
                        product.get('stock', 0),
                        product.get('department', ''),
                        product.get('sat_clave_prod_serv', '01010101'),
                        product.get('sat_clave_unidad', 'H87'),
                    ]
                    ws.append(row)
                    self.progress.setValue(i + 1)
                
                # Ajustar ancho de columnas
                for idx, col in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
                    max_length = 0
                    column_letter = get_column_letter(idx)
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except Exception as e:
                            logger.debug("Getting cell value length: %s", e)
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                ws.freeze_panes = "A2"
                wb.save(file_path)
            else:
                # CSV
                with open(file_path, 'w', newline='', encoding='utf-8') as f:
                    fieldnames = ['sku', 'name', 'price', 'stock', 'department', 'sat_clave_prod_serv', 'sat_clave_unidad']
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    
                    if include_headers:
                        writer.writeheader()
                    
                    for i, product in enumerate(products):
                        writer.writerow({
                            'sku': product.get('sku', ''),
                            'name': product.get('name', ''),
                            'price': product.get('price', 0),
                            'stock': product.get('stock', 0),
                            'department': product.get('department', ''),
                            'sat_clave_prod_serv': product.get('sat_clave_prod_serv', '01010101'),
                            'sat_clave_unidad': product.get('sat_clave_unidad', 'H87'),
                        })
                        self.progress.setValue(i + 1)
            
            self.result_text.append(f"✅ Exportados {len(products)} productos")
            self.result_text.append(f"📁 Archivo: {file_path}")
            self.status_label.setText("✅ Exportación completada")
            
        except Exception as e:
            self.result_text.append(f"❌ Error: {e}")
            self.status_label.setText("❌ Error en exportación")
    
    def export_customers(self, file_path, include_headers, export_format="CSV"):
        """Export customers to CSV or Excel"""
        self.status_label.setText("Exportando clientes...")
        self.result_text.append("👥 Obteniendo clientes...\n")
        
        try:
            customers = self.core.list_customers(limit=99999)
            self.progress.setMaximum(len(customers))
            
            if "Excel" in export_format:
                if not HAS_OPENPYXL:
                    self.result_text.append("❌ Error: Se requiere openpyxl para exportar a Excel.\n")
                    self.status_label.setText("❌ Error")
                    return
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Clientes"
                
                if include_headers:
                    headers = ['Nombre', 'Teléfono', 'Email', 'Límite Crédito']
                    ws.append(headers)
                    
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")
                
                for i, customer in enumerate(customers):
                    row = [
                        customer.get('name', ''),
                        customer.get('phone', ''),
                        customer.get('email', ''),
                        customer.get('credit_limit', 0)
                    ]
                    ws.append(row)
                    self.progress.setValue(i + 1)
                
                # Ajustar ancho de columnas
                for idx in range(1, 5):
                    column_letter = get_column_letter(idx)
                    ws.column_dimensions[column_letter].width = 20
                
                ws.freeze_panes = "A2"
                wb.save(file_path)
            else:
                # CSV
                with open(file_path, 'w', newline='', encoding='utf-8') as f:
                    fieldnames = ['name', 'phone', 'email', 'credit_limit']
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    
                    if include_headers:
                        writer.writeheader()
                    
                    for i, customer in enumerate(customers):
                        writer.writerow({
                            'name': customer.get('name', ''),
                            'phone': customer.get('phone', ''),
                            'email': customer.get('email', ''),
                            'credit_limit': customer.get('credit_limit', 0)
                        })
                        self.progress.setValue(i + 1)
            
            self.result_text.append(f"✅ Exportados {len(customers)} clientes")
            self.result_text.append(f"📁 Archivo: {file_path}")
            self.status_label.setText("✅ Exportación completada")
            
        except Exception as e:
            self.result_text.append(f"❌ Error: {e}")
            self.status_label.setText("❌ Error en exportación")
    
    def export_sales(self, file_path, include_headers, export_format="CSV"):
        """Export sales to CSV or Excel"""
        self.status_label.setText("Exportando ventas...")
        self.result_text.append("💰 Obteniendo ventas...\n")
        
        try:
            sales = self.core.get_sales(limit=10000)
            self.progress.setMaximum(len(sales))
            
            if "Excel" in export_format:
                if not HAS_OPENPYXL:
                    self.result_text.append("❌ Error: Se requiere openpyxl para exportar a Excel.\n")
                    self.status_label.setText("❌ Error")
                    return
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Ventas"
                
                if include_headers:
                    headers = ['ID', 'Fecha/Hora', 'Total', 'Método Pago', 'ID Turno']
                    ws.append(headers)
                    
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="2E75B5", end_color="2E75B5", fill_type="solid")
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")
                
                for i, sale in enumerate(sales):
                    row = [
                        sale.get('id', ''),
                        sale.get('timestamp', ''),
                        sale.get('total', 0),
                        sale.get('payment_method', ''),
                        sale.get('turn_id', '')
                    ]
                    ws.append(row)
                    self.progress.setValue(i + 1)
                
                # Ajustar ancho de columnas
                for idx in range(1, 6):
                    column_letter = get_column_letter(idx)
                    ws.column_dimensions[column_letter].width = 20
                
                ws.freeze_panes = "A2"
                wb.save(file_path)
            else:
                # CSV
                with open(file_path, 'w', newline='', encoding='utf-8') as f:
                    fieldnames = ['id', 'timestamp', 'total', 'payment_method', 'turn_id']
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    
                    if include_headers:
                        writer.writeheader()
                    
                    for i, sale in enumerate(sales):
                        writer.writerow({
                            'id': sale.get('id', ''),
                            'timestamp': sale.get('timestamp', ''),
                            'total': sale.get('total', 0),
                            'payment_method': sale.get('payment_method', ''),
                            'turn_id': sale.get('turn_id', '')
                        })
                        self.progress.setValue(i + 1)
            
            self.result_text.append(f"✅ Exportadas {len(sales)} ventas")
            self.result_text.append(f"📁 Archivo: {file_path}")
            self.status_label.setText("✅ Exportación completada")
            
        except Exception as e:
            self.result_text.append(f"❌ Error: {e}")
            self.status_label.setText("❌ Error en exportación")
