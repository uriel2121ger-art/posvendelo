"""
Import Wizard for Products and Customers
"""
import csv
import logging
import sys
from pathlib import Path

logger = logging.getLogger(__name__)


def debug_print(msg: str):
    """Print debug message to terminal and flush immediately."""
    print(f"[IMPORT] {msg}", file=sys.stderr, flush=True)
    logger.info(msg)

from PyQt6 import QtCore, QtWidgets

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Import debug logging utilities from path_utils
from app.utils.path_utils import agent_log_enabled, get_debug_log_path_str


class ImportWizard(QtWidgets.QWizard):
    """Wizard for importing products and customers from CSV"""
    
    def __init__(self, parent=None, core=None):
        super().__init__(parent)
        self.core = core
        self.setWindowTitle("Asistente de Importación")
        self.setWizardStyle(QtWidgets.QWizard.WizardStyle.ModernStyle)
        
        # Add pages
        self.addPage(IntroPage())
        self.addPage(FileSelectionPage())
        self.addPage(MappingPage())
        self.addPage(ImportPage(core=self.core))
        
        self.resize(600, 400)
        self._apply_theme()
    
    def showEvent(self, event):
        """Apply theme when shown."""
        super().showEvent(event)
        self._apply_theme()
    
    def _apply_theme(self):
        """Apply current theme colors."""
        try:
            from app.utils.theme_manager import theme_manager
            c = theme_manager.get_colors()
            self.setStyleSheet(f"""
                QWizard {{ background: {c['bg_secondary']}; }}
                QWizardPage {{ background: {c['bg_primary']}; }}
                QLabel {{ color: {c['text_primary']}; }}
                QLineEdit, QTextEdit, QComboBox {{
                    background: {c['bg_secondary']}; color: {c['text_primary']};
                    border: 1px solid {c['border']}; padding: 8px; border-radius: 4px;
                }}
                QPushButton {{
                    background: {c['btn_primary']}; color: white;
                    padding: 8px 16px; border-radius: 4px; font-weight: bold;
                }}
                QPushButton:hover {{ background: {c['btn_success']}; }}
                QProgressBar {{
                    border: 1px solid {c['border']}; border-radius: 4px;
                    background: {c['bg_secondary']}; text-align: center;
                }}
                QProgressBar::chunk {{ background: {c['btn_success']}; }}
            """)
        except Exception as e:
            logger.debug("Applying import wizard theme: %s", e)

class IntroPage(QtWidgets.QWizardPage):
    """Introduction page"""
    
    def __init__(self):
        super().__init__()
        self.setTitle("Importar Datos")
        self.setSubTitle("Importa productos, clientes o ventas desde archivos CSV")
        
        layout = QtWidgets.QVBoxLayout(self)
        
        info = QtWidgets.QLabel(
            "Este asistente te ayudará a importar:\n\n"
            "• Productos (SKU, nombre, precio, stock)\n"
            "• Clientes (nombre, teléfono, email)\n"
            "• Ventas Básicas (formato simple)\n"
            "• Ventas DETALLADAS (con productos por línea)\n\n"
            "Formatos soportados: CSV, Excel (.xlsx, .xls)"
        )
        layout.addWidget(info)
        
        # Import type
        self.type_combo = QtWidgets.QComboBox()
        self.type_combo.addItems([
            "Productos",
            "Clientes", 
            "Ventas (Formato Simple)",
            "Ventas DETALLADAS (con productos)",
            "Monedero Anónimo (puntos y transacciones)"
        ])
        
        type_layout = QtWidgets.QFormLayout()
        type_layout.addRow("Tipo de datos:", self.type_combo)
        layout.addLayout(type_layout)
        
        layout.addStretch()
        
        self.registerField("import_type", self.type_combo, "currentText")

class FileSelectionPage(QtWidgets.QWizardPage):
    """File selection page"""
    
    def __init__(self):
        super().__init__()
        self.setTitle("Seleccionar Archivo")
        self.setSubTitle("Selecciona el archivo CSV o Excel a importar")
        
        layout = QtWidgets.QVBoxLayout(self)
        
        # File selection
        file_layout = QtWidgets.QHBoxLayout()
        self.file_edit = QtWidgets.QLineEdit()
        self.file_edit.setPlaceholderText("Selecciona un archivo CSV o Excel...")
        
        browse_btn = QtWidgets.QPushButton("Examinar...")
        browse_btn.clicked.connect(self.browse_file)
        
        file_layout.addWidget(self.file_edit)
        file_layout.addWidget(browse_btn)
        layout.addLayout(file_layout)
        
        # Preview
        self.preview_label = QtWidgets.QLabel("Vista previa:")
        layout.addWidget(self.preview_label)
        
        self.preview_text = QtWidgets.QTextEdit()
        self.preview_text.setReadOnly(True)
        self.preview_text.setMaximumHeight(200)
        layout.addWidget(self.preview_text)
        
        self.registerField("file_path*", self.file_edit)
    
    def browse_file(self):
        """Browse for CSV or Excel file"""
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "Seleccionar Archivo", "", 
            "Archivos de Datos (*.csv *.xlsx *.xls);;CSV Files (*.csv);;Excel Files (*.xlsx *.xls)"
        )
        if file_path:
            self.file_edit.setText(file_path)
            self.load_preview(file_path)
    
    def load_preview(self, file_path):
        """Load file preview"""
        try:
            if file_path.lower().endswith(('.xlsx', '.xls')):
                if not HAS_OPENPYXL:
                    self.preview_text.setPlainText(
                        "⚠️ Se requiere openpyxl para leer archivos Excel.\n"
                        "Instala con: pip install openpyxl"
                    )
                    return
                
                # Validar archivo Excel
                import zipfile
                import os
                
                if not os.path.exists(file_path):
                    self.preview_text.setPlainText(f"❌ Error: Archivo no encontrado")
                    return
                
                if os.path.getsize(file_path) == 0:
                    self.preview_text.setPlainText(f"❌ Error: Archivo vacío")
                    return
                
                # Validar que .xlsx sea un ZIP válido
                if file_path.lower().endswith('.xlsx'):
                    try:
                        with zipfile.ZipFile(file_path, 'r') as zip_ref:
                            zip_files = zip_ref.namelist()
                            if not any('xl/workbook.xml' in f or '[Content_Types].xml' in f for f in zip_files):
                                self.preview_text.setPlainText("❌ Error: Archivo no parece ser un Excel válido")
                                return
                    except zipfile.BadZipFile:
                        self.preview_text.setPlainText(
                            "❌ Error: Archivo Excel corrupto.\n"
                            "Abre el archivo en Excel y guárdalo nuevamente."
                        )
                        return
                
                # Leer primeras filas de Excel
                wb = None
                try:
                    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                    ws = wb.active
                    preview_lines = []
                    for i, row in enumerate(ws.iter_rows(values_only=True, max_row=10)):
                        if i == 0:
                            preview_lines.append("Encabezados:")
                        row_str = " | ".join([str(cell) if cell is not None else "" for cell in row[:5]])
                        preview_lines.append(row_str)
                        if i >= 5:  # Mostrar solo 5 filas
                            break
                    self.preview_text.setPlainText("\n".join(preview_lines) + "\n...")
                except Exception as e:
                    self.preview_text.setPlainText(f"❌ Error leyendo Excel: {e}")
                finally:
                    if wb is not None:
                        wb.close()
            else:
                # CSV
                with open(file_path, 'r', encoding='utf-8') as f:
                    preview = f.read(500)  # First 500 chars
                self.preview_text.setPlainText(preview + "\n...")
        except Exception as e:
            self.preview_text.setPlainText(f"❌ Error: {e}")

class MappingPage(QtWidgets.QWizardPage):
    """Column mapping page with manual selection"""

    # Field definitions per import type
    PRODUCT_FIELDS = [
        ('sku', 'SKU / Código', True, ['sku', 'SKU', 'codigo', 'Codigo', 'code', 'Code', 'barcode']),
        ('name', 'Nombre', True, ['name', 'nombre', 'Name', 'producto', 'Producto', 'description']),
        ('price', 'Precio', True, ['price', 'precio', 'Price', 'cost', 'costo', 'precio_venta']),
        ('stock', 'Stock', False, ['stock', 'Stock', 'cantidad', 'qty', 'inventory', 'existencia']),
        ('category', 'Categoría', False, ['category', 'categoria', 'dept', 'department', 'departamento']),
        ('sat_clave', 'Clave SAT', False, ['sat_clave_prod_serv', 'clave_sat', 'sat_code', 'clave_producto']),
    ]

    CUSTOMER_FIELDS = [
        ('name', 'Nombre', True, ['name', 'nombre', 'Name', 'cliente', 'customer']),
        ('phone', 'Teléfono', False, ['phone', 'telefono', 'Phone', 'tel', 'celular', 'mobile']),
        ('email', 'Email', False, ['email', 'Email', 'correo', 'e-mail']),
        ('rfc', 'RFC', False, ['rfc', 'RFC', 'tax_id']),
        ('address', 'Dirección', False, ['address', 'direccion', 'domicilio', 'street']),
    ]

    # FIX 2026-01-31: Ningún campo es requerido - el código tiene fallbacks
    # Esto permite importar archivos con formatos variados
    SALES_FIELDS = [
        # Identificación de venta
        ('folio', 'Folio', False, ['FOLIO', 'Folio', 'folio', 'NUMERO', 'Numero', 'numero', 'TICKET', 'ticket']),
        ('serie', 'Serie', False, ['SERIE', 'Serie', 'serie', 'type', 'tipo']),
        ('date', 'Fecha', False, ['FECHA', 'Fecha', 'fecha', 'date', 'timestamp', 'created_at']),
        ('time', 'Hora', False, ['HORA', 'Hora', 'hora', 'time']),
        # Producto
        ('sku', 'SKU Producto', False, ['SKU', 'Sku', 'sku', 'CODIGO', 'Codigo', 'codigo', 'CÓDIGO_BARRAS', 'barcode']),
        ('product_name', 'Nombre Producto', False, ['PRODUCTO', 'Producto', 'producto', 'product', 'name', 'nombre']),
        ('quantity', 'Cantidad', False, ['CANTIDAD', 'Cantidad', 'cantidad', 'qty', 'quantity', 'units']),
        ('unit_price', 'Precio Unitario', False, ['PRECIO_UNITARIO', 'Precio_Unitario', 'precio_unitario', 'price', 'precio']),
        ('discount', 'Descuento', False, ['DESCUENTO', 'Descuento', 'descuento', 'discount']),
        ('line_total', 'Subtotal Línea', False, ['SUBTOTAL_PRODUCTO', 'Subtotal', 'subtotal', 'line_total']),
        # Claves SAT (CRÍTICO para facturación)
        ('sat_clave_prod', 'Clave SAT Producto', False, ['CLAVE_SAT', 'Clave_SAT', 'clave_sat', 'SAT_CLAVE_PROD_SERV', 'sat_clave_prod_serv', 'sat_code', 'clave_producto_sat']),
        ('sat_clave_unidad', 'Clave SAT Unidad', False, ['CLAVE_UNIDAD_SAT', 'Clave_Unidad_SAT', 'clave_unidad_sat', 'SAT_CLAVE_UNIDAD', 'sat_clave_unidad', 'sat_unit', 'unidad_sat']),
        ('sat_descripcion', 'Descripción SAT', False, ['DESCRIPCION_SAT', 'Descripcion_SAT', 'descripcion_sat', 'SAT_DESCRIPCION', 'sat_descripcion']),
        # Totales y pago
        ('sale_total', 'Total Venta', False, ['TOTAL_VENTA', 'Total_Venta', 'total_venta', 'total', 'Total']),
        ('payment', 'Método Pago', False, ['MÉTODO_PAGO', 'Metodo_Pago', 'metodo_pago', 'payment_method', 'payment']),
        # Cliente
        ('customer', 'Cliente', False, ['CLIENTE', 'Cliente', 'cliente', 'customer', 'customer_name']),
        ('customer_rfc', 'RFC Cliente', False, ['RFC_CLIENTE', 'Rfc_Cliente', 'rfc_cliente', 'rfc', 'RFC']),
        # Origen
        ('user', 'Usuario', False, ['USUARIO', 'Usuario', 'usuario', 'user', 'cajero', 'cashier']),
        ('branch', 'Sucursal', False, ['SUCURSAL', 'Sucursal', 'sucursal', 'branch', 'tienda']),
        ('terminal', 'Terminal/PC', False, ['TERMINAL', 'Terminal', 'terminal', 'PC', 'pc', 'terminal_id', 'pos_id', 'ORIGEN', 'Origen', 'origen']),
        ('notes', 'Notas', False, ['NOTAS', 'Notas', 'notas', 'notes', 'comments', 'observaciones']),
    ]

    def __init__(self):
        super().__init__()
        self.setTitle("Mapeo de Columnas")
        self.setSubTitle("Selecciona qué columna corresponde a cada campo")

        self.column_headers = []
        self.combo_boxes = {}

        main_layout = QtWidgets.QVBoxLayout(self)

        # Info label
        self.info_label = QtWidgets.QLabel(
            "💡 Selecciona la columna de tu archivo que corresponde a cada campo.\n"
            "Los campos marcados con * son obligatorios."
        )
        self.info_label.setStyleSheet("margin-bottom: 10px;")
        main_layout.addWidget(self.info_label)

        # Scroll area for mapping fields
        scroll = QtWidgets.QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setMaximumHeight(250)

        self.mapping_widget = QtWidgets.QWidget()
        self.mapping_layout = QtWidgets.QFormLayout(self.mapping_widget)
        self.mapping_layout.setSpacing(8)
        scroll.setWidget(self.mapping_widget)
        main_layout.addWidget(scroll)

        # Sales import type options
        self.sales_options_frame = QtWidgets.QFrame()
        sales_layout = QtWidgets.QVBoxLayout(self.sales_options_frame)
        sales_layout.setContentsMargins(0, 15, 0, 0)

        sales_label = QtWidgets.QLabel("📊 Tipo de importación de ventas:")
        sales_label.setStyleSheet("font-weight: bold;")
        sales_layout.addWidget(sales_label)

        self.rb_real = QtWidgets.QRadioButton("✅ Ventas REALES (respaldos, migración)")
        self.rb_simulated = QtWidgets.QRadioButton("🧪 Ventas SIMULADAS (testing)")
        self.rb_real.setChecked(True)

        sales_layout.addWidget(self.rb_real)
        sales_layout.addWidget(self.rb_simulated)

        main_layout.addWidget(self.sales_options_frame)
        self.sales_options_frame.hide()

        # Status label
        self.status_label = QtWidgets.QLabel("")
        self.status_label.setStyleSheet("margin-top: 10px; font-weight: bold;")
        main_layout.addWidget(self.status_label)

        main_layout.addStretch()

        # FIX 2026-01-31: Registrar checkbox con propiedad y señal correctas (PyQt6)
        self.registerField("sales_simulated", self.rb_simulated, "checked", self.rb_simulated.toggled)

    def _get_fields_for_type(self, import_type: str):
        """Get field definitions based on import type"""
        if "Producto" in import_type:
            return self.PRODUCT_FIELDS
        elif "Cliente" in import_type:
            return self.CUSTOMER_FIELDS
        elif "Ventas" in import_type:
            return self.SALES_FIELDS
        return self.PRODUCT_FIELDS

    def _auto_detect_column(self, possible_names: list) -> str:
        """Try to auto-detect the best matching column"""
        for name in possible_names:
            # Exact match (case insensitive)
            for header in self.column_headers:
                if header.lower() == name.lower():
                    return header
        # Partial match
        for name in possible_names:
            for header in self.column_headers:
                if name.lower() in header.lower():
                    return header
        return ""

    def _read_file_headers(self, file_path: str) -> list:
        """Read column headers from file"""
        import zipfile
        headers = []

        try:
            if file_path.lower().endswith(('.xlsx', '.xls')):
                if HAS_OPENPYXL:
                    # Validate Excel file
                    try:
                        with zipfile.ZipFile(file_path, 'r') as zf:
                            pass
                    except zipfile.BadZipFile:
                        return []

                    wb = None
                    try:
                        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                        ws = wb.active
                        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                        if first_row:
                            headers = [str(cell) if cell else f"Col_{i}" for i, cell in enumerate(first_row)]
                    finally:
                        if wb:
                            wb.close()
            else:
                # CSV
                encodings = ['utf-8', 'latin-1', 'cp1252']
                for enc in encodings:
                    try:
                        with open(file_path, 'r', encoding=enc) as f:
                            reader = csv.reader(f)
                            first_row = next(reader, None)
                            if first_row:
                                headers = [str(cell) if cell else f"Col_{i}" for i, cell in enumerate(first_row)]
                            break
                    except UnicodeDecodeError:
                        continue
        except Exception as e:
            logger.error(f"Error reading headers: {e}")

        return headers

    def initializePage(self):
        """Build mapping UI based on import type and file headers"""
        import_type = self.field("import_type")
        file_path = self.field("file_path")

        # Show/hide sales options
        if "Ventas" in str(import_type):
            self.sales_options_frame.show()
        else:
            self.sales_options_frame.hide()

        # Clear previous mappings
        while self.mapping_layout.count():
            item = self.mapping_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        self.combo_boxes.clear()

        # Read file headers
        self.column_headers = self._read_file_headers(file_path)

        if not self.column_headers:
            self.status_label.setText("⚠️ No se pudieron leer las columnas del archivo")
            self.status_label.setStyleSheet("color: orange; font-weight: bold;")
            return

        # Get fields for this import type
        fields = self._get_fields_for_type(str(import_type))

        # Create combo boxes for each field
        for field_id, field_label, required, possible_names in fields:
            label_text = f"{field_label} {'*' if required else ''}"
            label = QtWidgets.QLabel(label_text)
            if required:
                label.setStyleSheet("font-weight: bold;")

            combo = QtWidgets.QComboBox()
            combo.addItem("-- No importar --", "")
            for header in self.column_headers:
                combo.addItem(header, header)

            # Auto-detect best match
            auto_match = self._auto_detect_column(possible_names)
            if auto_match:
                index = combo.findData(auto_match)
                if index >= 0:
                    combo.setCurrentIndex(index)

            self.combo_boxes[field_id] = combo
            self.mapping_layout.addRow(label, combo)

        # Update status
        self.status_label.setText(f"✅ {len(self.column_headers)} columnas detectadas")
        self.status_label.setStyleSheet("color: green; font-weight: bold;")

    def get_mapping(self) -> dict:
        """Get the current column mapping"""
        mapping = {}
        for field_id, combo in self.combo_boxes.items():
            selected = combo.currentData()
            if selected:
                mapping[field_id] = selected
        return mapping

    def validatePage(self) -> bool:
        """Validate required fields are mapped"""
        try:
            import_type = self.field("import_type")
            fields = self._get_fields_for_type(str(import_type))

            missing = []
            for field_id, field_label, required, _ in fields:
                if required:
                    combo = self.combo_boxes.get(field_id)
                    if not combo or not combo.currentData():
                        missing.append(field_label)

            if missing:
                QtWidgets.QMessageBox.warning(
                    self,
                    "Campos Requeridos",
                    f"Por favor mapea los siguientes campos obligatorios:\n• " + "\n• ".join(missing)
                )
                return False

            # Store mapping in wizard for use by ImportPage
            wiz = self.wizard()
            if wiz:
                wiz.column_mapping = self.get_mapping()
                logger.info(f"Column mapping stored: {wiz.column_mapping}")
            else:
                logger.warning("Wizard reference not available, mapping not stored")

            return True

        except Exception as e:
            import traceback
            error_msg = f"Error en validación: {e}\n{traceback.format_exc()}"
            logger.error(error_msg)
            QtWidgets.QMessageBox.critical(
                self,
                "Error de Validación",
                f"Ocurrió un error al validar:\n{e}"
            )
            return False

class ImportPage(QtWidgets.QWizardPage):
    """Import execution page"""

    def __init__(self, core=None):
        super().__init__()
        self.core = core
        self.setTitle("Importando...")
        self.setSubTitle("Procesando archivo")
        self._import_running = False

        layout = QtWidgets.QVBoxLayout(self)

        # Gran etiqueta de estado visible
        self.status_label = QtWidgets.QLabel("⏳ PREPARANDO IMPORTACIÓN...")
        self.status_label.setStyleSheet("""
            QLabel {
                font-size: 18px;
                font-weight: bold;
                color: #2196F3;
                padding: 15px;
                background: #E3F2FD;
                border-radius: 8px;
                margin-bottom: 10px;
            }
        """)
        self.status_label.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.status_label)

        # Contador de registros
        self.counter_label = QtWidgets.QLabel("Registros: 0 / 0")
        self.counter_label.setStyleSheet("""
            QLabel {
                font-size: 14px;
                font-weight: bold;
                padding: 5px;
            }
        """)
        self.counter_label.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.counter_label)

        # Barra de progreso más grande
        self.progress = QtWidgets.QProgressBar()
        self.progress.setMinimumHeight(30)
        self.progress.setStyleSheet("""
            QProgressBar {
                border: 2px solid #2196F3;
                border-radius: 5px;
                text-align: center;
                font-weight: bold;
                font-size: 12px;
            }
            QProgressBar::chunk {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #4CAF50, stop:1 #8BC34A);
            }
        """)
        self.progress.setFormat("%v / %m (%p%)")
        layout.addWidget(self.progress)

        self.log_text = QtWidgets.QTextEdit()
        self.log_text.setReadOnly(True)
        layout.addWidget(self.log_text)

        self.setCommitPage(True)
        self.setButtonText(QtWidgets.QWizard.WizardButton.CommitButton, "Importar")

    def update_status(self, message: str, color: str = "#2196F3", bg: str = "#E3F2FD"):
        """Update the status label with visual feedback"""
        self.status_label.setText(message)
        self.status_label.setStyleSheet(f"""
            QLabel {{
                font-size: 18px;
                font-weight: bold;
                color: {color};
                padding: 15px;
                background: {bg};
                border-radius: 8px;
                margin-bottom: 10px;
            }}
        """)
        QtWidgets.QApplication.processEvents()

    def update_counter(self, current: int, total: int):
        """Update the counter label"""
        self.counter_label.setText(f"Registros: {current:,} / {total:,}")
        QtWidgets.QApplication.processEvents()
    
    def initializePage(self):
        """Start import when page is shown"""
        # #region agent log
        if agent_log_enabled():
            import json, time
            try:
                from app.utils.path_utils import get_debug_log_path_str
                with open(get_debug_log_path_str(), "a") as f:
                    f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"IMPORT_WIZARD","location":"import_wizard.py:initializePage","message":"Import wizard page initialized","data":{"core_db_is_none":self.core.db is None if self.core else None},"timestamp":int(time.time()*1000)})+"\n")
            except Exception as e: logger.debug("Writing import wizard init log: %s", e)
        # #endregion

        # Verificar que core y db estén disponibles
        if not self.core:
            self.log_text.append("❌ Error: Core no disponible")
            return
        
        if not self.core.db:
            # #region agent log
            if agent_log_enabled():
                import json, time
                try:
                    from app.utils.path_utils import get_debug_log_path_str
                    with open(get_debug_log_path_str(), "a") as f:
                        f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"IMPORT_WIZARD","location":"import_wizard.py:initializePage","message":"DB is None, attempting to initialize","data":{},"timestamp":int(time.time()*1000)})+"\n")
                except Exception as e: logger.debug("Writing DB init attempt log: %s", e)
            # #endregion
            # Intentar inicializar la base de datos
            if not self.core.initialize_database():
                self.log_text.append("❌ Error: No se pudo conectar a la base de datos.\n")
                self.log_text.append("Verifica la configuración de PostgreSQL.")
                return
        
        file_path = self.field("file_path")
        import_type = self.field("import_type")
        
        # #region agent log
        if agent_log_enabled():
            import json, time
            try:
                from app.utils.path_utils import get_debug_log_path_str
                with open(get_debug_log_path_str(), "a") as f:
                    f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"IMPORT_WIZARD","location":"import_wizard.py:initializePage","message":"Starting import","data":{"import_type":import_type,"file_path":file_path},"timestamp":int(time.time()*1000)})+"\n")
            except Exception as e: logger.debug("Writing starting import log: %s", e)
        # #endregion
        
        if import_type == "Productos":
            self.import_products(file_path)
        elif import_type == "Clientes":
            self.import_customers(file_path)
        elif "DETALLADAS" in import_type:
            self.import_sales_detailed(file_path)
        elif "Ventas" in import_type:
            self.import_sales(file_path)
        elif "Monedero" in import_type:
            self.import_anonymous_wallet(file_path)
    
    def import_products(self, file_path):
        """Import products from CSV or Excel"""
        # #region agent log
        if agent_log_enabled():
            import json, time
            try:
                from app.utils.path_utils import get_debug_log_path_str
                with open(get_debug_log_path_str(), "a") as f:
                    f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"IMPORT_WIZARD","location":"import_wizard.py:import_products","message":"Starting product import","data":{"file_path":file_path},"timestamp":int(time.time()*1000)})+"\n")
            except Exception as e: logger.debug("Writing starting product import log: %s", e)
        # #endregion
        
        self.log_text.append("📦 Importando productos...\n")
        
        # Verificar que core y engine estén disponibles
        if not self.core or not self.core.engine:
            self.log_text.append("❌ Error: Core o engine no disponible")
            return
        
        try:
            rows = []
            if file_path.lower().endswith(('.xlsx', '.xls')):
                if not HAS_OPENPYXL:
                    self.log_text.append("❌ Error: Se requiere openpyxl para importar Excel.\n")
                    self.log_text.append("Instala con: pip install openpyxl")
                    return
                
                # Validar archivo Excel
                import zipfile
                import os
                
                if not os.path.exists(file_path):
                    self.log_text.append(f"❌ Error: Archivo no encontrado: {file_path}")
                    return
                
                if os.path.getsize(file_path) == 0:
                    self.log_text.append("❌ Error: Archivo vacío")
                    return
                
                # Validar que .xlsx sea un ZIP válido
                if file_path.lower().endswith('.xlsx'):
                    try:
                        with zipfile.ZipFile(file_path, 'r') as zip_ref:
                            zip_files = zip_ref.namelist()
                            if not any('xl/workbook.xml' in f or '[Content_Types].xml' in f for f in zip_files):
                                self.log_text.append("❌ Error: Archivo no parece ser un Excel válido")
                                return
                    except zipfile.BadZipFile:
                        self.log_text.append("❌ Error: Archivo Excel corrupto. Abre el archivo en Excel y guárdalo nuevamente.")
                        return
                
                # Leer Excel
                wb = None
                try:
                    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                    ws = wb.active

                    # Leer encabezados de la primera fila
                    headers = []
                    # FIX CRÍTICO 2026-01-30: ws[1] NO funciona en read_only mode
                    # Usar iter_rows en su lugar
                    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                    if first_row:
                        headers = [str(cell) if cell is not None else "" for cell in first_row]
                    else:
                        self.log_text.append("❌ Error: No se encontró fila de encabezados")
                        return

                    # Leer datos
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        row_dict = {}
                        for i, cell_value in enumerate(row):
                            if i < len(headers):
                                row_dict[headers[i]] = str(cell_value) if cell_value is not None else ""
                        if any(row_dict.values()):  # Solo agregar si tiene datos
                            rows.append(row_dict)

                    self.log_text.append(f"✅ Archivo Excel leído: {len(rows)} filas\n")
                except Exception as e:
                    import traceback
                    self.log_text.append(f"❌ Error leyendo Excel: {e}")
                    self.log_text.append(f"   Detalle: {traceback.format_exc()[:500]}")
                    logger.error("Error leyendo Excel: %s", traceback.format_exc())
                    self.update_status("❌ ERROR LEYENDO EXCEL", "#f44336", "#FFEBEE")
                    self.update_counter(0, 0)
                    return
                finally:
                    if wb is not None:
                        wb.close()
            else:
                # CSV
                with open(file_path, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    rows = list(reader)
            
            if not rows:
                self.log_text.append("❌ Error: Archivo vacío o sin datos")
                return

            self.progress.setMaximum(len(rows))
            self.update_counter(0, len(rows))  # FIX 2026-02-01: Inicializar contador
            imported = 0
            updated = 0
            errors = 0
            skipped = 0

            # Get column mapping from wizard
            mapping = getattr(self.wizard(), 'column_mapping', {})

            # FIX 2026-02-04: Pre-load all existing SKUs to avoid N+1 queries
            existing_skus = set()
            if self.core.db:
                existing_result = self.core.db.execute_query("SELECT sku FROM products")
                existing_skus = {r['sku'] for r in existing_result if r.get('sku')}

            for i, row in enumerate(rows):
                try:
                    # Use mapped columns or fallback to auto-detection
                    sku_col = mapping.get('sku')
                    name_col = mapping.get('name')
                    price_col = mapping.get('price')
                    stock_col = mapping.get('stock')
                    cat_col = mapping.get('category')
                    sat_col = mapping.get('sat_clave')

                    # FIX 2026-01-31: Helper para buscar en múltiples variantes de columna
                    def get_col(row, mapped_col, fallbacks):
                        if mapped_col and mapped_col in row:
                            return row.get(mapped_col)
                        for fb in fallbacks:
                            if fb in row:
                                return row.get(fb)
                        return None

                    product_data = {
                        'sku': get_col(row, sku_col, ['sku', 'SKU', 'Sku', 'codigo', 'CODIGO', 'Codigo', 'code', 'CODE', 'barcode', 'BARCODE']) or f'IMP-{i}',
                        'name': get_col(row, name_col, ['name', 'nombre', 'NOMBRE', 'Name', 'Nombre', 'producto', 'PRODUCTO', 'Producto', 'description', 'descripcion']) or '',
                        'price': float(get_col(row, price_col, ['price', 'precio', 'PRECIO', 'Price', 'Precio', 'precio_venta', 'PRECIO_VENTA', 'unit_price']) or 0),
                        'stock': float(get_col(row, stock_col, ['stock', 'Stock', 'STOCK', 'cantidad', 'CANTIDAD', 'qty', 'existencia', 'EXISTENCIA']) or 0),
                        # SAT Catalog fields for CFDI 4.0
                        'sat_clave_prod_serv': get_col(row, sat_col, ['sat_clave_prod_serv', 'clave_sat', 'CLAVE_SAT', 'SAT_CODE', 'sat_code']) or '01010101',
                        'sat_clave_unidad': get_col(row, None, ['sat_clave_unidad', 'unidad_sat', 'UNIDAD_SAT', 'sat_unit']) or 'H87',
                    }

                    # Agregar costo si existe
                    cost_val = get_col(row, None, ['cost', 'costo', 'COSTO', 'Cost', 'Costo', 'cost_price', 'precio_costo'])
                    if cost_val:
                        product_data['cost'] = float(cost_val)

                    # Add category if mapped
                    if cat_col and row.get(cat_col):
                        product_data['category'] = row.get(cat_col)

                    # Validar datos mínimos
                    if not product_data['name']:
                        self.log_text.append(f"⚠️ Fila {i+1}: Sin nombre, usando SKU")
                        product_data['name'] = product_data['sku']

                    # FIX 2026-01-30: Verificar si SKU ya existe (SKU es UNIQUE)
                    # FIX 2026-02-04: Use pre-loaded SKUs set instead of N+1 queries
                    sku = product_data['sku']
                    if sku in existing_skus:
                        # Actualizar producto existente en lugar de fallar
                        self.core.db.execute_write(
                            """UPDATE products SET name = %s, price = %s,
                               sat_clave_prod_serv = %s, sat_clave_unidad = %s, synced = 0,
                               updated_at = CURRENT_TIMESTAMP
                               WHERE sku = %s""",
                            (product_data['name'], product_data['price'],
                             product_data['sat_clave_prod_serv'], product_data['sat_clave_unidad'], sku)
                        )
                        updated += 1
                        if len(rows) < 50:
                            self.log_text.append(f"🔄 {product_data['name']} (actualizado)")
                    else:
                        self.core.engine.create_product(product_data)
                        existing_skus.add(sku)  # FIX 2026-02-04: Track newly inserted SKU
                        imported += 1
                        if len(rows) < 50:
                            self.log_text.append(f"✅ {product_data['name']}")
                    
                except Exception as e:
                    errors += 1
                    self.log_text.append(f"❌ Fila {i+1}: {e}")
                    # #region agent log
                    if agent_log_enabled():
                        import json, time
                        try:
                            from app.utils.path_utils import get_debug_log_path_str
                            with open(get_debug_log_path_str(), "a") as f:
                                f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"IMPORT_WIZARD","location":"import_wizard.py:import_products","message":"Error importing product","data":{"row":i+1,"error":str(e)},"timestamp":int(time.time()*1000)})+"\n")
                        except Exception as e2: logger.debug("Writing product import error log: %s", e2)
                    # #endregion
                
                self.progress.setValue(i + 1)
                self.update_counter(i + 1, len(rows))  # FIX 2026-02-01: Actualizar contador
                QtWidgets.QApplication.processEvents()

            self.log_text.append(f"\n🎉 Importados: {imported} nuevos, {updated} actualizados / {len(rows)} filas")
            if errors > 0:
                self.log_text.append(f"⚠️ Errores: {errors}")

            # FIX 2026-01-31: Reset sequence after product import to avoid duplicate key errors
            try:
                self.core.db.execute_write(
                    "SELECT setval('products_id_seq', (SELECT COALESCE(MAX(id), 0) + 1 FROM products), false)"
                )
                self.log_text.append("✅ Secuencia de productos sincronizada")
                debug_print("Secuencia products_id_seq reseteada correctamente")
            except Exception as seq_err:
                self.log_text.append(f"⚠️ No se pudo sincronizar secuencia: {seq_err}")
                debug_print(f"Error reseteando secuencia: {seq_err}")

        except FileNotFoundError:
            self.log_text.append(f"\n❌ Error: Archivo no encontrado: {file_path}")
        except Exception as e:
            self.log_text.append(f"\n❌ Error: {e}")
            # #region agent log
            if agent_log_enabled():
                import json, time
                try:
                    from app.utils.path_utils import get_debug_log_path_str
                    with open(get_debug_log_path_str(), "a") as f:
                        f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"IMPORT_WIZARD","location":"import_wizard.py:import_products","message":"Fatal error in import","data":{"error":str(e)},"timestamp":int(time.time()*1000)})+"\n")
                except Exception as e2: logger.debug("Writing fatal import error log: %s", e2)
            # #endregion
    
    def import_customers(self, file_path):
        """Import customers from CSV or Excel"""
        self.log_text.append("👥 Importando clientes...\n")
        
        try:
            rows = []
            if file_path.lower().endswith(('.xlsx', '.xls')):
                if not HAS_OPENPYXL:
                    self.log_text.append("❌ Error: Se requiere openpyxl para importar Excel.\n")
                    return
                
                # Validar y leer Excel (similar a import_products)
                import zipfile
                import os
                
                if not os.path.exists(file_path) or os.path.getsize(file_path) == 0:
                    self.log_text.append("❌ Error: Archivo no válido")
                    return
                
                if file_path.lower().endswith('.xlsx'):
                    try:
                        with zipfile.ZipFile(file_path, 'r') as zip_ref:
                            zip_files = zip_ref.namelist()
                            if not any('xl/workbook.xml' in f or '[Content_Types].xml' in f for f in zip_files):
                                self.log_text.append("❌ Error: Archivo no parece ser un Excel válido")
                                return
                    except zipfile.BadZipFile:
                        self.log_text.append("❌ Error: Archivo Excel corrupto")
                        return
                
                wb = None
                try:
                    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                    ws = wb.active

                    headers = []
                    # FIX CRÍTICO 2026-01-30: ws[1] NO funciona en read_only mode
                    # Usar iter_rows en su lugar
                    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                    if first_row:
                        headers = [str(cell) if cell is not None else "" for cell in first_row]
                    else:
                        self.log_text.append("❌ Error: No se encontró fila de encabezados")
                        return

                    for row in ws.iter_rows(min_row=2, values_only=True):
                        row_dict = {}
                        for i, cell_value in enumerate(row):
                            if i < len(headers):
                                row_dict[headers[i]] = str(cell_value) if cell_value is not None else ""
                        if any(row_dict.values()):
                            rows.append(row_dict)

                    self.log_text.append(f"✅ Archivo Excel leído: {len(rows)} filas\n")
                except Exception as e:
                    import traceback
                    self.log_text.append(f"❌ Error leyendo Excel: {e}")
                    self.log_text.append(f"   Detalle: {traceback.format_exc()[:500]}")
                    logger.error("Error leyendo Excel: %s", traceback.format_exc())
                    self.update_status("❌ ERROR LEYENDO EXCEL", "#f44336", "#FFEBEE")
                    self.update_counter(0, 0)
                    return
                finally:
                    if wb is not None:
                        wb.close()
            else:
                # CSV - FIX 2026-01-30: Soporte para múltiples encodings
                encodings_to_try = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
                for encoding in encodings_to_try:
                    try:
                        with open(file_path, 'r', encoding=encoding) as f:
                            reader = csv.DictReader(f)
                            rows = list(reader)
                        self.log_text.append(f"✅ CSV leído con encoding {encoding}: {len(rows)} filas\n")
                        break
                    except UnicodeDecodeError:
                        continue
                    except Exception as e:
                        self.log_text.append(f"⚠️ Error con encoding {encoding}: {e}\n")
                        continue

            if not rows:
                self.log_text.append("❌ Error: Archivo vacío o sin datos")
                return

            self.progress.setMaximum(len(rows))
            self.update_counter(0, len(rows))  # FIX 2026-02-01: Inicializar contador
            imported = 0
            skipped_duplicates = 0

            # Get column mapping from wizard
            mapping = getattr(self.wizard(), 'column_mapping', {})

            for i, row in enumerate(rows):
                try:
                    # Use mapped columns or fallback to auto-detection
                    name_col = mapping.get('name')
                    phone_col = mapping.get('phone')
                    email_col = mapping.get('email')
                    rfc_col = mapping.get('rfc')
                    addr_col = mapping.get('address')

                    customer_data = {
                        'name': row.get(name_col, '') if name_col else row.get('name', row.get('nombre', row.get('Name', ''))),
                        'phone': row.get(phone_col, '') if phone_col else row.get('phone', row.get('telefono', '')),
                        'email': row.get(email_col, '') if email_col else row.get('email', row.get('Email', '')),
                        'rfc': row.get(rfc_col, '') if rfc_col else row.get('rfc', row.get('RFC', '')),
                    }

                    # Add address if mapped
                    if addr_col and row.get(addr_col):
                        customer_data['address'] = row.get(addr_col)

                    # FIX 2026-01-30: Verificar duplicados por RFC o nombre+teléfono
                    rfc = customer_data.get('rfc', '').strip()
                    name = customer_data.get('name', '').strip()
                    phone = customer_data.get('phone', '').strip()

                    if rfc and rfc.upper() not in ('', 'XAXX010101000', 'NONE', 'NULL'):
                        existing = list(self.core.db.execute_query(
                            "SELECT id FROM customers WHERE rfc = %s", (rfc,)
                        ))
                        if existing:
                            skipped_duplicates += 1
                            if skipped_duplicates <= 3:
                                self.log_text.append(f"⚠️ Fila {i+1}: RFC {rfc} ya existe, saltando")
                            continue
                    elif name and phone:
                        existing = list(self.core.db.execute_query(
                            "SELECT id FROM customers WHERE name = %s AND phone = %s",
                            (name, phone)
                        ))
                        if existing:
                            skipped_duplicates += 1
                            if skipped_duplicates <= 3:
                                self.log_text.append(f"⚠️ Fila {i+1}: Cliente {name} ({phone}) ya existe, saltando")
                            continue

                    self.core.engine.create_customer(customer_data)
                    imported += 1
                    if len(rows) < 50:
                        self.log_text.append(f"✅ {customer_data['name']}")

                except Exception as e:
                    self.log_text.append(f"❌ Fila {i+1}: {e}")

                self.progress.setValue(i + 1)
                self.update_counter(i + 1, len(rows))  # FIX 2026-02-01: Actualizar contador
                QtWidgets.QApplication.processEvents()

            self.log_text.append(f"\n🎉 Importados: {imported}/{len(rows)} clientes")
            if skipped_duplicates > 0:
                self.log_text.append(f"⏭️ Duplicados saltados: {skipped_duplicates}")
            
        except Exception as e:
            self.log_text.append(f"\n❌ Error: {e}")

    def import_sales(self, file_path):
        """
        Import sales from CSV.
        
        Status depends on user selection:
        - 'completed' for real sales (backups, migration)
        - 'simulated' for testing/demos
        
        IMPORTANT: Neither affects inventory!
        
        Expected CSV columns:
        - fecha/timestamp: Date and time of sale
        - total: Total amount
        - metodo/payment_method: Payment method (cash, card, etc)
        - items: Comma-separated item names or JSON
        - customer_id: (optional) Customer ID
        """
        # Get import type from wizard field
        is_simulated = self.field("sales_simulated")
        status = "simulated" if is_simulated else "completed"
        status_label = "SIMULADAS" if is_simulated else "REALES"
        note_text = "Importado para simulación" if is_simulated else "Importado desde respaldo"

        debug_print(f"=== INICIO IMPORTACIÓN SIMPLE ({status_label}) ===")
        self.log_text.append(f"💰 Importando ventas {status_label}...\n")
        self.log_text.append(f"📋 Estado: '{status}'\n")
        self.log_text.append("⚠️ Las ventas NO afectarán inventario\n\n")

        # FIX 2026-02-01: Inicializar contador al inicio para mostrar "procesando..."
        self.update_status("📂 LEYENDO ARCHIVO...", "#2196F3", "#E3F2FD")
        self.progress.setMaximum(0)  # Modo indeterminado mientras lee
        QtWidgets.QApplication.processEvents()

        try:
            from datetime import datetime
            import json
            import uuid as uuid_module

            from app.core import STATE

            # Get current context for proper integration
            user_id = STATE.user_id if STATE.user_id else 1
            branch_id = STATE.branch_id if STATE.branch_id else 1
            turn_id = STATE.current_turn_id if hasattr(STATE, 'current_turn_id') and STATE.current_turn_id else None

            # FIX 2026-01-30: Validar que user_id existe en la BD
            user_check = list(self.core.db.execute_query(
                "SELECT id FROM users WHERE id = %s", (user_id,)
            ))
            if not user_check:
                # Intentar encontrar cualquier usuario activo
                any_user = list(self.core.db.execute_query(
                    "SELECT id FROM users WHERE is_active = 1 LIMIT 1"
                ))
                if any_user:
                    user_id = any_user[0]['id']
                    self.log_text.append(f"⚠️ Usuario {STATE.user_id} no encontrado, usando ID={user_id}\n")
                else:
                    self.log_text.append("⚠️ No hay usuarios en la BD, usando user_id=1\n")
                    user_id = 1

            self.log_text.append(f"📍 Usuario: {user_id}, Sucursal: {branch_id}, Turno: {turn_id or 'N/A'}\n\n")
            
            # Leer archivo (CSV o Excel)
            rows = []
            if file_path.lower().endswith(('.xlsx', '.xls')):
                if not HAS_OPENPYXL:
                    self.log_text.append("❌ Error: Se requiere openpyxl para importar Excel.\n")
                    return
                
                import zipfile
                import os
                
                if not os.path.exists(file_path) or os.path.getsize(file_path) == 0:
                    self.log_text.append("❌ Error: Archivo no válido")
                    return
                
                if file_path.lower().endswith('.xlsx'):
                    try:
                        with zipfile.ZipFile(file_path, 'r') as zip_ref:
                            zip_files = zip_ref.namelist()
                            if not any('xl/workbook.xml' in f or '[Content_Types].xml' in f for f in zip_files):
                                self.log_text.append("❌ Error: Archivo no parece ser un Excel válido")
                                return
                    except zipfile.BadZipFile:
                        self.log_text.append("❌ Error: Archivo Excel corrupto")
                        return
                
                wb = None
                try:
                    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                    ws = wb.active

                    headers = []
                    # FIX CRÍTICO 2026-01-30: ws[1] NO funciona en read_only mode
                    # Usar iter_rows en su lugar
                    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                    if first_row:
                        headers = [str(cell) if cell is not None else "" for cell in first_row]
                    else:
                        self.log_text.append("❌ Error: No se encontró fila de encabezados")
                        return

                    for row in ws.iter_rows(min_row=2, values_only=True):
                        row_dict = {}
                        for i, cell_value in enumerate(row):
                            if i < len(headers):
                                row_dict[headers[i]] = str(cell_value) if cell_value is not None else ""
                        if any(row_dict.values()):
                            rows.append(row_dict)

                    self.log_text.append(f"✅ Archivo Excel leído: {len(rows)} filas\n")
                except Exception as e:
                    import traceback
                    self.log_text.append(f"❌ Error leyendo Excel: {e}")
                    self.log_text.append(f"   Detalle: {traceback.format_exc()[:500]}")
                    logger.error("Error leyendo Excel: %s", traceback.format_exc())
                    self.update_status("❌ ERROR LEYENDO EXCEL", "#f44336", "#FFEBEE")
                    self.update_counter(0, 0)
                    return
                finally:
                    if wb is not None:
                        wb.close()
            else:
                # CSV - FIX 2026-01-30: Mejor manejo de encoding y logging
                import os
                file_size = os.path.getsize(file_path)
                self.log_text.append(f"📁 Archivo CSV: {file_size:,} bytes\n")

                # Intentar diferentes encodings
                encodings_to_try = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
                rows = []
                encoding_used = None

                for encoding in encodings_to_try:
                    try:
                        with open(file_path, 'r', encoding=encoding) as f:
                            reader = csv.DictReader(f)
                            rows = list(reader)
                            encoding_used = encoding
                            break
                    except UnicodeDecodeError:
                        continue
                    except Exception as e:
                        self.log_text.append(f"⚠️ Error con encoding {encoding}: {e}\n")
                        continue

                if not rows and not encoding_used:
                    self.log_text.append("❌ No se pudo leer el archivo con ningún encoding\n")
                    return

                self.log_text.append(f"✅ CSV leído con encoding {encoding_used}: {len(rows)} filas\n")

                # FIX 2026-01-30: Mostrar columnas detectadas para diagnóstico
                if rows:
                    columns = list(rows[0].keys())
                    self.log_text.append(f"📋 Columnas detectadas: {', '.join(columns[:10])}{'...' if len(columns) > 10 else ''}\n")

            # FIX 2026-01-30: Validar que hay datos
            if not rows:
                self.log_text.append("❌ Error: No se encontraron filas para importar\n")
                return

            # FIX 2026-01-30: Diagnóstico exhaustivo
            self.log_text.append(f"\n{'═'*60}")
            self.log_text.append("🔍 DIAGNÓSTICO DE IMPORTACIÓN SIMPLE")
            self.log_text.append(f"{'═'*60}")
            self.log_text.append(f"📊 Total filas a importar: {len(rows)}")

            # Mostrar columnas detectadas
            if rows:
                all_cols = list(rows[0].keys())
                self.log_text.append(f"📋 Columnas detectadas ({len(all_cols)}): {all_cols}")

                # Mostrar muestra de datos
                self.log_text.append("\n📑 MUESTRA (primeras 3 filas):")
                for i, sample in enumerate(rows[:3]):
                    self.log_text.append(f"   Fila {i+1}: {dict(list(sample.items())[:5])}")

            self.log_text.append(f"{'─'*60}\n")

            self.progress.setMaximum(len(rows))
            self.update_counter(0, len(rows))  # FIX 2026-02-01: Inicializar contador
            imported = 0
            errors_db = 0
            last_error = None

            # FIX 2026-01-30: Contadores de validación
            skipped_duplicates = 0
            skipped_invalid = 0

            # FIX 2026-01-31: Obtener mapeo de columnas
            mapping = getattr(self.wizard(), 'column_mapping', {})

            # Helper function para usar mapeo o fallback
            def get_mapped(row, field_id, fallbacks):
                """Get value from row using mapped column or fallback to auto-detection"""
                mapped_col = mapping.get(field_id)
                if mapped_col and mapped_col in row:
                    return row.get(mapped_col)
                for col in fallbacks:
                    if col in row:
                        return row.get(col)
                return None

            for i, row in enumerate(rows):
                try:
                    # FIX 2026-01-31: Usar get_mapped para todos los campos
                    # Parse timestamp
                    fecha = get_mapped(row, 'date', ['FECHA', 'Fecha', 'fecha', 'date', 'timestamp', 'created_at'])
                    hora = get_mapped(row, 'time', ['HORA', 'Hora', 'hora', 'time'])
                    if fecha and hora:
                        timestamp = f"{fecha} {hora}"
                    elif fecha:
                        timestamp = f"{fecha} 00:00:00"
                    else:
                        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

                    # Parse total - FIX 2026-01-30: Manejo robusto de valores numéricos
                    total_str = get_mapped(row, 'sale_total', ['TOTAL_VENTA', 'Total_Venta', 'total_venta', 'total', 'Total']) or '0'
                    try:
                        total_clean = str(total_str).replace('$', '').replace(',', '').replace(' ', '').strip()
                        total = float(total_clean) if total_clean and total_clean.lower() not in ('none', 'null', '') else 0.0
                    except (ValueError, TypeError):
                        self.log_text.append(f"⚠️ Fila {i+1}: Total inválido '{total_str}', usando 0")
                        total = 0.0

                    # Calculate tax (assume 16% IVA)
                    tax_rate = 0.16
                    subtotal = total / (1 + tax_rate) if total > 0 else 0.0
                    tax = total - subtotal

                    # Parse payment method
                    payment = get_mapped(row, 'payment', ['MÉTODO_PAGO', 'Metodo_Pago', 'metodo_pago', 'payment_method', 'payment', 'metodo', 'Método']) or 'cash'
                    payment_map = {
                        'efectivo': 'cash', 'tarjeta': 'card', 'credito': 'credit',
                        'crédito': 'credit', 'mixto': 'mixed', 'transferencia': 'transfer'
                    }
                    payment = payment_map.get(payment.lower(), payment.lower()) if payment else 'cash'

                    # ═══════════════════════════════════════════════════════════════
                    # FIX CRÍTICO 2026-01-30: VALIDACIÓN DE INTEGRIDAD REFERENCIAL
                    # ═══════════════════════════════════════════════════════════════

                    # Customer (optional) - VALIDAR QUE EXISTE
                    customer_str = get_mapped(row, 'customer', ['CLIENTE', 'Cliente', 'cliente', 'customer', 'customer_name', 'customer_id', 'cliente_id'])
                    customer_id = None
                    if customer_str:
                        try:
                            customer_id = int(customer_str)
                            customer_check = list(self.core.db.execute_query(
                                "SELECT id FROM customers WHERE id = %s", (customer_id,)
                            ))
                            if not customer_check:
                                customer_id = None
                        except (ValueError, TypeError):
                            customer_id = None

                    # FIX CRÍTICO 2026-02-01: Ventas importadas NUNCA deben tener turn_id
                    # El turn_id causaba que el cuadre de caja sumara ventas históricas
                    # Ver docs/ISSUES_CRITICOS_SYNC_2026-02-01.md Issue #1
                    final_turn_id = None  # SIEMPRE NULL para importaciones

                    # FIX 2026-01-30: Parsear serie del CSV si existe
                    serie = get_mapped(row, 'serie', ['SERIE', 'Serie', 'serie', 'type']) or 'A'
                    if serie not in ('A', 'B'):
                        serie = 'A'

                    # FIX 2026-01-30: Parsear origin_pc del CSV (PC de procedencia)
                    origin_pc = get_mapped(row, 'terminal', ['TERMINAL', 'Terminal', 'terminal', 'PC', 'pc', 'origin_pc', 'ORIGIN_PC'])
                    if origin_pc and str(origin_pc).lower() in ('none', 'null', ''):
                        origin_pc = None

                    # FIX 2026-01-30: Parsear branch_id del CSV (sucursal) - VALIDAR QUE EXISTE
                    csv_branch_str = get_mapped(row, 'branch', ['SUCURSAL_ID', 'branch_id', 'BRANCH_ID', 'SUCURSAL', 'Sucursal', 'sucursal_id'])
                    csv_branch_id = None
                    if csv_branch_str:
                        try:
                            csv_branch_id = int(csv_branch_str)
                        except (ValueError, TypeError):
                            csv_branch_id = None
                    if csv_branch_id:
                        try:
                            csv_branch_id = int(csv_branch_id)
                            # Verificar que la sucursal existe
                            branch_check = list(self.core.db.execute_query(
                                "SELECT id FROM branches WHERE id = %s", (csv_branch_id,)
                            ))
                            if not branch_check:
                                csv_branch_id = None  # Usar default
                        except Exception:
                            csv_branch_id = None
                    final_branch_id = csv_branch_id or branch_id

                    # FIX 2026-01-31: Usuario del CSV
                    user_str = get_mapped(row, 'user', ['USUARIO', 'Usuario', 'usuario', 'user', 'cajero', 'cashier', 'USER_ID', 'user_id'])
                    csv_user_id = None
                    if user_str:
                        try:
                            csv_user_id = int(user_str)
                            user_check = list(self.core.db.execute_query(
                                "SELECT id FROM users WHERE id = %s", (csv_user_id,)
                            ))
                            if not user_check:
                                csv_user_id = None
                        except (ValueError, TypeError):
                            csv_user_id = None
                    final_user_id = csv_user_id or user_id

                    # Notas del CSV
                    notas = get_mapped(row, 'notes', ['NOTAS', 'Notas', 'notas', 'notes', 'comments', 'observaciones'])

                    # FIX 2026-01-30: Parsear folio original del CSV si existe
                    csv_folio = get_mapped(row, 'folio', ['FOLIO', 'Folio', 'folio', 'NUMERO', 'numero', 'folio_visible'])
                    if csv_folio and str(csv_folio).lower() not in ('none', 'null', ''):
                        folio_visible = str(csv_folio)
                    else:
                        folio_visible = f"IMP-{datetime.now().strftime('%Y%m%d')}-{i+1:04d}"

                    # Generate unique UUID (or use existing from CSV)
                    csv_uuid = row.get('uuid', row.get('UUID', None))
                    if csv_uuid and str(csv_uuid).lower() not in ('none', 'null', ''):
                        sale_uuid = str(csv_uuid)
                    else:
                        sale_uuid = str(uuid_module.uuid4())

                    # ═══════════════════════════════════════════════════════════════
                    # FIX CRÍTICO 2026-01-30: VALIDACIÓN DE DUPLICADOS
                    # ═══════════════════════════════════════════════════════════════

                    # Verificar si UUID ya existe
                    existing_uuid = list(self.core.db.execute_query(
                        "SELECT id FROM sales WHERE uuid = %s", (sale_uuid,)
                    ))
                    if existing_uuid:
                        skipped_duplicates += 1
                        if skipped_duplicates <= 3:
                            self.log_text.append(f"⚠️ Fila {i+1}: UUID {sale_uuid[:8]}... ya existe, saltando")
                        continue

                    # Verificar si folio_visible ya existe (solo si viene del CSV)
                    if csv_folio:
                        existing_folio = list(self.core.db.execute_query(
                            "SELECT id FROM sales WHERE folio_visible = %s AND serie = %s",
                            (folio_visible, serie)
                        ))
                        if existing_folio:
                            skipped_duplicates += 1
                            if skipped_duplicates <= 3:
                                self.log_text.append(f"⚠️ Fila {i+1}: Folio {folio_visible} serie {serie} ya existe, saltando")
                            continue

                    # Insert sale with ALL required fields including origin_pc, branch_id and synced
                    # FIX 2026-01-30: Incluir synced=0 para sincronización bidireccional
                    sale_sql = """
                        INSERT INTO sales (
                            uuid, timestamp, subtotal, tax, total,
                            payment_method, customer_id, user_id, turn_id,
                            serie, folio_visible, status, notes, origin_pc, branch_id,
                            synced, sync_status
                        )
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 0, 'pending')
                        RETURNING id
                    """
                    # #region agent log
                    if agent_log_enabled():
                        import json, time
                        try:
                            from app.utils.path_utils import get_debug_log_path_str
                            with open(get_debug_log_path_str(), "a") as f:
                                f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"IMPORT_WIZARD","location":"import_wizard.py:import_sales","message":"Inserting sale","data":{"total":total,"payment":payment,"status":status},"timestamp":int(time.time()*1000)})+"\n")
                        except Exception as e: logger.debug("Writing inserting sale log: %s", e)
                    # #endregion
                    
                    # FIX 2026-01-31: Usar final_user_id y notas del CSV
                    note_final = notas if notas else note_text
                    result = self.core.db.execute_write(sale_sql, (
                        sale_uuid, timestamp, round(subtotal, 2), round(tax, 2), total,
                        payment, customer_id, final_user_id, final_turn_id,
                        serie, folio_visible, status, note_final, origin_pc, final_branch_id
                    ))
                    # FIX 2026-01-31: Extraer sale_id correctamente del RETURNING
                    sale_id = None
                    if result:
                        if isinstance(result, (list, tuple)) and len(result) > 0:
                            first_row = result[0]
                            if isinstance(first_row, dict):
                                sale_id = first_row.get('id')
                            elif isinstance(first_row, (list, tuple)) and len(first_row) > 0:
                                sale_id = first_row[0]
                            else:
                                sale_id = first_row
                        elif isinstance(result, int):
                            sale_id = result
                    
                    # #region agent log
                    if agent_log_enabled():
                        import json, time
                        try:
                            from app.utils.path_utils import get_debug_log_path_str
                            with open(get_debug_log_path_str(), "a") as f:
                                f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"IMPORT_WIZARD","location":"import_wizard.py:import_sales","message":"Sale inserted","data":{"sale_id":sale_id},"timestamp":int(time.time()*1000)})+"\n")
                        except Exception as e: logger.debug("Writing sale inserted log: %s", e)
                    # #endregion
                    
                    if not sale_id or sale_id == 0:
                        # #region agent log
                        if agent_log_enabled():
                            import json, time
                            try:
                                from app.utils.path_utils import get_debug_log_path_str
                                with open(get_debug_log_path_str(), "a") as f:
                                    f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"IMPORT_WIZARD","location":"import_wizard.py:import_sales","message":"Warning: sale_id is 0 or None","data":{},"timestamp":int(time.time()*1000)})+"\n")
                            except Exception as e: logger.debug("Writing sale_id warning log: %s", e)
                        # #endregion
                        self.log_text.append(f"⚠️ Fila {i+1}: Venta insertada pero sale_id no disponible")
                        sale_id = None  # No insertar items si no hay sale_id
                    
                    # Parse and insert items if provided
                    items_str = row.get('items', row.get('Items', row.get('productos', '')))
                    if items_str and sale_id:
                        try:
                            # Try JSON first
                            items_list = json.loads(items_str)
                        except Exception:
                            # Fallback: comma-separated
                            items_list = [{'name': name.strip(), 'qty': 1, 'price': 0} 
                                         for name in items_str.split(',') if name.strip()]
                        
                        for item in items_list:
                            if isinstance(item, dict):
                                # FIX 2026-01-30: Incluir name y usar NULL en lugar de 0
                                item_sql = """
                                    INSERT INTO sale_items (sale_id, product_id, name, qty, price, subtotal, synced)
                                    VALUES (%s, %s, %s, %s, %s, %s, 0)
                                """
                                qty = float(item.get('qty', item.get('quantity', 1)))
                                price = float(item.get('price', item.get('precio', 0)))
                                item_name = item.get('name', item.get('nombre', ''))
                                # FIX 2026-01-30: Usar None (NULL) si no hay product_id válido
                                raw_pid = item.get('product_id', item.get('producto_id'))
                                product_id = int(raw_pid) if raw_pid and str(raw_pid).isdigit() and int(raw_pid) > 0 else None
                                try:
                                    self.core.db.execute_write(item_sql, (
                                        sale_id, product_id, item_name, qty, price, qty * price
                                    ))
                                except Exception as item_error:
                                    self.log_text.append(f"   ⚠️ Error insertando item: {item_error}")
                    
                    imported += 1
                    # Solo mostrar cada venta individual si son pocas (< 50)
                    if len(rows) < 50:
                        self.log_text.append(f"✅ Venta #{sale_id}: ${total:.2f} ({payment})")

                except Exception as e:
                    errors_db += 1
                    last_error = str(e)
                    self.log_text.append(f"❌ Fila {i+1}: {e}")
                    import traceback
                    tb = traceback.format_exc()
                    logger.error(f"Error importando fila {i+1}: {tb}")
                    # Mostrar detalles de primeros 5 errores
                    if errors_db <= 5:
                        self.log_text.append(f"   Detalle: {tb[:300]}")

                self.progress.setValue(i + 1)
                self.update_counter(i + 1, len(rows))  # FIX 2026-02-01: Actualizar contador

                # FIX 2026-01-30: Mostrar progreso cada 50 registros
                if (i + 1) % 50 == 0:
                    self.log_text.append(f"📈 Progreso: {i+1}/{len(rows)} | OK: {imported} | Err: {errors_db}")

                QtWidgets.QApplication.processEvents()

            # ══════════════════════════════════════════════════════════════════
            # RESUMEN FINAL
            # ══════════════════════════════════════════════════════════════════
            debug_print(f"=== RESUMEN SIMPLE: {imported} ventas de {len(rows)} filas, {skipped_duplicates} duplicados, {errors_db} errores ===")
            self.log_text.append(f"\n{'═'*60}")
            self.log_text.append("🎉 IMPORTACIÓN COMPLETADA - RESUMEN")
            self.log_text.append(f"{'═'*60}")
            self.log_text.append(f"   📁 Filas en archivo: {len(rows)}")
            self.log_text.append(f"   ✅ Ventas importadas: {imported}")
            self.log_text.append(f"   ⏭️ Duplicados saltados: {skipped_duplicates}")
            self.log_text.append(f"   ❌ Errores: {errors_db}")
            if last_error and errors_db > 0:
                self.log_text.append(f"   📝 Último error: {last_error[:100]}")
            self.log_text.append(f"\n⚠️ Status de ventas: '{status}'")
            self.log_text.append("   Las ventas NO afectaron el inventario.")

        except Exception as e:
            import traceback
            self.log_text.append(f"\n❌ Error fatal: {e}")
            self.log_text.append(f"   {traceback.format_exc()[:500]}")

    def import_sales_detailed(self, file_path):
        """
        Import sales from DETAILED CSV format.
        
        This format has one row per PRODUCT, not per sale.
        Multiple rows with the same FOLIO belong to the same sale.
        
        Expected columns (from detailed export):
        - FOLIO: Sale identifier (groups products)
        - SERIE: A or B
        - FECHA: Date
        - HORA: Time
        - CLIENTE: Customer name
        - RFC_CLIENTE: Customer RFC
        - SKU: Product SKU
        - CÓDIGO_BARRAS: Barcode
        - PRODUCTO: Product name
        - CANTIDAD: Quantity
        - PRECIO_UNITARIO: Unit price
        - DESCUENTO: Discount
        - SUBTOTAL_PRODUCTO: Line subtotal
        - MÉTODO_PAGO: Payment method
        - TOTAL_VENTA: Sale total
        - USUARIO: User
        - SUCURSAL: Branch
        - NOTAS: Notes
        """
        is_simulated = self.field("sales_simulated")
        status = "simulated" if is_simulated else "completed"
        status_label = "SIMULADAS" if is_simulated else "REALES"
        note_text = "Importado (detallado)" if is_simulated else "Restaurado desde backup detallado"

        debug_print(f"=== INICIO IMPORTACIÓN DETALLADA ({status_label}) ===")
        self.log_text.append(f"📋 Importando ventas DETALLADAS {status_label}...\n")
        self.log_text.append(f"📊 Formato: Una línea por producto\n")
        self.log_text.append("⚠️ Las ventas NO afectarán inventario\n\n")

        # FIX 2026-02-01: Inicializar contador al inicio para mostrar "procesando..."
        self.update_status("📂 LEYENDO ARCHIVO...", "#2196F3", "#E3F2FD")
        self.progress.setMaximum(0)  # Modo indeterminado mientras lee
        QtWidgets.QApplication.processEvents()

        try:
            from collections import defaultdict
            from datetime import datetime
            import uuid as uuid_module

            from app.core import STATE

            user_id = STATE.user_id if STATE.user_id else 1
            branch_id = STATE.branch_id if STATE.branch_id else 1
            turn_id = STATE.current_turn_id if hasattr(STATE, 'current_turn_id') and STATE.current_turn_id else None

            # FIX 2026-01-30: Validar que user_id existe en la BD
            user_check = list(self.core.db.execute_query(
                "SELECT id FROM users WHERE id = %s", (user_id,)
            ))
            if not user_check:
                any_user = list(self.core.db.execute_query(
                    "SELECT id FROM users WHERE is_active = 1 LIMIT 1"
                ))
                if any_user:
                    user_id = any_user[0]['id']
                    self.log_text.append(f"⚠️ Usuario {STATE.user_id} no encontrado, usando ID={user_id}\n")
                else:
                    user_id = 1

            # Leer archivo (CSV o Excel)
            rows = []
            if file_path.lower().endswith(('.xlsx', '.xls')):
                if not HAS_OPENPYXL:
                    self.log_text.append("❌ Error: Se requiere openpyxl para importar Excel.\n")
                    return
                
                import zipfile
                import os
                
                if not os.path.exists(file_path) or os.path.getsize(file_path) == 0:
                    self.log_text.append("❌ Error: Archivo no válido")
                    return
                
                if file_path.lower().endswith('.xlsx'):
                    try:
                        with zipfile.ZipFile(file_path, 'r') as zip_ref:
                            zip_files = zip_ref.namelist()
                            if not any('xl/workbook.xml' in f or '[Content_Types].xml' in f for f in zip_files):
                                self.log_text.append("❌ Error: Archivo no parece ser un Excel válido")
                                return
                    except zipfile.BadZipFile:
                        self.log_text.append("❌ Error: Archivo Excel corrupto")
                        return
                
                wb = None
                try:
                    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                    ws = wb.active

                    headers = []
                    # FIX CRÍTICO 2026-01-30: ws[1] NO funciona en read_only mode
                    # Usar iter_rows en su lugar
                    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                    if first_row:
                        headers = [str(cell) if cell is not None else "" for cell in first_row]
                    else:
                        self.log_text.append("❌ Error: No se encontró fila de encabezados")
                        return

                    for row in ws.iter_rows(min_row=2, values_only=True):
                        row_dict = {}
                        for i, cell_value in enumerate(row):
                            if i < len(headers):
                                row_dict[headers[i]] = str(cell_value) if cell_value is not None else ""
                        if any(row_dict.values()):
                            rows.append(row_dict)

                    self.log_text.append(f"✅ Archivo Excel leído: {len(rows)} filas\n")
                except Exception as e:
                    import traceback
                    self.log_text.append(f"❌ Error leyendo Excel: {e}")
                    self.log_text.append(f"   Detalle: {traceback.format_exc()[:500]}")
                    logger.error("Error leyendo Excel: %s", traceback.format_exc())
                    self.update_status("❌ ERROR LEYENDO EXCEL", "#f44336", "#FFEBEE")
                    self.update_counter(0, 0)
                    return
                finally:
                    if wb is not None:
                        wb.close()
            else:
                # CSV - FIX 2026-01-30: Mejor manejo de encoding y logging
                import os
                file_size = os.path.getsize(file_path)
                self.log_text.append(f"📁 Archivo CSV: {file_size:,} bytes\n")

                # Intentar diferentes encodings
                encodings_to_try = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
                rows = []
                encoding_used = None

                for encoding in encodings_to_try:
                    try:
                        with open(file_path, 'r', encoding=encoding) as f:
                            reader = csv.DictReader(f)
                            rows = list(reader)
                            encoding_used = encoding
                            break
                    except UnicodeDecodeError:
                        continue
                    except Exception as e:
                        self.log_text.append(f"⚠️ Error con encoding {encoding}: {e}\n")
                        continue

                if rows:
                    self.log_text.append(f"✅ CSV leído con encoding {encoding_used}: {len(rows)} filas\n")
                    columns = list(rows[0].keys())
                    self.log_text.append(f"📋 Columnas detectadas: {', '.join(columns[:10])}{'...' if len(columns) > 10 else ''}\n")

            if not rows:
                debug_print("ERROR: Archivo vacío o no se pudo leer")
                self.log_text.append("❌ Archivo vacío o no se pudo leer")
                self.update_status("❌ ERROR: ARCHIVO VACÍO", "#f44336", "#FFEBEE")
                self.update_counter(0, 0)
                return

            debug_print(f"Archivo leído: {len(rows)} filas")

            # Get column mapping from wizard
            mapping = getattr(self.wizard(), 'column_mapping', {})
            if mapping:
                debug_print(f"Mapeo de columnas: {mapping}")

            # Helper function to get value using mapping or fallback
            def get_mapped(row, field_id, fallbacks):
                """Get value from row using mapped column or fallback to auto-detection"""
                mapped_col = mapping.get(field_id)
                if mapped_col and mapped_col in row:
                    return row.get(mapped_col)
                # Fallback to auto-detection
                for col in fallbacks:
                    if col in row:
                        return row.get(col)
                return None

            # ══════════════════════════════════════════════════════════════════
            # DIAGNÓSTICO EXHAUSTIVO 2026-01-30: Detectar problema de ~100 registros
            # ══════════════════════════════════════════════════════════════════
            self.log_text.append(f"\n{'═'*60}")
            self.log_text.append("🔍 DIAGNÓSTICO DE IMPORTACIÓN")
            self.log_text.append(f"{'═'*60}")
            self.log_text.append(f"📊 TOTAL FILAS LEÍDAS DEL ARCHIVO: {len(rows)}")

            # Show mapping if configured
            if mapping:
                self.log_text.append(f"🗺️ MAPEO MANUAL: {mapping}")

            # Mostrar primeras columnas detectadas
            if rows:
                all_columns = list(rows[0].keys())
                self.log_text.append(f"📋 COLUMNAS ({len(all_columns)}): {all_columns}")

                # Detectar columna de FOLIO (use mapping first)
                folio_col_found = mapping.get('folio')
                if not folio_col_found:
                    folio_candidates = ['FOLIO', 'Folio', 'folio', 'NUMERO', 'Numero', 'numero', 'TICKET', 'Ticket']
                    for col in folio_candidates:
                        if col in all_columns:
                            folio_col_found = col
                            break

                if folio_col_found:
                    self.log_text.append(f"✅ Columna FOLIO: '{folio_col_found}'")
                else:
                    self.log_text.append(f"⚠️ NO se encontró columna FOLIO. Buscando: {folio_candidates}")
                    self.log_text.append(f"   Columnas disponibles: {all_columns[:15]}...")

                # Mostrar primeras 3 filas como muestra
                self.log_text.append("\n📑 MUESTRA DE DATOS (primeras 3 filas):")
                for i, sample_row in enumerate(rows[:3]):
                    folio_val = sample_row.get('FOLIO', sample_row.get('Folio', sample_row.get('folio', 'N/A')))
                    producto_val = sample_row.get('PRODUCTO', sample_row.get('Producto', sample_row.get('producto', 'N/A')))
                    total_val = sample_row.get('TOTAL_VENTA', sample_row.get('Total_Venta', sample_row.get('total', 'N/A')))
                    self.log_text.append(f"   Fila {i+1}: FOLIO='{folio_val}' | PRODUCTO='{str(producto_val)[:30]}' | TOTAL='{total_val}'")

            self.log_text.append(f"{'─'*60}\n")

            # Agrupar filas por FOLIO con diagnóstico detallado
            sales_grouped = defaultdict(list)
            rows_without_folio = 0
            folio_values_seen = set()
            empty_folio_samples = []

            for idx, row in enumerate(rows):
                folio = get_mapped(row, 'folio', ['FOLIO', 'Folio', 'folio', 'NUMERO', 'numero', 'TICKET'])
                folio_str = str(folio).strip() if folio is not None else ''

                if folio_str and folio_str.lower() not in ('none', 'null', ''):
                    sales_grouped[folio_str].append(row)
                    folio_values_seen.add(folio_str)
                else:
                    rows_without_folio += 1
                    # Guardar muestra de filas sin FOLIO
                    if len(empty_folio_samples) < 3:
                        producto = get_mapped(row, 'product_name', ['PRODUCTO', 'Producto', 'producto', 'product'])
                        empty_folio_samples.append({
                            'row_num': idx + 1,
                            'folio_raw': repr(folio),
                            'producto': str(producto or '')[:30]
                        })

            # Diagnóstico de agrupación
            debug_print(f"Agrupación: {len(sales_grouped)} FOLIOs únicos, {rows_without_folio} filas sin FOLIO, {len(rows)} filas total")
            self.log_text.append(f"{'═'*60}")
            self.log_text.append("📊 RESULTADO DE AGRUPACIÓN POR FOLIO:")
            self.log_text.append(f"   ✅ FOLIOs únicos encontrados: {len(sales_grouped)}")
            self.log_text.append(f"   ❌ Filas SIN FOLIO válido: {rows_without_folio}")
            self.log_text.append(f"   📄 Total filas procesadas: {len(rows)}")

            # Verificar proporción
            if len(sales_grouped) > 0:
                avg_products_per_sale = len(rows) / len(sales_grouped)
                self.log_text.append(f"   📈 Promedio productos/venta: {avg_products_per_sale:.1f}")

            # Si hay muchas filas sin FOLIO, mostrar muestras
            if rows_without_folio > 0 and empty_folio_samples:
                self.log_text.append(f"\n   ⚠️ MUESTRAS de filas SIN FOLIO:")
                for sample in empty_folio_samples:
                    self.log_text.append(f"      Fila {sample['row_num']}: FOLIO={sample['folio_raw']} | PRODUCTO='{sample['producto']}'")

            # Mostrar algunos FOLIOs encontrados como muestra
            if folio_values_seen:
                sample_folios = list(folio_values_seen)[:5]
                self.log_text.append(f"\n   📝 Muestras de FOLIOs encontrados: {sample_folios}")

            self.log_text.append(f"{'═'*60}\n")

            # FIX CRÍTICO: Si NO hay ventas agrupadas, intentar modo alternativo
            if len(sales_grouped) == 0 and len(rows) > 0:
                self.log_text.append("⚠️ MODO FALLBACK: No se encontraron FOLIOs válidos.")
                self.log_text.append("   Intentando tratar cada fila como una venta individual...")
                # Crear un FOLIO sintético para cada fila
                for idx, row in enumerate(rows):
                    synthetic_folio = f"IMPORT-{datetime.now().strftime('%Y%m%d')}-{idx+1:06d}"
                    sales_grouped[synthetic_folio] = [row]
                self.log_text.append(f"   Creadas {len(sales_grouped)} ventas con FOLIOs sintéticos\n")

            self.log_text.append(f"📦 Ventas a importar: {len(sales_grouped)}")
            self.log_text.append(f"📄 Líneas de productos: {len(rows)}\n")
            
            self.progress.setMaximum(len(sales_grouped))
            imported = 0
            items_imported = 0
            errors_db = 0  # Errores de base de datos
            errors_no_id = 0  # Ventas sin ID retornado
            last_error_msg = None
            skipped_duplicates = 0  # FIX 2026-01-30: Contador de duplicados

            # FIX 2026-01-30: Contadores detallados para diagnóstico
            self.log_text.append(f"\n{'─'*60}")
            self.log_text.append("🚀 INICIANDO IMPORTACIÓN DE VENTAS...")
            self.log_text.append(f"{'─'*60}\n")

            # Actualizar estado visual
            self.update_status("🚀 IMPORTANDO VENTAS...", "#FF9800", "#FFF3E0")
            self.update_counter(0, len(sales_grouped))
            QtWidgets.QApplication.processEvents()  # FIX 2026-02-01: Actualizar UI antes de iniciar loop

            logger.info("🚀 Iniciando importación de %d ventas...", len(sales_grouped))

            for idx, (folio, product_rows) in enumerate(sales_grouped.items()):
                try:
                    # Tomar datos de la venta desde la primera fila
                    first_row = product_rows[0]

                    # FIX 2026-01-31: Usar get_mapped para todos los campos
                    # Fecha y hora
                    fecha = get_mapped(first_row, 'date', ['FECHA', 'Fecha', 'fecha', 'date', 'timestamp'])
                    hora = get_mapped(first_row, 'time', ['HORA', 'Hora', 'hora', 'time'])
                    if fecha and hora:
                        timestamp = f"{fecha} {hora}"
                    elif fecha:
                        timestamp = f"{fecha} 00:00:00"
                    else:
                        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

                    # Total de la venta - FIX 2026-01-30: Manejo robusto
                    total_str = get_mapped(first_row, 'sale_total', ['TOTAL_VENTA', 'Total_Venta', 'total_venta', 'total', 'Total']) or '0'
                    try:
                        total_clean = str(total_str).replace('$', '').replace(',', '').replace(' ', '').strip()
                        total = float(total_clean) if total_clean and total_clean.lower() not in ('none', 'null', '') else 0.0
                    except (ValueError, TypeError):
                        self.log_text.append(f"⚠️ {folio}: Total inválido '{total_str}', usando 0")
                        total = 0.0

                    # Serie
                    serie = get_mapped(first_row, 'serie', ['SERIE', 'Serie', 'serie', 'type']) or 'A'
                    if serie not in ('A', 'B'):
                        serie = 'A'

                    # Método de pago
                    payment = get_mapped(first_row, 'payment', ['MÉTODO_PAGO', 'Metodo_Pago', 'metodo_pago', 'payment_method', 'payment']) or 'cash'
                    payment_map = {
                        'efectivo': 'cash', 'tarjeta': 'card', 'credito': 'credit',
                        'crédito': 'credit', 'mixto': 'mixed', 'transferencia': 'transfer',
                        'cash': 'cash', 'card': 'card', 'credit': 'credit'
                    }
                    payment = payment_map.get(payment.lower(), payment.lower()) if payment else 'cash'

                    # Cliente
                    cliente_nombre = get_mapped(first_row, 'customer', ['CLIENTE', 'Cliente', 'cliente', 'customer', 'customer_name']) or ''

                    # Calcular subtotal e IVA
                    tax_rate = 0.16
                    subtotal = total / (1 + tax_rate) if total > 0 else 0.0
                    tax = total - subtotal

                    # FIX 2026-01-30: Parsear origin_pc del CSV (PC de procedencia)
                    origin_pc = get_mapped(first_row, 'terminal', ['TERMINAL', 'Terminal', 'terminal', 'PC', 'pc', 'ORIGIN_PC', 'origin_pc'])
                    if origin_pc and str(origin_pc).lower() in ('none', 'null', ''):
                        origin_pc = None

                    # FIX 2026-01-30: Parsear branch_id del CSV (sucursal) - CON VALIDACIÓN
                    branch_str = get_mapped(first_row, 'branch', ['SUCURSAL_ID', 'branch_id', 'BRANCH_ID', 'SUCURSAL', 'Sucursal'])
                    csv_branch_id = None
                    if branch_str:
                        try:
                            csv_branch_id = int(branch_str)
                            # Verificar que la sucursal existe
                            branch_check = list(self.core.db.execute_query(
                                "SELECT id FROM branches WHERE id = %s", (csv_branch_id,)
                            ))
                            if not branch_check:
                                csv_branch_id = None  # Usar default
                        except Exception:
                            csv_branch_id = None
                    final_branch_id = csv_branch_id or branch_id

                    # Usuario del CSV
                    user_str = get_mapped(first_row, 'user', ['USUARIO', 'Usuario', 'usuario', 'user', 'cajero', 'cashier', 'USER_ID'])
                    csv_user_id = None
                    if user_str:
                        try:
                            csv_user_id = int(user_str)
                            user_check = list(self.core.db.execute_query(
                                "SELECT id FROM users WHERE id = %s", (csv_user_id,)
                            ))
                            if not user_check:
                                csv_user_id = None
                        except Exception:
                            csv_user_id = None
                    final_user_id = csv_user_id or user_id

                    # FIX CRÍTICO 2026-02-04: Ventas importadas NUNCA deben tener turn_id
                    # El turn_id causaba que el cuadre de caja sumara ventas históricas
                    # Ver docs/ISSUES_CRITICOS_SYNC_2026-02-01.md Issue #1
                    final_turn_id = None  # SIEMPRE NULL para importaciones

                    # Notas
                    notas = get_mapped(first_row, 'notes', ['NOTAS', 'Notas', 'notas', 'notes', 'comments', 'observaciones'])

                    # FIX 2026-01-30: Generar o usar UUID existente
                    csv_uuid = first_row.get('UUID', first_row.get('uuid', None))
                    if csv_uuid and str(csv_uuid).lower() not in ('none', 'null', ''):
                        sale_uuid = str(csv_uuid)
                    else:
                        sale_uuid = str(uuid_module.uuid4())

                    # ═══════════════════════════════════════════════════════════════
                    # FIX CRÍTICO 2026-01-30: VALIDACIÓN DE DUPLICADOS
                    # ═══════════════════════════════════════════════════════════════

                    # Verificar si UUID ya existe
                    existing_uuid = list(self.core.db.execute_query(
                        "SELECT id FROM sales WHERE uuid = %s", (sale_uuid,)
                    ))
                    if existing_uuid:
                        skipped_duplicates += 1
                        if skipped_duplicates <= 3:
                            self.log_text.append(f"⚠️ {folio}: UUID {sale_uuid[:8]}... ya existe, saltando")
                        continue

                    # Verificar si folio_visible ya existe
                    folio_visible = folio if folio else f"IMP-{datetime.now().strftime('%Y%m%d')}-{idx+1:04d}"
                    existing_folio = list(self.core.db.execute_query(
                        "SELECT id FROM sales WHERE folio_visible = %s AND serie = %s",
                        (folio_visible, serie)
                    ))
                    if existing_folio:
                        skipped_duplicates += 1
                        if skipped_duplicates <= 3:
                            self.log_text.append(f"⚠️ {folio}: Folio {folio_visible} serie {serie} ya existe, saltando")
                        continue

                    # Insertar la venta con TODOS los campos incluyendo origin_pc, branch_id y synced
                    # FIX 2026-01-30: Incluir synced=0 para sincronización bidireccional
                    sale_sql = """
                        INSERT INTO sales (
                            uuid, timestamp, subtotal, tax, total,
                            payment_method, user_id, turn_id, serie,
                            folio_visible, status, notes, origin_pc, branch_id,
                            synced, sync_status
                        )
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 0, 'pending')
                        RETURNING id
                    """
                    # #region agent log
                    if agent_log_enabled():
                        import json, time
                        try:
                            from app.utils.path_utils import get_debug_log_path_str
                            with open(get_debug_log_path_str(), "a") as f:
                                f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"IMPORT_WIZARD","location":"import_wizard.py:import_sales_detailed","message":"Inserting detailed sale","data":{"folio":folio,"total":total,"status":status},"timestamp":int(time.time()*1000)})+"\n")
                        except Exception as e: logger.debug("Writing inserting detailed sale log: %s", e)
                    # #endregion
                    
                    # FIX 2026-01-31: Usar final_user_id y notas del CSV
                    # FIX 2026-02-04: Usar final_turn_id (NULL) en lugar de turn_id activo
                    note_final = notas if notas else f"{note_text} - Folio original: {folio}"
                    result = self.core.db.execute_write(sale_sql, (
                        sale_uuid, timestamp, round(subtotal, 2), round(tax, 2), total,
                        payment, final_user_id, final_turn_id, serie,
                        folio_visible, status, note_final,
                        origin_pc, final_branch_id
                    ))
                    # FIX 2026-01-31: Extraer sale_id correctamente del RETURNING
                    sale_id = None
                    if result:
                        if isinstance(result, (list, tuple)) and len(result) > 0:
                            first_row = result[0]
                            if isinstance(first_row, dict):
                                sale_id = first_row.get('id')
                            elif isinstance(first_row, (list, tuple)) and len(first_row) > 0:
                                sale_id = first_row[0]
                            else:
                                sale_id = first_row
                        elif isinstance(result, int):
                            sale_id = result
                    
                    # #region agent log
                    if agent_log_enabled():
                        import json, time
                        try:
                            from app.utils.path_utils import get_debug_log_path_str
                            with open(get_debug_log_path_str(), "a") as f:
                                f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"IMPORT_WIZARD","location":"import_wizard.py:import_sales_detailed","message":"Detailed sale inserted","data":{"sale_id":sale_id,"folio":folio},"timestamp":int(time.time()*1000)})+"\n")
                        except Exception as e: logger.debug("Writing detailed sale inserted log: %s", e)
                    # #endregion
                    
                    if not sale_id or sale_id == 0:
                        errors_no_id += 1
                        # #region agent log
                        if agent_log_enabled():
                            import json, time
                            try:
                                from app.utils.path_utils import get_debug_log_path_str
                                with open(get_debug_log_path_str(), "a") as f:
                                    f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"IMPORT_WIZARD","location":"import_wizard.py:import_sales_detailed","message":"Warning: sale_id is 0 or None","data":{"folio":folio},"timestamp":int(time.time()*1000)})+"\n")
                            except Exception as e: logger.debug("Writing detailed sale_id warning log: %s", e)
                        # #endregion
                        if errors_no_id <= 5:  # Solo mostrar primeras 5
                            self.log_text.append(f"⚠️ {folio}: Venta insertada pero sale_id no disponible")
                        # Aún así contar como importado, la venta SÍ se insertó
                        imported += 1
                        continue  # Saltar items si no hay sale_id
                    
                    # Insertar cada producto
                    for prod_row in product_rows:
                        try:
                            # FIX 2026-01-31: Usar get_mapped para productos
                            sku = get_mapped(prod_row, 'sku', ['SKU', 'Sku', 'sku', 'CODIGO', 'Codigo', 'codigo', 'CÓDIGO_BARRAS', 'barcode']) or ''
                            nombre = get_mapped(prod_row, 'product_name', ['PRODUCTO', 'Producto', 'producto', 'product', 'name', 'nombre']) or ''

                            # Si no tiene producto, saltar pero contar
                            if not nombre or nombre == "(Sin detalle de productos)" or str(nombre).lower() in ('none', 'null', ''):
                                # FIX 2026-01-30: Contar productos saltados
                                if not hasattr(self, '_skipped_items'):
                                    self._skipped_items = 0
                                self._skipped_items += 1
                                continue

                            # FIX 2026-01-30: Manejo robusto de valores numéricos
                            try:
                                cant_str = get_mapped(prod_row, 'quantity', ['CANTIDAD', 'Cantidad', 'cantidad', 'qty', 'quantity']) or '1'
                                cantidad = float(str(cant_str).replace(',', '.').strip()) if str(cant_str).lower() not in ('none', 'null', '') else 1.0
                            except (ValueError, TypeError):
                                cantidad = 1.0

                            try:
                                precio_str = get_mapped(prod_row, 'unit_price', ['PRECIO_UNITARIO', 'Precio_Unitario', 'precio_unitario', 'price', 'precio']) or '0'
                                precio_clean = str(precio_str).replace('$', '').replace(',', '').strip()
                                precio = float(precio_clean) if precio_clean and precio_clean.lower() not in ('none', 'null', '') else 0.0
                            except (ValueError, TypeError):
                                precio = 0.0

                            try:
                                descuento_str = get_mapped(prod_row, 'discount', ['DESCUENTO', 'Descuento', 'descuento', 'discount']) or '0'
                                descuento_clean = str(descuento_str).replace('$', '').replace(',', '').strip()
                                descuento = float(descuento_clean) if descuento_clean and descuento_clean.lower() not in ('none', 'null', '') else 0.0
                            except (ValueError, TypeError):
                                descuento = 0.0

                            line_subtotal = cantidad * precio - descuento

                            # FIX 2026-01-31: Extraer claves SAT del CSV
                            sat_clave_prod = get_mapped(prod_row, 'sat_clave_prod', ['CLAVE_SAT', 'Clave_SAT', 'clave_sat', 'SAT_CLAVE_PROD_SERV', 'sat_clave_prod_serv', 'sat_code']) or '01010101'
                            sat_clave_unidad = get_mapped(prod_row, 'sat_clave_unidad', ['CLAVE_UNIDAD_SAT', 'Clave_Unidad_SAT', 'clave_unidad_sat', 'SAT_CLAVE_UNIDAD', 'sat_clave_unidad', 'sat_unit']) or 'H87'
                            sat_descripcion = get_mapped(prod_row, 'sat_descripcion', ['DESCRIPCION_SAT', 'Descripcion_SAT', 'descripcion_sat', 'SAT_DESCRIPCION', 'sat_descripcion']) or ''

                            # Limpiar valores SAT
                            if sat_clave_prod and str(sat_clave_prod).lower() in ('none', 'null', ''):
                                sat_clave_prod = '01010101'
                            if sat_clave_unidad and str(sat_clave_unidad).lower() in ('none', 'null', ''):
                                sat_clave_unidad = 'H87'
                            if sat_descripcion and str(sat_descripcion).lower() in ('none', 'null'):
                                sat_descripcion = ''

                            # FIX 2026-01-30: Búsqueda mejorada de product_id
                            # Primero por SKU, luego por nombre, si no NULL
                            product_id = None  # NULL en lugar de 0 para integridad referencial
                            if sku:
                                try:
                                    prod = list(self.core.db.execute_query(
                                        "SELECT id FROM products WHERE sku = %s LIMIT 1",
                                        (sku,)
                                    ))
                                    if prod:
                                        product_id = prod[0]['id']
                                except Exception:
                                    pass

                            # Si no encontró por SKU, intentar por nombre
                            if product_id is None and nombre:
                                try:
                                    prod = list(self.core.db.execute_query(
                                        "SELECT id FROM products WHERE name ILIKE %s LIMIT 1",
                                        (nombre,)
                                    ))
                                    if prod:
                                        product_id = prod[0]['id']
                                except Exception:
                                    pass

                            # FIX 2026-01-31: Insertar item CON claves SAT
                            item_sql = """
                                INSERT INTO sale_items (sale_id, product_id, name, qty, price, subtotal, discount, sat_clave_prod_serv, sat_descripcion, synced)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, 0)
                            """
                            try:
                                self.core.db.execute_write(item_sql, (
                                    sale_id, product_id, nombre, cantidad, precio, line_subtotal, descuento, sat_clave_prod, sat_descripcion
                                ))
                                items_imported += 1
                            except Exception as item_error:
                                self.log_text.append(f"   ⚠️ Error insertando producto '{nombre}': {item_error}")
                                # #region agent log
                                if agent_log_enabled():
                                    import json, time
                                    try:
                                        from app.utils.path_utils import get_debug_log_path_str
                                        with open(get_debug_log_path_str(), "a") as f:
                                            f.write(json.dumps({"sessionId":"debug-session","runId":"run1","hypothesisId":"IMPORT_WIZARD","location":"import_wizard.py:import_sales_detailed","message":"Error inserting item","data":{"sale_id":sale_id,"product_name":nombre,"error":str(item_error)},"timestamp":int(time.time()*1000)})+"\n")
                                    except Exception as e: logger.debug("Writing item insert error log: %s", e)
                                # #endregion
                            
                        except Exception as item_error:
                            self.log_text.append(f"   ⚠️ Error en producto: {item_error}")
                    
                    imported += 1
                    # Solo mostrar cada venta si son pocas
                    if len(sales_grouped) < 50:
                        self.log_text.append(f"✅ {folio}: {len(product_rows)} productos - ${total:.2f}")

                except Exception as e:
                    errors_db += 1
                    last_error_msg = str(e)
                    self.log_text.append(f"❌ {folio}: {e}")
                    import traceback
                    tb = traceback.format_exc()
                    logger.error(f"Error importando venta {folio}: {tb}")
                    # FIX 2026-01-30: Mostrar más detalle del error
                    if errors_db <= 5:  # Solo mostrar primeros 5 errores detallados
                        self.log_text.append(f"   Detalle: {tb[:300]}")

                self.progress.setValue(idx + 1)
                self.update_counter(idx + 1, len(sales_grouped))

                # FIX 2026-01-30: Mostrar progreso cada 50 registros (más frecuente para diagnóstico)
                if (idx + 1) % 50 == 0:
                    progress_msg = f"Progreso: {idx+1}/{len(sales_grouped)} | OK: {imported} | Err: {errors_db} | Items: {items_imported}"
                    debug_print(progress_msg)
                    logger.info("📈 %s", progress_msg)  # FIX 2026-02-01: Log a terminal
                    self.log_text.append(f"📈 {progress_msg}")
                    self.update_status(f"🔄 IMPORTANDO... {idx+1}/{len(sales_grouped)}", "#FF9800", "#FFF3E0")

                QtWidgets.QApplication.processEvents()

            # ══════════════════════════════════════════════════════════════════
            # RESUMEN FINAL DETALLADO
            # ══════════════════════════════════════════════════════════════════
            debug_print(f"=== RESUMEN: {imported} ventas importadas, {items_imported} productos, {errors_db} errores, {skipped_duplicates} duplicados ===")
            logger.info("✅ IMPORTACIÓN COMPLETADA: %d ventas, %d items, %d errores, %d duplicados", imported, items_imported, errors_db, skipped_duplicates)

            # Actualizar estado final
            if errors_db == 0:
                self.update_status(f"✅ COMPLETADO: {imported} ventas importadas", "#4CAF50", "#E8F5E9")
            else:
                self.update_status(f"⚠️ COMPLETADO con {errors_db} errores", "#FF9800", "#FFF3E0")
            self.update_counter(len(sales_grouped), len(sales_grouped))

            self.log_text.append(f"\n{'═'*60}")
            self.log_text.append("🎉 IMPORTACIÓN COMPLETADA - RESUMEN DETALLADO")
            self.log_text.append(f"{'═'*60}")
            self.log_text.append(f"   📁 Filas leídas del archivo: {len(rows)}")
            self.log_text.append(f"   📦 Ventas únicas (por FOLIO): {len(sales_grouped)}")
            self.log_text.append(f"   ✅ Ventas importadas: {imported}")
            # FIX 2026-01-30: Reportar duplicados saltados
            if skipped_duplicates > 0:
                self.log_text.append(f"   🔄 Duplicados saltados: {skipped_duplicates}")
            self.log_text.append(f"   ❌ Errores de BD: {errors_db}")
            self.log_text.append(f"   ⚠️ Ventas sin ID: {errors_no_id}")
            self.log_text.append(f"   📦 Productos importados: {items_imported}")
            # FIX 2026-01-30: Reportar productos sin nombre saltados
            skipped_items = getattr(self, '_skipped_items', 0)
            if skipped_items > 0:
                self.log_text.append(f"   ⚠️ Productos sin nombre saltados: {skipped_items}")
            self._skipped_items = 0  # Reset para próxima importación

            # FIX 2026-01-30: Mostrar último error si hubo errores
            if errors_db > 0 and last_error_msg:
                self.log_text.append(f"\n   📝 Último error: {last_error_msg[:150]}")

            # FIX 2026-01-30: Diagnóstico de por qué podrían faltar registros
            rows_sin_folio = getattr(self, '_rows_without_folio', 0)
            if len(sales_grouped) < len(rows) / 10:  # Si hay muchas menos ventas que filas
                self.log_text.append(f"\n{'─'*60}")
                self.log_text.append("ℹ️ EXPLICACIÓN: Formato DETALLADO agrupa por FOLIO")
                self.log_text.append(f"   Si tu archivo tiene {len(rows)} filas pero {len(sales_grouped)} ventas,")
                self.log_text.append(f"   significa que cada venta tiene ~{len(rows)/max(len(sales_grouped),1):.1f} productos en promedio.")
                self.log_text.append(f"   ¡Esto es CORRECTO si tu formato es una línea por producto!")

            self.log_text.append(f"\n⚠️ Las ventas tienen status='{status}'")
            self.log_text.append("   y NO afectaron el inventario.")
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            self.log_text.append(f"\n❌ Error: {e}")
    
    def import_anonymous_wallet(self, file_path):
        """
        Importa monedero anónimo desde CSV o Excel.
        
        Columnas esperadas:
        - wallet_id: ID único del monedero (requerido)
        - phone: Teléfono (opcional)
        - nickname: Apodo (opcional)
        - points_balance: Saldo actual de puntos
        - total_earned: Total ganado
        - total_redeemed: Total canjeado
        - visit_count: Número de visitas
        - last_visit: Última visita
        - status: active, suspended, blocked
        """
        self.log_text.append("💳 Importando monedero anónimo...\n")
        
        try:
            rows = []
            if file_path.lower().endswith(('.xlsx', '.xls')):
                if not HAS_OPENPYXL:
                    self.log_text.append("❌ Error: Se requiere openpyxl para importar Excel.\n")
                    return
                
                import zipfile
                import os
                
                if not os.path.exists(file_path) or os.path.getsize(file_path) == 0:
                    self.log_text.append("❌ Error: Archivo no válido")
                    return
                
                if file_path.lower().endswith('.xlsx'):
                    try:
                        with zipfile.ZipFile(file_path, 'r') as zip_ref:
                            zip_files = zip_ref.namelist()
                            if not any('xl/workbook.xml' in f or '[Content_Types].xml' in f for f in zip_files):
                                self.log_text.append("❌ Error: Archivo no parece ser un Excel válido")
                                return
                    except zipfile.BadZipFile:
                        self.log_text.append("❌ Error: Archivo Excel corrupto")
                        return
                
                wb = None
                try:
                    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                    ws = wb.active

                    headers = []
                    # FIX CRÍTICO 2026-01-30: ws[1] NO funciona en read_only mode
                    # Usar iter_rows en su lugar
                    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                    if first_row:
                        headers = [str(cell) if cell is not None else "" for cell in first_row]
                    else:
                        self.log_text.append("❌ Error: No se encontró fila de encabezados")
                        return

                    for row in ws.iter_rows(min_row=2, values_only=True):
                        row_dict = {}
                        for i, cell_value in enumerate(row):
                            if i < len(headers):
                                row_dict[headers[i]] = str(cell_value) if cell_value is not None else ""
                        if any(row_dict.values()):
                            rows.append(row_dict)

                    self.log_text.append(f"✅ Archivo Excel leído: {len(rows)} filas\n")
                except Exception as e:
                    import traceback
                    self.log_text.append(f"❌ Error leyendo Excel: {e}")
                    self.log_text.append(f"   Detalle: {traceback.format_exc()[:500]}")
                    logger.error("Error leyendo Excel: %s", traceback.format_exc())
                    self.update_status("❌ ERROR LEYENDO EXCEL", "#f44336", "#FFEBEE")
                    self.update_counter(0, 0)
                    return
                finally:
                    if wb is not None:
                        wb.close()
            else:
                # CSV
                with open(file_path, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    rows = list(reader)
            
            if not rows:
                self.log_text.append("❌ Error: Archivo vacío o sin datos")
                return

            self.progress.setMaximum(len(rows))
            self.update_counter(0, len(rows))  # FIX 2026-02-01: Inicializar contador
            imported = 0
            skipped = 0
            errors = 0

            for i, row in enumerate(rows):
                try:
                    # Obtener wallet_id (requerido)
                    wallet_id = row.get('wallet_id', row.get('Wallet ID', row.get('wallet_id', ''))).strip()
                    if not wallet_id:
                        self.log_text.append(f"⚠️ Fila {i+1}: Sin wallet_id, saltando...")
                        skipped += 1
                        continue
                    
                    # Verificar si ya existe
                    check_sql = "SELECT id FROM anonymous_wallet WHERE wallet_id = %s"
                    existing = self.core.db.execute_query(check_sql, (wallet_id,))
                    
                    if existing:
                        self.log_text.append(f"⚠️ Wallet {wallet_id} ya existe, saltando...")
                        skipped += 1
                        continue
                    
                    # Preparar datos
                    wallet_data = {
                        'wallet_id': wallet_id,
                        'wallet_hash': row.get('wallet_hash', row.get('hash', wallet_id)),
                        'phone': row.get('phone', row.get('telefono', row.get('Phone', ''))),
                        'nickname': row.get('nickname', row.get('apodo', row.get('Nickname', ''))),
                        'points_balance': int(float(row.get('points_balance', row.get('saldo', row.get('Saldo', 0))))) or 0,
                        'balance': float(row.get('balance', row.get('balance', 0))) or 0.0,
                        'total_earned': int(float(row.get('total_earned', row.get('total_ganado', row.get('Total Earned', 0))))) or 0,
                        'total_redeemed': int(float(row.get('total_redeemed', row.get('total_canjeado', row.get('Total Redeemed', 0))))) or 0,
                        'total_spent': float(row.get('total_spent', row.get('total_gastado', row.get('Total Spent', 0)))) or 0.0,
                        'last_visit': row.get('last_visit', row.get('ultima_visita', row.get('Last Visit', ''))),
                        'last_activity': row.get('last_activity', row.get('ultima_actividad', '')),
                        'visit_count': int(float(row.get('visit_count', row.get('visitas', row.get('Visit Count', 0))))) or 0,
                        'status': row.get('status', row.get('estado', row.get('Status', 'active')))
                    }
                    
                    # Insertar wallet
                    insert_sql = """
                        INSERT INTO anonymous_wallet (
                            wallet_id, wallet_hash, phone, nickname,
                            points_balance, balance, total_earned, total_redeemed,
                            total_spent, last_visit, last_activity, visit_count, status, synced
                        )
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 0)
                    """

                    self.core.db.execute_write(insert_sql, (
                        wallet_data['wallet_id'],
                        wallet_data['wallet_hash'],
                        wallet_data['phone'] if wallet_data['phone'] else None,
                        wallet_data['nickname'] if wallet_data['nickname'] else None,
                        wallet_data['points_balance'],
                        wallet_data['balance'],
                        wallet_data['total_earned'],
                        wallet_data['total_redeemed'],
                        wallet_data['total_spent'],
                        wallet_data['last_visit'] if wallet_data['last_visit'] else None,
                        wallet_data['last_activity'] if wallet_data['last_activity'] else None,
                        wallet_data['visit_count'],
                        wallet_data['status']
                    ))
                    
                    imported += 1
                    self.log_text.append(f"✅ Wallet {wallet_id}: {wallet_data['points_balance']} puntos")
                
                except Exception as e:
                    errors += 1
                    self.log_text.append(f"❌ Fila {i+1}: {e}")

                self.progress.setValue(i + 1)
                self.update_counter(i + 1, len(rows))  # FIX 2026-02-01: Actualizar contador
                QtWidgets.QApplication.processEvents()

            self.log_text.append(f"\n🎉 Importados: {imported}/{len(rows)} monederos")
            if skipped > 0:
                self.log_text.append(f"⚠️ Saltados: {skipped} (ya existían)")
            if errors > 0:
                self.log_text.append(f"❌ Errores: {errors}")
        
        except Exception as e:
            self.log_text.append(f"\n❌ Error: {e}")