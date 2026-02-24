"""
Módulo de Exportación Inteligente - Reportes de Ventas con Auditoría
Separación Serie A (Fiscal) / Serie B (Sombra) con fórmulas dinámicas
"""

from typing import Any, Dict, List, Optional, Tuple
from datetime import date, datetime
from decimal import Decimal
import logging
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
    from openpyxl.styles import Alignment, Border, Fill, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

logger = logging.getLogger(__name__)

class IntelligentExporter:
    """
    Exportador inteligente con:
    - Separación automática Serie A (Fiscal) / Serie B (Sombra)
    - Fórmulas dinámicas de IVA
    - Filtros de auditoría con estados coloreados
    - Gráficos de resumen
    """
    
    # Estilos
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    
    SERIE_A_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    SERIE_B_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    
    CANCELLED_FILL = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    CANCELLED_FONT = Font(color="FFFFFF", bold=True, strikethrough=True)
    
    TAX_RATE = Decimal("0.16")  # 16% IVA México
    TAX_FACTOR = Decimal("1.16")  # 1 + TAX_RATE for dividing out IVA
    
    def __init__(self, db_manager):
        """
        Inicializa el exportador.
        
        Args:
            db_manager: DatabaseManager instance para acceso a la base de datos
        """
        self.db = db_manager
        # Use dynamic workspace root for export directory
        try:
            from app.utils.path_utils import get_workspace_root
            workspace_root = get_workspace_root()
            self.export_dir = workspace_root / "data" / "exports"
        except Exception:
            # Fallback to relative path if path_utils is not available
            self.export_dir = Path("data/exports")
        self.export_dir.mkdir(parents=True, exist_ok=True)
    
    def export_sales_report(
        self,
        start_date: date,
        end_date: date,
        include_cancelled: bool = True,
        separate_series: bool = True,
        with_charts: bool = True
    ) -> str:
        """
        Exporta reporte de ventas con inteligencia fiscal.
        
        Args:
            start_date: Fecha inicial
            end_date: Fecha final
            include_cancelled: Incluir ventas canceladas
            separate_series: Separar Serie A y B en hojas diferentes
            with_charts: Incluir gráficos de resumen
        
        Returns:
            Ruta del archivo Excel generado
        """
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl requerido: pip install openpyxl")
        
        wb = Workbook()
        
        # Obtener datos
        sales = self._fetch_sales(start_date, end_date, include_cancelled)
        
        if separate_series:
            # Separar por serie
            serie_a = [s for s in sales if s.get('serie', 'A') == 'A']
            serie_b = [s for s in sales if s.get('serie', 'A') == 'B']
            
            # Hoja Serie A (Fiscal)
            ws_a = wb.active
            ws_a.title = "📗 Serie A - Fiscal"
            self._create_sales_sheet(ws_a, serie_a, "SERIE A - VENTAS FISCALES")
            
            # Hoja Serie B (Sombra)
            ws_b = wb.create_sheet("📙 Serie B - Sombra")
            self._create_sales_sheet(ws_b, serie_b, "SERIE B - VENTAS SOMBRA")
            
            # Hoja Resumen Global
            ws_summary = wb.create_sheet("📊 Resumen Global")
            self._create_summary_sheet(ws_summary, serie_a, serie_b, start_date, end_date)
        else:
            # Una sola hoja con todas las ventas
            ws = wb.active
            ws.title = "Ventas"
            self._create_sales_sheet(ws, sales, f"REPORTE DE VENTAS {start_date} - {end_date}")
        
        # Hoja de auditoría (ventas canceladas)
        if include_cancelled:
            cancelled = [s for s in sales if s.get('status') == 'cancelled']
            if cancelled:
                ws_audit = wb.create_sheet("🔴 Auditoría")
                self._create_audit_sheet(ws_audit, cancelled)
        
        # Guardar archivo
        filename = f"Ventas_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
        filepath = self.export_dir / filename
        wb.save(filepath)
        
        logger.info(f"✅ Reporte exportado: {filepath}")
        return str(filepath)
    
    def _fetch_sales(self, start_date: date, end_date: date, include_cancelled: bool) -> List[Dict]:
        """Obtiene ventas de la base de datos."""
        status_filter = "" if include_cancelled else "AND (s.status IS NULL OR s.status != 'cancelled')"
        
        query = f"""
            SELECT 
                s.id,
                s.folio,
                COALESCE(s.serie, 'A') as serie,
                s.timestamp as fecha,
                s.total,
                s.subtotal,
                s.tax as iva,
                s.discount as descuento,
                s.payment_method as metodo_pago,
                s.status,
                c.name as cliente,
                u.username as cajero,
                s.turn_id as turno
            FROM sales s
            LEFT JOIN customers c ON s.customer_id = c.id
            LEFT JOIN users u ON s.user_id = u.id
            WHERE s.timestamp::date >= %s AND s.timestamp::date <= %s
            {status_filter}
            ORDER BY s.timestamp DESC
        """
        
        # Use DatabaseManager.execute_query() instead of direct cursor
        result = self.db.execute_query(query, (start_date.isoformat(), end_date.isoformat()))
        
        if not result:
            return []
        
        # Convert Row objects to dicts
        sales = []
        for row in result:
            sale = dict(row)
            sales.append(sale)
        
        return sales
    
    def _create_sales_sheet(self, ws, sales: List[Dict], title: str):
        """Crea una hoja de ventas con formato profesional."""
        
        # Título
        ws.merge_cells('A1:L1')
        ws['A1'] = title
        ws['A1'].font = Font(size=16, bold=True, color="1F4E79")
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Fecha de generación
        ws['A2'] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A2'].font = Font(italic=True, color="666666")
        
        # Headers
        headers = ['Folio', 'Serie', 'Fecha', 'Total', 'Subtotal', 'IVA (16%)', 
                   'Descuento', 'Método Pago', 'Cliente', 'Cajero', 'Turno', 'Estado']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
        
        # Datos con fórmulas dinámicas
        for row_idx, sale in enumerate(sales, 5):
            ws.cell(row=row_idx, column=1, value=sale.get('folio', sale.get('id')))
            ws.cell(row=row_idx, column=2, value=sale.get('serie', 'A'))
            ws.cell(row=row_idx, column=3, value=sale.get('fecha'))
            
            # Total
            total_cell = ws.cell(row=row_idx, column=4, value=float(sale.get('total', 0)))
            total_cell.number_format = '$#,##0.00'
            
            # Subtotal - Calcular si no existe (using Decimal for precision)
            subtotal = sale.get('subtotal')
            if subtotal is None:
                total_dec = Decimal(str(sale.get('total', 0)))
                subtotal = float((total_dec / self.TAX_FACTOR).quantize(Decimal('0.01')))
            
            subtotal_cell = ws.cell(row=row_idx, column=5, value=float(subtotal))
            subtotal_cell.number_format = '$#,##0.00'
            
            # IVA - FÓRMULA DINÁMICA
            iva_cell = ws.cell(row=row_idx, column=6)
            iva_cell.value = f"=E{row_idx}*0.16"  # Subtotal * 16%
            iva_cell.number_format = '$#,##0.00'
            
            # Descuento
            desc_cell = ws.cell(row=row_idx, column=7, value=float(sale.get('descuento', 0)))
            desc_cell.number_format = '$#,##0.00'
            
            ws.cell(row=row_idx, column=8, value=sale.get('metodo_pago', 'efectivo'))
            ws.cell(row=row_idx, column=9, value=sale.get('cliente', 'Público General'))
            ws.cell(row=row_idx, column=10, value=sale.get('cajero', '-'))
            ws.cell(row=row_idx, column=11, value=sale.get('turno', '-'))
            
            # Estado con color
            status = sale.get('status', 'completed')
            status_cell = ws.cell(row=row_idx, column=12, value=status or 'completada')
            
            if status == 'cancelled':
                status_cell.fill = self.CANCELLED_FILL
                status_cell.font = self.CANCELLED_FONT
                # Aplicar tachado a toda la fila
                for col in range(1, 13):
                    ws.cell(row=row_idx, column=col).font = Font(strikethrough=True)
        
        # Totales al final
        if sales:
            total_row = len(sales) + 5
            ws.cell(row=total_row, column=3, value="TOTALES:").font = Font(bold=True)
            
            # Suma de totales
            ws.cell(row=total_row, column=4).value = f"=SUM(D5:D{total_row-1})"
            ws.cell(row=total_row, column=4).number_format = '$#,##0.00'
            ws.cell(row=total_row, column=4).font = Font(bold=True)
            
            # Suma de subtotales
            ws.cell(row=total_row, column=5).value = f"=SUM(E5:E{total_row-1})"
            ws.cell(row=total_row, column=5).number_format = '$#,##0.00'
            ws.cell(row=total_row, column=5).font = Font(bold=True)
            
            # Suma de IVA
            ws.cell(row=total_row, column=6).value = f"=SUM(F5:F{total_row-1})"
            ws.cell(row=total_row, column=6).number_format = '$#,##0.00'
            ws.cell(row=total_row, column=6).font = Font(bold=True)
        
        # Ajustar anchos de columna
        column_widths = [10, 8, 20, 15, 15, 15, 12, 15, 25, 15, 10, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def _create_summary_sheet(self, ws, serie_a: List[Dict], serie_b: List[Dict], 
                               start_date: date, end_date: date):
        """Crea hoja de resumen con comparativa Serie A vs B."""
        
        ws.merge_cells('A1:F1')
        ws['A1'] = f"📊 RESUMEN GLOBAL: {start_date} - {end_date}"
        ws['A1'].font = Font(size=18, bold=True)
        
        # Calcular totales
        total_a = sum(float(s.get('total', 0)) for s in serie_a if s.get('status') != 'cancelled')
        total_b = sum(float(s.get('total', 0)) for s in serie_b if s.get('status') != 'cancelled')
        
        cancelled_a = len([s for s in serie_a if s.get('status') == 'cancelled'])
        cancelled_b = len([s for s in serie_b if s.get('status') == 'cancelled'])
        
        # Tabla de resumen (using Decimal for IVA calculations)
        total_a_dec = Decimal(str(total_a))
        total_b_dec = Decimal(str(total_b))
        total_combined_dec = total_a_dec + total_b_dec

        subtotal_a = float((total_a_dec / self.TAX_FACTOR).quantize(Decimal('0.01')))
        subtotal_b = float((total_b_dec / self.TAX_FACTOR).quantize(Decimal('0.01')))
        subtotal_combined = float((total_combined_dec / self.TAX_FACTOR).quantize(Decimal('0.01')))

        iva_a = float((total_a_dec - (total_a_dec / self.TAX_FACTOR)).quantize(Decimal('0.01')))
        iva_b = float((total_b_dec - (total_b_dec / self.TAX_FACTOR)).quantize(Decimal('0.01')))
        iva_combined = float((total_combined_dec - (total_combined_dec / self.TAX_FACTOR)).quantize(Decimal('0.01')))

        data = [
            ['', 'Serie A (Fiscal)', 'Serie B (Sombra)', 'TOTAL'],
            ['Núm. Ventas', len(serie_a), len(serie_b), len(serie_a) + len(serie_b)],
            ['Total Ventas', total_a, total_b, total_a + total_b],
            ['Subtotal', subtotal_a, subtotal_b, subtotal_combined],
            ['IVA (16%)', iva_a, iva_b, iva_combined],
            ['Canceladas', cancelled_a, cancelled_b, cancelled_a + cancelled_b],
        ]
        
        for row_idx, row_data in enumerate(data, 3):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                if row_idx == 3:
                    cell.fill = self.HEADER_FILL
                    cell.font = self.HEADER_FONT
                elif col_idx == 1:
                    cell.font = Font(bold=True)
                elif isinstance(value, (int, float)) and col_idx > 1 and row_idx > 4:
                    cell.number_format = '$#,##0.00'
        
        # Colorear columnas por serie
        for row in range(4, 9):
            ws.cell(row=row, column=2).fill = self.SERIE_A_FILL
            ws.cell(row=row, column=3).fill = self.SERIE_B_FILL
        
        # Estadísticas adicionales
        ws['A11'] = "📈 ANÁLISIS"
        ws['A11'].font = Font(size=14, bold=True)
        
        total_global = total_a + total_b
        if total_global > 0:
            ws['A12'] = f"• Serie A representa {(total_a/total_global)*100:.1f}% de las ventas"
            ws['A13'] = f"• Serie B representa {(total_b/total_global)*100:.1f}% de las ventas"
        
        if cancelled_a + cancelled_b > 0:
            ws['A15'] = "⚠️ ALERTA DE AUDITORÍA"
            ws['A15'].font = Font(size=12, bold=True, color="FF0000")
            ws['A16'] = f"Se detectaron {cancelled_a + cancelled_b} ventas canceladas que requieren revisión"
        
        # Ajustar anchos
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def _create_audit_sheet(self, ws, cancelled_sales: List[Dict]):
        """Crea hoja de auditoría para ventas canceladas (detección de fraude)."""
        
        ws['A1'] = "🔴 AUDITORÍA - VENTAS CANCELADAS"
        ws['A1'].font = Font(size=16, bold=True, color="FF0000")
        
        ws['A2'] = "⚠️ Revisar estas transacciones para detectar posibles irregularidades"
        ws['A2'].font = Font(italic=True, color="FF0000")
        
        headers = ['Folio', 'Serie', 'Fecha', 'Total', 'Cajero', 'Cliente', '⚠️ REQUIERE REVISIÓN']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.fill = self.CANCELLED_FILL
            cell.font = Font(color="FFFFFF", bold=True)
        
        for row_idx, sale in enumerate(cancelled_sales, 5):
            ws.cell(row=row_idx, column=1, value=sale.get('folio', sale.get('id')))
            ws.cell(row=row_idx, column=2, value=sale.get('serie', 'A'))
            ws.cell(row=row_idx, column=3, value=sale.get('fecha'))
            
            total_cell = ws.cell(row=row_idx, column=4, value=float(sale.get('total', 0)))
            total_cell.number_format = '$#,##0.00'
            
            ws.cell(row=row_idx, column=5, value=sale.get('cajero', '-'))
            ws.cell(row=row_idx, column=6, value=sale.get('cliente', '-'))
            ws.cell(row=row_idx, column=7, value="🔍 PENDIENTE")
            
            # Resaltar filas sospechosas (montos altos)
            if float(sale.get('total', 0)) > 1000:
                for col in range(1, 8):
                    ws.cell(row=row_idx, column=col).fill = PatternFill(
                        start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
                    )
        
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 18
